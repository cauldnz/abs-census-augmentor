"""Real-Data-First probe for the new-datasets work (SA3 boundary, AIHW Mental
Health Prescriptions, ABS Building Approvals).

Run with::

    uv run python tools/probe_new_datasets.py

Downloads one representative slice of each upstream and dumps the schema
(columns, sheet names, code formats, file sizes). Outputs to stdout so the
findings can be pasted directly into the spec PR.

This is a discovery probe, NOT part of the test suite. Once the production
fetchers register, the equivalent real-data smoke moves into
``tools/verify_real_parsers.py``.

See ``CLAUDE.md`` "Real Data First".
"""

from __future__ import annotations

import io
import sys
import zipfile
from pathlib import Path

import requests

# --------------------------------------------------------------------------
# URLs probed live 2026-06-01:
# - SA3 boundary: confirmed 200 + application/zip via HEAD
# - LGA 2025 boundary: confirmed 200 + application/zip via HEAD
# - ABS BA Mar 2026 NSW SA2: confirmed 200 + xlsx via HEAD
# - AIHW MH Prescriptions: found via the NMHSPF "Regional activity data" page

PROBE_DIR = Path(__file__).resolve().parent.parent / "data" / "_probe"
PROBE_DIR.mkdir(parents=True, exist_ok=True)

USER_AGENT = "abs-census-augmentor-probe/0.1 (chris.auld@auld.nz)"

SA3_URL = (
    "https://www.abs.gov.au/statistics/standards/"
    "australian-statistical-geography-standard-asgs/"
    "edition-3-july-2021-june-2026/access-and-downloads/"
    "digital-boundary-files/SA3_2021_AUST_SHP_GDA2020.zip"
)
LGA_URL = (
    "https://www.abs.gov.au/statistics/standards/"
    "australian-statistical-geography-standard-asgs/"
    "edition-3-july-2021-june-2026/access-and-downloads/"
    "digital-boundary-files/LGA_2025_AUST_GDA2020.zip"
)
BA_NSW_SA2_URL = (
    "https://www.abs.gov.au/statistics/industry/building-and-construction/"
    "building-approvals-australia/mar-2026/87310do002_202603.xlsx"
)
AIHW_RX_ZIP_URL = (
    "https://www.aihw.gov.au/getmedia/"
    "464b35c8-9573-4a02-a508-0757c66feeb4/"
    "Mental-health-related-prescriptions-2024-25.zip"
)


def _fetch(url: str, cache_name: str) -> bytes:
    cache_path = PROBE_DIR / cache_name
    if cache_path.exists() and cache_path.stat().st_size > 1024:
        print(f"  [cached] {cache_path.name} ({cache_path.stat().st_size:,} bytes)")
        return cache_path.read_bytes()
    print(f"  fetching {url} ...")
    r = requests.get(url, headers={"User-Agent": USER_AGENT}, timeout=120, stream=True)
    r.raise_for_status()
    blob = r.content
    cache_path.write_bytes(blob)
    print(f"  wrote {cache_path.name} ({len(blob):,} bytes)")
    return blob


def probe_sa3_boundary() -> None:
    print("=" * 70)
    print("(1) ABS SA3 boundary — ASGS Edition 3 / GDA2020 / 2021")
    print("=" * 70)
    blob = _fetch(SA3_URL, "sa3-edition3.zip")
    with zipfile.ZipFile(io.BytesIO(blob)) as zf:
        names = zf.namelist()
        print(f"  ZIP contents ({len(names)} files):")
        for n in names:
            info = zf.getinfo(n)
            print(f"    {n}  ({info.file_size:,} bytes)")

    # Read the .dbf attribute table without geometry — we only need columns.
    import geopandas as gpd  # noqa: PLC0415

    zip_path = PROBE_DIR / "sa3-edition3.zip"
    gdf = gpd.read_file(f"zip://{zip_path}")
    print(f"  loaded GeoDataFrame: {len(gdf)} rows, {len(gdf.columns)} cols")
    print(f"  columns: {list(gdf.columns)}")
    print(f"  dtypes:\n{gdf.dtypes}")
    print("  first 2 rows (geometry omitted):")
    print(gdf.drop(columns=["geometry"]).head(2).to_string())
    print(f"  CRS: {gdf.crs}")
    print()


def probe_lga_boundary() -> None:
    print("=" * 70)
    print("(2) ABS LGA 2025 boundary — ASGS Edition 3 / GDA2020")
    print("=" * 70)
    blob = _fetch(LGA_URL, "lga-2025.zip")
    with zipfile.ZipFile(io.BytesIO(blob)) as zf:
        names = zf.namelist()
        print(f"  ZIP contents ({len(names)} files):")
        for n in names:
            info = zf.getinfo(n)
            print(f"    {n}  ({info.file_size:,} bytes)")

    import geopandas as gpd  # noqa: PLC0415

    zip_path = PROBE_DIR / "lga-2025.zip"
    gdf = gpd.read_file(f"zip://{zip_path}")
    print(f"  loaded GeoDataFrame: {len(gdf)} rows, {len(gdf.columns)} cols")
    print(f"  columns: {list(gdf.columns)}")
    print(f"  dtypes:\n{gdf.dtypes}")
    print("  first 2 rows (geometry omitted):")
    print(gdf.drop(columns=["geometry"]).head(2).to_string())
    print(f"  CRS: {gdf.crs}")
    print()


def probe_abs_ba_nsw_sa2() -> None:
    print("=" * 70)
    print("(3) ABS Building Approvals — March 2026 NSW SA2 cube (87310do002)")
    print("=" * 70)
    blob = _fetch(BA_NSW_SA2_URL, "abs-ba-mar2026-nsw-sa2.xlsx")
    import openpyxl  # noqa: PLC0415

    wb = openpyxl.load_workbook(io.BytesIO(blob), read_only=True, data_only=True)
    print(f"  Sheets ({len(wb.sheetnames)}): {wb.sheetnames}")
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"  --- sheet: {sheet_name} ---")
        rows = list(ws.iter_rows(values_only=True))
        print(f"    rows: {len(rows)}")
        # First ~15 rows so we see preamble + header + first data rows.
        for i, row in enumerate(rows[:15]):
            cells = [str(c)[:50] if c is not None else "" for c in row[:12]]
            print(f"    row {i}: {cells}")
    print()


def probe_aihw_rx() -> None:
    print("=" * 70)
    print("(4) AIHW Mental Health Prescriptions 2024-25 ZIP (NMHSPF)")
    print("=" * 70)
    blob = _fetch(AIHW_RX_ZIP_URL, "aihw-mh-rx-2024-25.zip")
    with zipfile.ZipFile(io.BytesIO(blob)) as zf:
        names = zf.namelist()
        print(f"  ZIP contents ({len(names)} files):")
        for n in names:
            info = zf.getinfo(n)
            print(f"    {n}  ({info.file_size:,} bytes)")

        # Look at the first .xlsx file's sheet structure
        xlsx_files = [n for n in names if n.lower().endswith(".xlsx")]
        print(f"  xlsx files found: {len(xlsx_files)}")
        if not xlsx_files:
            print("  (no xlsx in ZIP — checking for csv/other)")
            csv_files = [n for n in names if n.lower().endswith(".csv")]
            print(f"  csv files found: {len(csv_files)}")
            for cn in csv_files[:3]:
                with zf.open(cn) as f:
                    head = f.read(1500).decode("utf-8", errors="replace")
                print(f"  --- csv: {cn} (first 1500 bytes) ---")
                print(head)
                print()
            return

        # For up to first 3 xlsx files, dump sheet names + first few rows
        import openpyxl  # noqa: PLC0415

        for xname in xlsx_files[:3]:
            with zf.open(xname) as f:
                xbytes = f.read()
            wb = openpyxl.load_workbook(io.BytesIO(xbytes), read_only=True, data_only=True)
            print(f"  --- xlsx: {xname} ---")
            print(f"    sheets: {wb.sheetnames}")
            # Spot the SA3 sheet if present
            for sn in wb.sheetnames:
                lower = sn.lower()
                if "sa3" in lower or "sa 3" in lower or "statistical area" in lower:
                    ws = wb[sn]
                    rows = list(ws.iter_rows(values_only=True))
                    print(f"    --- SA3-ish sheet: {sn} ({len(rows)} rows) ---")
                    for i, row in enumerate(rows[:12]):
                        cells = [str(c)[:50] if c is not None else "" for c in row[:10]]
                        print(f"      row {i}: {cells}")
            print()


def main() -> int:
    probes = [
        ("SA3 boundary", probe_sa3_boundary),
        ("LGA 2025 boundary", probe_lga_boundary),
        ("ABS Building Approvals SA2 NSW Mar 2026", probe_abs_ba_nsw_sa2),
        ("AIHW Mental Health Prescriptions", probe_aihw_rx),
    ]
    failures = 0
    for label, fn in probes:
        try:
            fn()
        except Exception as e:
            failures += 1
            print(f"  *** {label} FAILED: {e!r}")
            print()
    print(f"Done. {len(probes) - failures}/{len(probes)} probes succeeded.")
    return 0 if failures == 0 else 1


if __name__ == "__main__":
    sys.exit(main())
