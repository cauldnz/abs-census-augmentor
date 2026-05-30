"""DSS Payment Demographic Data fetcher (spec §20, dataset id ``dss_payments``).

DSS publishes quarterly XLSX workbooks via the data.gov.au CKAN API.
We:

1. Hit the CKAN ``package_show`` endpoint to enumerate releases.
2. Pick the requested release (default: latest by last_modified).
3. Download the XLSX (~1 MB).
4. Parse the ``SA2`` sheet into a SA2-keyed DataFrame, snake-case-ifying
   the payment-type column names.

Suppressed cells (per ABS perturbation rules) come through with whatever
the source publishes — ``5`` is the suppression floor in modern DSS
publications. Downstream callers should treat ``5`` as potentially
suppressed rather than as a literal count of 5 recipients.
"""

from __future__ import annotations

import logging
import re
from pathlib import Path

import pandas as pd
import requests

from ._xlsx_base import _AbsXlsxDataset

_log = logging.getLogger(__name__)

CKAN_PACKAGE_URL = (
    "https://data.gov.au/data/api/3/action/package_show?id=dss-payment-demographic-data"
)


# Match a release period in a resource name like "September 2025" or
# "December 2024". The CKAN package's resource names are the only
# stable signal we have for picking a specific quarter.
_PERIOD_RE = re.compile(
    r"\b(?P<month>January|February|March|April|May|June|July|"
    r"August|September|October|November|December)"
    r"\s+(?P<year>20\d{2})\b",
    re.IGNORECASE,
)

# Map "September" → quarter ID "Q3" (DSS publishes month-end snapshots
# but the underlying data is quarterly).
_MONTH_TO_QUARTER: dict[str, str] = {
    "march": "Q1",
    "june": "Q2",
    "september": "Q3",
    "december": "Q4",
}


class DssDataSource(_AbsXlsxDataset):
    """Fetch + load the DSS Payment Demographic Data SA2 file.

    Implements the :class:`DatasetFetcher` Protocol via the shared
    :class:`_AbsXlsxDataset` base.

    The ``release`` parameter accepts ``"latest"`` (default), or a
    ``"YYYY-Qn"`` string like ``"2025-Q4"``. Resolution picks the
    matching CKAN resource by name; if no match, raises with the
    available release list.
    """

    _label = "DSS payment demographics"
    _cache_glob = "dss-*.xlsx"

    def __init__(
        self,
        *,
        release: str = "latest",
        root: Path,
        ckan_url: str = CKAN_PACKAGE_URL,
        session: requests.Session | None = None,
        chunk_size: int = 256 * 1024,
        timeout: float = 60.0,
    ) -> None:
        super().__init__(
            release=release,
            root=root,
            session=session,
            chunk_size=chunk_size,
            timeout=timeout,
        )
        self._ckan_url = ckan_url

    # ---- hooks ---------------------------------------------------------

    def _filename_stem(self, release: str) -> str:
        return f"dss-{release}"

    def _post_parse(self, df: pd.DataFrame) -> pd.DataFrame:
        # Tag rows with the release identifier so downstream consumers
        # can tell which quarter's snapshot they're looking at.
        df["release_quarter"] = self.resolved_release
        return df

    def _resolve_release(self) -> None:
        if self._resolved_release is not None:
            return

        _log.debug("Resolving DSS release via CKAN: %s", self._ckan_url)
        resp = self._session.get(self._ckan_url, timeout=self._timeout)
        resp.raise_for_status()
        payload = resp.json()
        if not payload.get("success"):
            raise RuntimeError(f"CKAN package_show returned success=false: {payload}")
        resources = payload.get("result", {}).get("resources", [])
        if not resources:
            raise RuntimeError(
                "No resources listed in the DSS CKAN package — has the dataset moved?"
            )

        # Build (release_id, last_modified, url) tuples for every
        # resource that names a quarter.
        candidates: list[tuple[str, str, str]] = []
        for res in resources:
            name = res.get("name", "")
            url = res.get("url", "")
            last_modified = res.get("last_modified", "") or ""
            release_id = _release_id_from_name(name)
            if not release_id or not url:
                continue
            # Skip non-XLSX resources just in case.
            fmt = (res.get("format") or "").lower()
            if "xlsx" not in fmt and "excel" not in fmt:
                continue
            candidates.append((release_id, last_modified, url))

        if not candidates:
            raise RuntimeError(
                "No DSS resources matched the expected name pattern "
                f"({_PERIOD_RE.pattern!r}). Available names: "
                f"{[r.get('name', '') for r in resources[:5]]}..."
            )

        # Pick the requested release (or latest by date in the id).
        requested = self._release_request
        if requested == "latest":
            # Sort lexicographically by release_id (YYYY-Qn) — that
            # matches chronological order.
            picked = max(candidates, key=lambda t: (t[0], t[1]))
        else:
            matching = [c for c in candidates if c[0] == requested]
            if not matching:
                raise RuntimeError(
                    f"DSS release {requested!r} not found. "
                    f"Available: "
                    f"{sorted({c[0] for c in candidates}, reverse=True)[:8]}"
                )
            picked = matching[0]

        self._resolved_release = picked[0]
        self._resolved_url = picked[2]
        _log.info(
            "Resolved DSS release=%s, url=%s",
            self._resolved_release,
            self._resolved_url,
        )

    # ---- parsing -------------------------------------------------------

    def _parse_xlsx(self, xlsx_path: Path) -> pd.DataFrame:
        import openpyxl  # noqa: PLC0415

        # Lazy import — keeps the 63 KB static mapping out of memory for
        # callers who never parse a pre-Q2-2023 DSS file.
        from ._dss_sa2_5digit_edition_2 import (  # noqa: PLC0415
            SA2_5DIG_TO_MAIN_EDITION_2,
        )

        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
        if "SA2" not in wb.sheetnames:
            wb.close()
            raise RuntimeError(
                f"DSS workbook at {xlsx_path} has no 'SA2' sheet. Sheets: {wb.sheetnames}"
            )

        sa2 = wb["SA2"]
        rows: list[list[object]] = []
        for r in sa2.iter_rows(values_only=True):
            rows.append(list(r))

        # Header is row 2 (0-indexed); some publications have varied
        # preamble lengths so we scan for the row containing 'SA2'
        # and 'SA2 name' in adjacent cells.
        header_idx = -1
        for i in range(min(15, len(rows))):
            row = rows[i]
            if len(row) < 2:
                continue
            cell_a = "" if row[0] is None else str(row[0]).strip().lower()
            cell_b = "" if row[1] is None else str(row[1]).strip().lower()
            if cell_a == "sa2" and "name" in cell_b:
                header_idx = i
                break

        if header_idx < 0:
            wb.close()
            raise RuntimeError(f"Could not find SA2 header row in {xlsx_path}")

        header = ["" if c is None else str(c).strip() for c in rows[header_idx]]
        # Normalise payment-type column names to snake_case + suffix.
        col_names: list[str | None] = ["sa2_code_2021", None]  # name slot
        for raw in header[2:]:
            if not raw:
                col_names.append(None)
                continue
            col_names.append(_payment_column_name(raw))

        # Issue #99: pre-Q2-2023 DSS releases use 5-digit ``SA2_5DIG16``
        # codes (e.g. ``11007`` Braidwood) rather than the 9-digit
        # ``SA2_MAIN16`` form (``101021007``) that Q2-2023+ files
        # adopted. Detect on first data row and convert via the
        # bundled static mapping. ABS Edition 2 codes are frozen so
        # the mapping never goes stale.
        unknown_5digit_codes: list[str] = []
        records: list[dict[str, object]] = []
        for row in rows[header_idx + 1 :]:
            if len(row) < 2:
                continue
            sa2_raw = row[0]
            sa2 = "" if sa2_raw is None else str(sa2_raw).strip()
            if len(sa2) == 5 and sa2.isdigit():
                # Older DSS release with 5-digit SA2 codes — convert to
                # the 9-digit Edition 2 form so the resulting index
                # joins cleanly with the cross-edition spatial lookup
                # (which produces 9-digit codes).
                converted = SA2_5DIG_TO_MAIN_EDITION_2.get(sa2)
                if converted is None:
                    unknown_5digit_codes.append(sa2)
                    continue
                sa2 = converted
            elif not (len(sa2) == 9 and sa2.isdigit()):
                continue
            rec: dict[str, object] = {"sa2_code_2021": sa2}
            for col_idx, col_name in enumerate(col_names):
                if col_name is None or col_idx == 0:
                    continue
                if col_idx >= len(row):
                    rec[col_name] = None
                    continue
                rec[col_name] = _coerce_dss_cell(row[col_idx])
            records.append(rec)

        wb.close()
        if unknown_5digit_codes:
            _log.warning(
                "DSS workbook %s contained %d 5-digit SA2 codes not in the "
                "bundled Edition 2 mapping (sample: %s); those rows were "
                "skipped. If you see this against a fresh ABS release the "
                "mapping may need regenerating from a current Edition 2 "
                "boundary download.",
                xlsx_path,
                len(unknown_5digit_codes),
                unknown_5digit_codes[:5],
            )
        if not records:
            raise RuntimeError(f"No SA2 data rows in {xlsx_path}")

        df = pd.DataFrame.from_records(records)
        return df.set_index("sa2_code_2021")


# ---- helpers ------------------------------------------------------------


def _release_id_from_name(name: str) -> str | None:
    """Extract a ``YYYY-Qn`` release id from a CKAN resource name.

    DSS resource names look like
    ``"Expanded DSS Benefit and Payment Recipient Demographics - December 2025"``.
    Returns ``"2025-Q4"`` for that example, or ``None`` if no period
    pattern matches.
    """
    match = _PERIOD_RE.search(name or "")
    if not match:
        return None
    month = match.group("month").lower()
    year = match.group("year")
    quarter = _MONTH_TO_QUARTER.get(month)
    if not quarter:
        return None
    return f"{year}-{quarter}"


def _payment_column_name(raw: str) -> str:
    """Normalise a payment-type header into snake_case + ``_recipients``.

    "Age Pension" → "age_pension_recipients"
    "ABSTUDY (Living allowance)" → "abstudy_living_allowance_recipients"
    "Carer Allowance (Child Health Care Card only)" →
        "carer_allowance_child_health_care_card_only_recipients"
    """
    s = raw.strip().lower()
    # Drop punctuation; collapse whitespace.
    s = re.sub(r"[^a-z0-9\s]+", " ", s)
    s = re.sub(r"\s+", "_", s).strip("_")
    if not s:
        return ""
    return f"{s}_recipients"


def _coerce_dss_cell(cell: object) -> object:
    """DSS counts come as int or as the string 'np' / blank for missing."""
    if cell is None:
        return None
    if isinstance(cell, (int, float)):
        return cell
    s = str(cell).strip()
    if not s or s.lower() in ("np", "na", "n/a", "-", "..", ".", "<20", "nan", "null"):
        return None
    if re.fullmatch(r"-?\d+", s):
        return int(s)
    if re.fullmatch(r"-?\d+\.\d+", s):
        return float(s)
    return s


# ---- fetcher registration ------------------------------------------------


def _build_fetcher(root: Path, release: str | None = None) -> DssDataSource:
    kwargs: dict[str, object] = {"root": root}
    if release is not None:
        kwargs["release"] = release
    return DssDataSource(**kwargs)  # type: ignore[arg-type]


def _register() -> None:
    from . import registry  # noqa: PLC0415

    registry.register_fetcher("dss_payments", _build_fetcher)


_register()
