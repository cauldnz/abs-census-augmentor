"""ABS Building Approvals (catalogue 8731.0) fetcher (spec §20).

ABS publishes building approvals monthly as **8 per-state XLSX cubes**
(one per state/territory) at SA2 level. Each release cycle ships both
the complete previous financial year and the current FYTD; the SA2
cubes have product numbers ``do002`` / ``do003`` (NSW), ``do006`` /
``do007`` (VIC), and so on through ``do030`` / ``do031`` (ACT).

The fan-out shape means this fetcher doesn't fit the
:class:`_AbsXlsxDataset` single-file base; it implements
:class:`DatasetFetcher` directly. Cache layout:

- ``<root>/abs-ba-<release>-<state>.xlsx`` — one XLSX per state, deterministic name
- ``<root>/abs-ba-<release>.parquet`` — combined SA2 rows, written after parse

``release`` is the Australian financial year (e.g. ``"2024-25"``). Each
release maps to one of two product-number series per state — the complete
FY uses the ``do002``/``do006``/... series, the FYTD uses
``do003``/``do007``/.... The current release ("latest") is the most
recent complete FY.

Real-data probe: see ``tools/probe_new_datasets.py`` for the
representative download + schema dump that this implementation mirrors.
"""

from __future__ import annotations

import logging
import re
from pathlib import Path

import pandas as pd
import requests

from .._http_retry import retry_stream_get

_log = logging.getLogger(__name__)

ABS_BA_LANDING_URL = (
    "https://www.abs.gov.au/statistics/industry/building-and-construction/"
    "building-approvals-australia/latest-release"
)

# Per-state SA2 product numbers: (state_label, complete-FY product, FYTD product).
# Confirmed live 2026-06-01 via tools/probe_new_datasets.py against the
# March 2026 release. Eight entries — every Australian state + territory.
_STATE_PRODUCTS: list[tuple[str, str, str]] = [
    ("NSW", "do002", "do003"),
    ("VIC", "do006", "do007"),
    ("QLD", "do010", "do011"),
    ("SA", "do014", "do015"),
    ("WA", "do018", "do019"),
    ("TAS", "do022", "do023"),
    ("NT", "do026", "do027"),
    ("ACT", "do030", "do031"),
]

# Output column names for the 9 ABS metrics in column order.
# Header row 4 (0-indexed) of Table_1 carries the source labels; we map
# each to a snake_case output column with the units encoded in the suffix.
_OUTPUT_COLUMNS: list[str] = [
    "new_houses_count",
    "new_other_residential_building_count",
    "total_dwellings_count",
    "value_new_houses",
    "value_new_other_residential_building",
    "value_alterations_additions_conversions",
    "value_total_residential_building",
    "value_non_residential_building",
    "value_total_building",
]

# Source header labels we use to validate the layout matches the probe.
# ABS sometimes adjusts label punctuation slightly between releases;
# matching is case-insensitive + whitespace-normalised, and lookups use
# the first contiguous prefix to be lenient about " & " vs " and "
# variants. The order is the column order of Table_1.
_EXPECTED_HEADER_PREFIXES: list[str] = [
    "new houses",
    "new other residential",
    "total dwellings",
    "value of new houses",
    "value of new other residential",
    "value of alterations",
    "value of total residential",
    "value of non-residential",
    "value of total building",
]


class AbsBaDataSource:
    """Fetch + load ABS Building Approvals at SA2 (catalogue 8731.0).

    Implements the :class:`DatasetFetcher` Protocol. Eight per-state XLSX
    cubes per release; the combined SA2-keyed DataFrame is cached as a
    parquet sidecar after the first parse.

    ``release`` is the Australian financial year (e.g. ``"2024-25"``,
    ``"2025-26"``). ``"latest"`` resolves to the most recent **complete**
    financial year. A specific FYTD request is honoured if explicitly
    asked for.
    """

    def __init__(
        self,
        *,
        release: str = "latest",
        root: Path,
        landing_url: str = ABS_BA_LANDING_URL,
        session: requests.Session | None = None,
        chunk_size: int = 256 * 1024,
        timeout: float = 60.0,
    ) -> None:
        self._release_request = str(release)
        self._root = Path(root)
        self._landing_url = landing_url
        self._session = session if session is not None else requests.Session()
        self._chunk_size = chunk_size
        self._timeout = timeout
        self._resolved_release: str | None = None
        # Resolved month suffix and per-state file URLs cached after
        # _resolve_release() runs once.
        self._yyyymm: str | None = None
        self._state_urls: dict[str, str] = {}

    # ---- DatasetFetcher protocol --------------------------------------

    @property
    def resolved_release(self) -> str:
        if self._resolved_release is None:
            self._resolve_release()
        assert self._resolved_release is not None
        return self._resolved_release

    @property
    def is_cached(self) -> bool:
        """True if all 8 per-state XLSX cache files exist for the resolved
        release. When the release hasn't been resolved yet, fall back to a
        glob probe so callers can check cheaply.
        """
        if self._resolved_release is not None:
            return all(p.exists() for p in self._state_xlsx_paths().values())
        return self._root.exists() and any(self._root.glob("abs-ba-*.xlsx"))

    def _state_xlsx_paths(self) -> dict[str, Path]:
        """Map state label -> on-disk XLSX path for the resolved release."""
        return {
            state: self._root / f"abs-ba-{self.resolved_release}-{state}.xlsx"
            for state, *_ in _STATE_PRODUCTS
        }

    @property
    def _parquet_path(self) -> Path:
        return self._root / f"abs-ba-{self.resolved_release}.parquet"

    def fetch(self, refresh: bool = False) -> Path:
        """Download all 8 per-state cubes for the resolved release.

        Returns the cache *directory* (callers usually just want
        ``load()``; the explicit return is for consistency with the
        single-file fetchers).
        """
        self._resolve_release()
        self._root.mkdir(parents=True, exist_ok=True)
        paths = self._state_xlsx_paths()
        for state, dest in paths.items():
            if dest.exists() and not refresh:
                _log.debug("ABS BA %s cached at %s", state, dest)
                continue
            url = self._state_urls.get(state)
            if not url:
                raise RuntimeError(
                    f"No resolved URL for state {state!r}; landing-page scrape "
                    f"may have missed this state. URLs found: "
                    f"{sorted(self._state_urls)}"
                )
            tmp = dest.with_suffix(dest.suffix + ".tmp")
            _log.info(
                "Downloading ABS BA %s (%s) from %s",
                state,
                self.resolved_release,
                url,
            )
            with retry_stream_get(
                self._session,
                url,
                timeout=self._timeout,
                label=f"ABS BA {state}",
            ) as response:
                response.raise_for_status()
                with tmp.open("wb") as f:
                    for chunk in response.iter_content(chunk_size=self._chunk_size):
                        if chunk:
                            f.write(chunk)
            tmp.replace(dest)
        return self._root

    def load(self) -> pd.DataFrame:
        """Return a DataFrame indexed by sa2_code_2021, columns per
        ``_OUTPUT_COLUMNS`` + ``reference_financial_year``. Combined across
        all 8 state cubes; parquet-cached after the first parse.
        """
        if self._resolved_release is not None and self._parquet_path.exists():
            return pd.read_parquet(self._parquet_path).set_index("sa2_code_2021")

        self.fetch()
        frames: list[pd.DataFrame] = []
        for state, xlsx_path in self._state_xlsx_paths().items():
            try:
                state_df = self._parse_state_xlsx(xlsx_path)
            except Exception as e:
                raise RuntimeError(
                    f"Failed to parse ABS BA cube for {state} at {xlsx_path}: {e!r}"
                ) from e
            if state_df.empty:
                _log.warning("ABS BA %s parsed to empty DataFrame at %s", state, xlsx_path)
                continue
            frames.append(state_df)

        if not frames:
            raise RuntimeError(
                f"All {len(_STATE_PRODUCTS)} ABS BA state cubes parsed to empty "
                f"DataFrames; data is missing for release {self.resolved_release}"
            )

        df = pd.concat(frames, ignore_index=True)
        # Tag rows with the reference FY so consumers can see which release.
        df["reference_financial_year"] = self.resolved_release
        df = df.set_index("sa2_code_2021")
        # De-dupe defensively — each SA2 should appear in exactly one state file.
        if df.index.has_duplicates:
            dupes = df.index[df.index.duplicated()].unique().tolist()
            raise RuntimeError(
                f"ABS BA combined DataFrame has duplicate SA2 codes "
                f"({len(dupes)} cases, e.g. {dupes[:3]!r}); a state's cube "
                f"includes SA2s that also appear in another state's file. "
                f"Investigate the upstream ABS release."
            )
        df.reset_index().to_parquet(self._parquet_path, index=False)
        return df

    # ---- release resolution -------------------------------------------

    def _resolve_release(self) -> None:
        if self._resolved_release is not None:
            return

        _log.debug("Resolving ABS BA release via %s", self._landing_url)
        resp = self._session.get(self._landing_url, timeout=self._timeout)
        resp.raise_for_status()
        html = resp.text

        # Per-state SA2 download URLs look like
        # /statistics/industry/building-and-construction/building-approvals-australia/
        # mar-2026/87310do002_202603.xlsx
        pattern = re.compile(
            r'href="([^"]*building-approvals-australia/'
            r"[a-z]{3}-\d{4}/"
            r"87310(?P<product>do\d{3})_(?P<yyyymm>\d{6})\.xlsx)\"?",
            re.IGNORECASE,
        )

        # Build product -> (yyyymm, url) — keeping only the latest yyyymm
        # if multiple appear (the page sometimes carries archive links).
        product_to_url: dict[str, tuple[str, str]] = {}
        for m in pattern.finditer(html):
            href = m.group(1)
            product = m.group("product").lower()
            yyyymm = m.group("yyyymm")
            url = (
                href
                if href.startswith("http")
                else ("https://www.abs.gov.au" + (href if href.startswith("/") else "/" + href))
            )
            existing = product_to_url.get(product)
            if existing is None or yyyymm > existing[0]:
                product_to_url[product] = (yyyymm, url)

        if not product_to_url:
            raise RuntimeError(
                f"Could not find any ABS BA per-state XLSX links on {self._landing_url}. "
                f"ABS may have changed the page layout."
            )

        # Pick latest yyyymm across all found products — they should all
        # match (ABS releases all per-state files together each month).
        latest_yyyymm = max(yyyymm for yyyymm, _ in product_to_url.values())

        # Decide which series to use for this release request:
        # - "latest" or a release matching the do002-series FY → use complete-FY series
        # - A release matching the do003-series FY → use FYTD series
        # The complete-FY release covers (yyyymm year - 1 if month < 7 else yyyymm year),
        # but rather than computing it we read both Contents sheets — actually,
        # the simpler rule: the Contents sheet of do002 says "2024-2025"; we
        # parse the Contents sheet to learn the FY each series covers, and
        # match the user's request against that. Concretely, fetch both
        # do002 and do003 of NSW, peek at Contents, then route.
        # For this v1 implementation, take a simpler stance: each monthly
        # release has fixed FY mapping — do002 is "complete FY" (previous FY
        # ended last June), do003 is current FYTD. Compute from yyyymm.
        year = int(latest_yyyymm[:4])
        month = int(latest_yyyymm[4:])
        # ABS files released in Jul-Dec X cover FY (X-1)-(X). Files released
        # in Jan-Jun X cover FY (X-1)-X also (the previous-but-one full FY
        # plus the current FYTD).
        # The do002 series is the most recently *complete* FY:
        #   - Releases in Jul X to Jun X+1 -> complete FY ending Jun X
        #     -> FY label "(X-1)-X"
        complete_fy_end = year if month >= 7 else year - 1
        complete_fy_label = f"{complete_fy_end - 1}-{str(complete_fy_end)[-2:]}"
        # The do003 series is the current FYTD:
        #   - FY label "X-(X+1)" if released Jul X..Jun X+1
        ytd_fy_label = f"{complete_fy_end}-{str(complete_fy_end + 1)[-2:]}"

        if self._release_request == "latest":
            chosen_label = complete_fy_label
            chosen_series = "complete"
        elif self._release_request == complete_fy_label:
            chosen_label = complete_fy_label
            chosen_series = "complete"
        elif self._release_request == ytd_fy_label:
            chosen_label = ytd_fy_label
            chosen_series = "ytd"
        else:
            raise RuntimeError(
                f"ABS BA release {self._release_request!r} not available in "
                f"the {latest_yyyymm} ABS release. Available: complete FY "
                f"{complete_fy_label!r} or FYTD {ytd_fy_label!r}. Older "
                f"financial years are not published in current ABS releases; "
                f"check the ABS archive for historical 8731.0 cubes."
            )

        # Pick the per-state product code based on the chosen series.
        for state, complete_product, ytd_product in _STATE_PRODUCTS:
            product = complete_product if chosen_series == "complete" else ytd_product
            entry = product_to_url.get(product)
            if entry is None:
                raise RuntimeError(
                    f"ABS BA landing page is missing the {product!r} cube "
                    f"(state {state}) for release {chosen_label}. Products "
                    f"found: {sorted(product_to_url)}."
                )
            self._state_urls[state] = entry[1]

        self._resolved_release = chosen_label
        self._yyyymm = latest_yyyymm
        _log.info(
            "Resolved ABS BA release=%s (series=%s, yyyymm=%s, %d states)",
            chosen_label,
            chosen_series,
            latest_yyyymm,
            len(self._state_urls),
        )

    # ---- parsing -------------------------------------------------------

    @staticmethod
    def _parse_state_xlsx(xlsx_path: Path) -> pd.DataFrame:
        """Parse one per-state SA2 cube into a DataFrame.

        Layout (live-probed 2026-06-01):
        - Sheet ``Table_1``
        - Row 4 (0-indexed) = column headers (mixed metric labels)
        - Row 5 = units row (``no.`` for counts, ``$'000`` for values)
        - Row 6 onwards = data
        - Column A holds mixed-level codes (state, GCC, SA4, SA3, SA2);
          we filter strictly to 9-digit numeric codes to drop aggregates.
        - Columns C onwards (0-indexed 2+) hold the 9 metric values.
        """
        import openpyxl  # noqa: PLC0415

        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
        if "Table_1" not in wb.sheetnames:
            wb.close()
            raise RuntimeError(
                f"ABS BA workbook {xlsx_path} has no 'Table_1' sheet. Sheets: {wb.sheetnames}"
            )
        ws = wb["Table_1"]
        rows: list[list[object]] = [list(r) for r in ws.iter_rows(values_only=True)]
        wb.close()

        # Locate the header row by scanning for the first row that contains
        # at least 4 of the expected column-header prefixes. Defensive against
        # small layout shifts across monthly releases.
        header_idx = -1
        for i in range(min(15, len(rows))):
            row = rows[i]
            row_text = " | ".join("" if c is None else str(c).strip().lower() for c in row)
            hits = sum(1 for p in _EXPECTED_HEADER_PREFIXES if p in row_text)
            if hits >= 4:
                header_idx = i
                break
        if header_idx < 0:
            raise RuntimeError(
                f"Could not find ABS BA column-header row in {xlsx_path}. "
                f"Looked for at least 4 of {_EXPECTED_HEADER_PREFIXES} in the "
                f"first 15 rows."
            )

        # Map each header column to its output name. Columns 0-1 are the
        # SA2 code + name; columns 2 onwards hold the metrics in the
        # order matching _OUTPUT_COLUMNS / _EXPECTED_HEADER_PREFIXES.
        header_row = rows[header_idx]
        col_to_output: dict[int, str] = {}
        # Order-preserving prefix match — for each header cell from col 2
        # onwards, take the next unused _EXPECTED_HEADER_PREFIXES entry.
        prefix_iter = iter(zip(_EXPECTED_HEADER_PREFIXES, _OUTPUT_COLUMNS, strict=False))
        for col_idx in range(2, len(header_row)):
            cell = header_row[col_idx]
            if cell is None:
                continue
            text = str(cell).strip().lower()
            if not text:
                continue
            try:
                prefix, output = next(prefix_iter)
            except StopIteration:
                break
            # Sanity: the cell's lower-cased text must contain the expected
            # prefix (case-insensitive); if it doesn't, the layout has shifted
            # and we should fail loud rather than silently map the wrong column.
            if prefix not in text:
                raise RuntimeError(
                    f"ABS BA header column {col_idx} text {text!r} does not "
                    f"contain expected prefix {prefix!r} (mapped to output "
                    f"{output!r}). The XLSX layout may have shifted; "
                    f"check {xlsx_path}."
                )
            col_to_output[col_idx] = output

        if len(col_to_output) != len(_OUTPUT_COLUMNS):
            raise RuntimeError(
                f"ABS BA parser matched {len(col_to_output)} of "
                f"{len(_OUTPUT_COLUMNS)} expected output columns in "
                f"{xlsx_path}. Header row was: {header_row!r}"
            )

        # Parse data rows. Skip the units row (header_idx + 1) and read
        # rows from header_idx + 2 onwards. Filter strictly to 9-digit
        # numeric SA2 codes; col A also holds 1-digit state codes,
        # alphanumeric GCC codes ('1GSYD'), 3-digit SA4 codes, and 5-digit
        # SA3 codes — all are aggregates we don't want.
        records: list[dict[str, object]] = []
        for row in rows[header_idx + 2 :]:
            if not row:
                continue
            sa2_raw = row[0]
            sa2 = "" if sa2_raw is None else str(sa2_raw).strip()
            if not (len(sa2) == 9 and sa2.isdigit()):
                continue
            rec: dict[str, object] = {"sa2_code_2021": sa2}
            for col_idx, output_name in col_to_output.items():
                if col_idx < len(row):
                    rec[output_name] = _coerce_number(row[col_idx])
                else:
                    rec[output_name] = None
            records.append(rec)

        # An empty state cube is suspicious in real ABS data (every state +
        # territory has SA2s) but synthetic test fixtures may legitimately
        # ship empty per-state files. Return an empty DataFrame with the
        # expected columns rather than raising — load() raises if ALL
        # 8 states come back empty.
        if not records:
            _log.warning(
                "ABS BA cube %s contained no 9-digit SA2 rows. Real ABS state "
                "cubes always have at least a few SA2s; investigate the upstream "
                "file if this happens against live data.",
                xlsx_path,
            )
            return pd.DataFrame(
                {col: pd.Series(dtype="object") for col in ["sa2_code_2021", *_OUTPUT_COLUMNS]}
            )
        return pd.DataFrame.from_records(records)


def _coerce_number(cell: object) -> object:
    """ABS BA cells are int counts or float values; treat blank/dash as None.
    Source publishes raw counts (no suppression at SA2 level), but defensively
    handle the same null sentinels as the other ABS datasets.
    """
    if cell is None:
        return None
    if isinstance(cell, bool):
        # openpyxl can occasionally surface Python bools; treat as int.
        return int(cell)
    if isinstance(cell, (int, float)):
        return cell
    s = str(cell).strip()
    if not s or s.lower() in ("np", "na", "n/a", "-", "..", ".", "nan", "null"):
        return None
    if re.fullmatch(r"-?\d+", s):
        return int(s)
    if re.fullmatch(r"-?\d+\.\d+", s):
        return float(s)
    return None


# ---- fetcher registration ------------------------------------------------


def _build_fetcher(root: Path, release: str | None = None) -> AbsBaDataSource:
    kwargs: dict[str, object] = {"root": root}
    if release is not None:
        kwargs["release"] = release
    return AbsBaDataSource(**kwargs)  # type: ignore[arg-type]


def _register() -> None:
    from . import registry  # noqa: PLC0415

    registry.register_fetcher("abs_building_approvals", _build_fetcher)


_register()
