"""ABS Counts of Australian Businesses (catalogue 8165.0 / CABEE)
fetcher (spec §20, dataset id ``abs_business_counts``).

ABS publishes business counts annually as a single national XLSX data
cube. Cube **DC8** ("Businesses by Industry Division by Statistical Area
Level 2 by Annualised Employment Size Ranges") is SA2-native — no
downscale needed. The augmentor surfaces the per-SA2 business counts by
employment-size band, summed across all industry divisions.

Real-data findings (live-probed 2026-06-10 against the
``jul2021-jun2025`` release):

- One workbook (``8165DC08.xlsx``, ~8 MB) carries **three reference
  years** in three sheets: ``Table 1`` = June 2025, ``Table 2`` = June
  2024, ``Table 3 `` (note the trailing space) = June 2023. The
  ``release`` selects which year/sheet the parser surfaces.
- Each ``Table`` sheet: rows 1-3 banner, row 4 title, a **2-row header
  band** (row 5 = size-band labels, row 6 = ``Code``/``Label``/``no.``),
  data from row 7. Columns: Industry Code, Industry Label, **SA2 Code**
  (9-digit), SA2 Label, then 5 employment-size bands (Non employing /
  1-4 / 5-19 / 20-199 / 200+) and a Total.
- **Long format**: one row per (industry division × SA2). There are 20
  industry rows per SA2 (ANZSIC divisions A-S + ``X`` "Currently
  Unknown") and NO per-SA2 total row — so the per-SA2 figure is the
  **sum across the 20 industry rows**. The only ``Total All Industries``
  rows are national totals at the file tail with a blank SA2 code.
- Footnote rows + national-total rows have a blank/non-numeric SA2 code,
  so a strict 9-digit-code filter drops them cleanly.
- ABS perturbs cell values (footnote (b)); division/state/size/Australia
  totals are not perturbed. So the summed size bands may not add to the
  summed Total exactly — surfaced as published.

URL is fully deterministic (no HTML scrape, per CLAUDE.md): hardcoded per
release. A new annual workbook (e.g. ``jul2022-jun2026`` from Aug 2026)
needs a new entry in ``_ABS_CAB_RELEASES``.
"""

from __future__ import annotations

import logging
import re
from pathlib import Path

import pandas as pd
import requests

from .._http_retry import retry_stream_get

_log = logging.getLogger(__name__)

# Reference year -> (workbook url, sheet name, a title marker to verify
# the sheet is the year we think it is). All three years currently live
# in the one jul2021-jun2025 workbook. New releases add a year here.
_ABS_CAB_RELEASES: dict[str, dict[str, str]] = {
    "2025": {
        "url": (
            "https://www.abs.gov.au/statistics/economy/business-indicators/"
            "counts-australian-businesses-including-entries-and-exits/"
            "jul2021-jun2025/8165DC08.xlsx"
        ),
        "sheet": "Table 1",
        "title_marker": "June 2025",
    },
    "2024": {
        "url": (
            "https://www.abs.gov.au/statistics/economy/business-indicators/"
            "counts-australian-businesses-including-entries-and-exits/"
            "jul2021-jun2025/8165DC08.xlsx"
        ),
        "sheet": "Table 2",
        "title_marker": "June 2024",
    },
    "2023": {
        "url": (
            "https://www.abs.gov.au/statistics/economy/business-indicators/"
            "counts-australian-businesses-including-entries-and-exits/"
            "jul2021-jun2025/8165DC08.xlsx"
        ),
        "sheet": "Table 3 ",
        "title_marker": "June 2023",
    },
}

# Employment-size band headers (row 5, cols E-J / 0-indexed 4-9), in
# order, matched case-insensitively against the real header band, and the
# snake_case output column each maps to.
_SIZE_HEADER_PREFIXES: list[str] = [
    "non employing",
    "1-4",
    "5-19",
    "20-199",
    "200+",
    "total",
]
_OUTPUT_COLUMNS: list[str] = [
    "business_count_non_employing",
    "business_count_1_4_employees",
    "business_count_5_19_employees",
    "business_count_20_199_employees",
    "business_count_200_plus_employees",
    "business_count_total",
]
# 0-indexed sheet columns holding the SA2 code and the 6 size-band values.
_SA2_CODE_COL = 2
_FIRST_VALUE_COL = 4


class AbsBusinessCountsDataSource:
    """Fetch + load ABS Counts of Australian Businesses at SA2 (8165.0).

    Implements the :class:`DatasetFetcher` Protocol. SA2-native — no
    SA2→SA4 downscale. The per-SA2 business counts by employment-size
    band (summed across industry divisions) are parquet-cached after the
    first parse.

    Args:
        release: Reference year (``"2025"`` / ``"2024"`` / ``"2023"``) or
            ``"latest"`` (the most recent year in the registry).
        root: Cache directory for the workbook + parquet sidecar.
        session: Optional ``requests.Session`` (tests pass a hermetic one).
        chunk_size / timeout: As per the other ABS fetchers.
    """

    def __init__(
        self,
        *,
        release: str = "latest",
        root: Path,
        session: requests.Session | None = None,
        chunk_size: int = 256 * 1024,
        timeout: float = 60.0,
    ) -> None:
        self._release_request = str(release)
        self._root = Path(root)
        self._session = session if session is not None else requests.Session()
        self._chunk_size = chunk_size
        self._timeout = timeout
        self._resolved_release: str | None = None
        self._resolved_url: str | None = None
        self._resolved_sheet: str | None = None
        self._resolved_marker: str | None = None

    # ---- DatasetFetcher protocol --------------------------------------

    @property
    def resolved_release(self) -> str:
        if self._resolved_release is None:
            self._resolve_release()
        assert self._resolved_release is not None
        return self._resolved_release

    @property
    def is_cached(self) -> bool:
        if self._resolved_release is not None:
            return self._workbook_path.exists()
        return self._root.exists() and any(self._root.glob("abs-cab-dc08-*.xlsx"))

    @property
    def _workbook_slug(self) -> str:
        """Stable slug for the workbook (shared across the 3 years it
        carries), derived from the URL's release path segment."""
        url = self._resolved_url or ""
        # .../counts-.../jul2021-jun2025/8165DC08.xlsx -> "jul2021-jun2025"
        parts = url.rstrip("/").split("/")
        return parts[-2] if len(parts) >= 2 else "release"

    @property
    def _workbook_path(self) -> Path:
        return self._root / f"abs-cab-dc08-{self._workbook_slug}.xlsx"

    @property
    def _parquet_path(self) -> Path:
        return self._root / f"abs-cab-{self.resolved_release}.parquet"

    def fetch(self, refresh: bool = False) -> Path:
        """Download the ABS CAB workbook (shared across its 3 years)."""
        self._resolve_release()
        if self._workbook_path.exists() and not refresh:
            _log.debug("ABS CAB cached at %s", self._workbook_path)
            return self._workbook_path

        self._root.mkdir(parents=True, exist_ok=True)
        tmp = self._workbook_path.with_suffix(self._workbook_path.suffix + ".tmp")
        url = self._resolved_url or ""
        _log.info("Downloading ABS CAB (%s) from %s", self.resolved_release, url)
        with retry_stream_get(
            self._session,
            url,
            timeout=self._timeout,
            label="ABS CAB",
        ) as response:
            response.raise_for_status()
            with tmp.open("wb") as f:
                for chunk in response.iter_content(chunk_size=self._chunk_size):
                    if chunk:
                        f.write(chunk)
        tmp.replace(self._workbook_path)
        _log.info("Saved ABS CAB to %s", self._workbook_path)
        return self._workbook_path

    def load(self) -> pd.DataFrame:
        """Return a DataFrame indexed by ``sa2_code_2021`` with the
        per-SA2 business counts by employment-size band + reference year.
        """
        if self._resolved_release is not None and self._parquet_path.exists():
            return pd.read_parquet(self._parquet_path).set_index("sa2_code_2021")

        workbook = self.fetch()
        df = self._parse_workbook(
            workbook,
            sheet=self._resolved_sheet or "",
            title_marker=self._resolved_marker or "",
        )
        df["reference_period"] = self.resolved_release
        df = df.set_index("sa2_code_2021")
        df.reset_index().to_parquet(self._parquet_path, index=False)
        return df

    # ---- release resolution -------------------------------------------

    def _resolve_release(self) -> None:
        if self._resolved_release is not None:
            return

        if self._release_request == "latest":
            picked = max(_ABS_CAB_RELEASES)
        elif self._release_request in _ABS_CAB_RELEASES:
            picked = self._release_request
        else:
            raise RuntimeError(
                f"ABS CAB release {self._release_request!r} not in the registry. "
                f"Available: {sorted(_ABS_CAB_RELEASES)}. New annual workbooks "
                f"need an entry added to _ABS_CAB_RELEASES in "
                f"src/census_augment/datasets/_abs_cab.py."
            )
        entry = _ABS_CAB_RELEASES[picked]
        self._resolved_release = picked
        self._resolved_url = entry["url"]
        self._resolved_sheet = entry["sheet"]
        self._resolved_marker = entry["title_marker"]
        _log.info(
            "Resolved ABS CAB release=%s, sheet=%s, url=%s",
            picked,
            entry["sheet"],
            entry["url"],
        )

    # ---- parsing -------------------------------------------------------

    @staticmethod
    def _parse_workbook(workbook_path: Path, *, sheet: str, title_marker: str) -> pd.DataFrame:
        """Parse one ``Table`` sheet of the DC8 cube into a DataFrame of
        per-SA2 business counts by employment-size band.

        Sums the 20 industry-division rows per SA2 (there is no per-SA2
        total row in the source). Validates the title marker + size-band
        header band against the real layout (live-probed 2026-06-10).
        """
        import openpyxl  # noqa: PLC0415

        wb = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
        if sheet not in wb.sheetnames:
            available = wb.sheetnames
            wb.close()
            raise RuntimeError(
                f"ABS CAB workbook {workbook_path} has no {sheet!r} sheet. "
                f"Sheets: {available}. The release→sheet map in "
                f"_ABS_CAB_RELEASES may be stale."
            )
        ws = wb[sheet]
        rows: list[list[object]] = [list(r) for r in ws.iter_rows(values_only=True)]
        wb.close()

        # Drift guard: the per-sheet TITLE row must name the reference
        # period we mapped this sheet to. Check the title row specifically,
        # NOT the whole banner — the workbook subtitle row always reads
        # "... June 2021 to June 2025" (the latest year) on every sheet, so
        # a banner-wide check would never catch a sheet→year mismatch. The
        # title row is the one naming "Statistical Area Level 2" (the
        # subtitle row doesn't).
        title_line: str | None = None
        for row in rows[:8]:
            joined = " ".join("" if c is None else str(c) for c in row)
            if "statistical area level 2" in joined.lower():
                title_line = joined
                break
        if title_line is None:
            raise RuntimeError(
                f"ABS CAB sheet {sheet!r} in {workbook_path} has no title row "
                f"naming 'Statistical Area Level 2' in its first 8 rows — the "
                f"layout may have changed; re-probe the cube."
            )
        if title_marker and title_marker not in title_line:
            raise RuntimeError(
                f"ABS CAB sheet {sheet!r} in {workbook_path} title does not name "
                f"{title_marker!r} — the release→sheet map in _ABS_CAB_RELEASES "
                f"may be stale (ABS re-ordered the sheets?). Title seen: "
                f"{title_line[:160]!r}"
            )

        # Find the size-band header row: SA2 code column header == 'SA2'
        # and the first value column starts the 'Non employing' band.
        header_idx = -1
        for i in range(min(15, len(rows))):
            row = rows[i]
            if len(row) <= _FIRST_VALUE_COL:
                continue
            code_hdr = str(row[_SA2_CODE_COL]).strip().lower() if row[_SA2_CODE_COL] else ""
            first_val_hdr = (
                str(row[_FIRST_VALUE_COL]).strip().lower() if row[_FIRST_VALUE_COL] else ""
            )
            if code_hdr == "sa2" and first_val_hdr.startswith("non employ"):
                header_idx = i
                break
        if header_idx < 0:
            raise RuntimeError(
                f"Could not find the size-band header row in ABS CAB sheet "
                f"{sheet!r} of {workbook_path} (looked for 'SA2' in col "
                f"{_SA2_CODE_COL} + 'Non employing' in col {_FIRST_VALUE_COL})."
            )

        # Validate each size-band header against the expected order so a
        # column reshuffle fails loud rather than mislabelling silently.
        header_row = rows[header_idx]
        for offset, prefix in enumerate(_SIZE_HEADER_PREFIXES):
            col = _FIRST_VALUE_COL + offset
            cell = str(header_row[col]).strip().lower() if col < len(header_row) else ""
            if prefix not in cell:
                raise RuntimeError(
                    f"ABS CAB size-band header col {col} is {cell!r}, expected "
                    f"to contain {prefix!r} (-> {_OUTPUT_COLUMNS[offset]!r}). "
                    f"Layout may have shifted in {workbook_path}."
                )

        # Data rows start 2 below the band-label row (skip the Code/Label/
        # no. sub-header). Keep only 9-digit SA2 codes (drops the national
        # 'Total All Industries' rows with blank SA2 + footnotes).
        value_cols = list(range(_FIRST_VALUE_COL, _FIRST_VALUE_COL + len(_OUTPUT_COLUMNS)))
        records: list[dict[str, object]] = []
        for row in rows[header_idx + 2 :]:
            if len(row) <= _SA2_CODE_COL:
                continue
            sa2_raw = row[_SA2_CODE_COL]
            sa2 = "" if sa2_raw is None else str(sa2_raw).strip()
            if not (len(sa2) == 9 and sa2.isdigit()):
                continue
            rec: dict[str, object] = {"sa2_code_2021": sa2}
            for output_name, col in zip(_OUTPUT_COLUMNS, value_cols, strict=True):
                rec[output_name] = _coerce_count(row[col] if col < len(row) else None)
            records.append(rec)

        if not records:
            raise RuntimeError(
                f"ABS CAB sheet {sheet!r} in {workbook_path} produced no 9-digit "
                f"SA2 rows. The layout may have changed — re-probe the cube."
            )

        raw = pd.DataFrame.from_records(records)
        # Sum the 20 industry-division rows per SA2 (no per-SA2 total row
        # exists in the source). NaN treated as 0 by groupby.sum.
        grouped = raw.groupby("sa2_code_2021", as_index=False)[_OUTPUT_COLUMNS].sum()
        for count_col in _OUTPUT_COLUMNS:
            grouped[count_col] = grouped[count_col].astype("Int64")
        return grouped


def _coerce_count(cell: object) -> object:
    """ABS CAB cells are integer counts; treat blank/dash/np as None."""
    if cell is None:
        return None
    if isinstance(cell, bool):
        return int(cell)
    if isinstance(cell, (int, float)):
        return cell
    s = str(cell).strip()
    if not s or s.lower() in ("np", "na", "n/a", "-", "..", ".", "nan", "null"):
        return None
    if re.fullmatch(r"-?\d+", s):
        return int(s)
    if re.fullmatch(r"-?\d+\.\d+", s):
        return float(s)
    return None


# ---- fetcher registration ------------------------------------------------


def _build_fetcher(root: Path, release: str | None = None) -> AbsBusinessCountsDataSource:
    kwargs: dict[str, object] = {"root": root}
    if release is not None:
        kwargs["release"] = release
    return AbsBusinessCountsDataSource(**kwargs)  # type: ignore[arg-type]


def _register() -> None:
    from . import registry  # noqa: PLC0415

    registry.register_fetcher("abs_business_counts", _build_fetcher)


_register()
