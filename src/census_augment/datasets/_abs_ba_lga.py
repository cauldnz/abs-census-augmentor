"""ABS Building Approvals (catalogue 8731.0) LGA-keyed fetcher (spec §20).

Sibling to the SA2-native :class:`AbsBaDataSource` in ``_abs_ba``. ABS
publishes the same source data at both SA2 and LGA granularities; this
fetcher reads the LGA cubes and downscales LGA values to SA2 via an
area-weighted spatial correspondence
(:class:`census_augment.correspondence.LgaSa2Correspondence`).

First production dataset to exercise the LGA-SA2 spatial correspondence
machinery introduced in v2.2.0 / PR #107.

Key real-data findings (live-probed 2026-06-01) vs the SA2 cube:

- LGA cube data sheet is **``Table 1``** (with space), NOT ``Table_1``
  (with underscore) as in the SA2 cube. The two parsers do NOT share
  the sheet-name string.
- Otherwise row layout matches: header at row 4 (0-indexed), units at
  row 5, data from row 6.
- Column A is 5-digit LGA codes (10000-89999) — much simpler than the
  SA2 cube's mixed-level codes; the parser filters strictly to 5-digit
  numerics to drop the single state-aggregate row at the top.

Per-state product codes are offset by 2 from the SA2 series
(``do004``/``do008``/... for complete FY; ``do005``/``do009``/... for
FYTD).

Cross-level downscale requires a :class:`LgaSa2Correspondence`
attached before ``load()`` — analogous to AIHW MH Prescriptions'
``attach_sa2_to_sa4_mapping``. Without it, ``load()`` raises a clear
error explaining how to attach one.
"""

from __future__ import annotations

import logging
import re
from pathlib import Path

import pandas as pd
import requests

from .._http_retry import retry_stream_get
from ._abs_ba import ABS_BA_LANDING_URL

_log = logging.getLogger(__name__)


# Per-state LGA product numbers: (state_label, complete-FY product, FYTD
# product). Offset by 2 from the SA2 series. Confirmed live 2026-06-01
# via the NSW LGA Mar 2026 cube schema probe; the other states follow
# the same naming convention by symmetry with the SA2 series.
_STATE_PRODUCTS: list[tuple[str, str, str]] = [
    ("NSW", "do004", "do005"),
    ("VIC", "do008", "do009"),
    ("QLD", "do012", "do013"),
    ("SA", "do016", "do017"),
    ("WA", "do020", "do021"),
    ("TAS", "do024", "do025"),
    ("NT", "do028", "do029"),
    ("ACT", "do032", "do033"),
]

# Output column names in column order. Identical to the SA2 fetcher
# since the metric set is the same — only the input geography differs.
_OUTPUT_COLUMNS: list[str] = [
    "new_houses_count",
    "new_other_residential_building_count",
    "total_dwellings_count",
    "value_new_houses",
    "value_new_other_residential_building",
    "value_alterations_additions_conversions",
    "value_total_residential_building",
    "value_non_residential_building",
    "value_total_building",
]

# Source header label prefixes for resilient header-row detection
# (identical to the SA2 fetcher's set).
_EXPECTED_HEADER_PREFIXES: list[str] = [
    "new houses",
    "new other residential",
    "total dwellings",
    "value of new houses",
    "value of new other residential",
    "value of alterations",
    "value of total residential",
    "value of non-residential",
    "value of total building",
]


class AbsBaLgaDataSource:
    """Fetch + load ABS Building Approvals LGA cubes, downscaled to SA2.

    Implements the :class:`DatasetFetcher` Protocol. Eight per-state LGA
    XLSX cubes per release; LGA-keyed values are downscaled to SA2 via
    a :class:`LgaSa2Correspondence` attached before ``load()``.
    """

    def __init__(
        self,
        *,
        release: str = "latest",
        root: Path,
        landing_url: str = ABS_BA_LANDING_URL,
        session: requests.Session | None = None,
        chunk_size: int = 256 * 1024,
        timeout: float = 60.0,
    ) -> None:
        self._release_request = str(release)
        self._root = Path(root)
        self._landing_url = landing_url
        self._session = session if session is not None else requests.Session()
        self._chunk_size = chunk_size
        self._timeout = timeout
        self._resolved_release: str | None = None
        self._state_urls: dict[str, str] = {}
        # LGA-SA2 correspondence; attached via ``attach_correspondence``.
        # None means "not yet attached" — load() raises until it's set.
        # Typed as ``object`` to avoid a hard import on the
        # correspondence module at fetcher-registration time; callers
        # always attach an actual :class:`LgaSa2Correspondence`.
        self._correspondence: object | None = None

    # ---- correspondence attachment -----------------------------------

    def attach_correspondence(self, corr: object) -> None:
        """Attach the LGA-SA2 spatial correspondence used to downscale.

        Pipeline.from_config wires this automatically by fetching the
        LGA boundary, intersecting it with the SA2 boundary, and
        caching the result. Library callers can construct a
        :class:`LgaSa2Correspondence` themselves via
        ``census_augment.correspondence.compute_lga_sa2_correspondence``
        and attach it here.

        Args:
            corr: A :class:`LgaSa2Correspondence` (or any object exposing
                ``downscale_counts(lga_values) -> dict[str, float]``).
        """
        from ..correspondence import LgaSa2Correspondence  # noqa: PLC0415

        if not isinstance(corr, LgaSa2Correspondence):
            raise TypeError(
                f"attach_correspondence expects a LgaSa2Correspondence; got {type(corr).__name__}"
            )
        self._correspondence = corr
        _log.debug(
            "AbsBaLgaDataSource: attached correspondence with %d (SA2, LGA) pairs",
            len(corr.weights),
        )

    # ---- DatasetFetcher protocol --------------------------------------

    @property
    def resolved_release(self) -> str:
        if self._resolved_release is None:
            self._resolve_release()
        assert self._resolved_release is not None
        return self._resolved_release

    @property
    def is_cached(self) -> bool:
        if self._resolved_release is not None:
            return all(p.exists() for p in self._state_xlsx_paths().values())
        return self._root.exists() and any(self._root.glob("abs-ba-lga-*.xlsx"))

    def _state_xlsx_paths(self) -> dict[str, Path]:
        return {
            state: self._root / f"abs-ba-lga-{self.resolved_release}-{state}.xlsx"
            for state, *_ in _STATE_PRODUCTS
        }

    @property
    def _parquet_path(self) -> Path:
        return self._root / f"abs-ba-lga-{self.resolved_release}.parquet"

    def fetch(self, refresh: bool = False) -> Path:
        """Download all 8 per-state LGA cubes for the resolved release."""
        self._resolve_release()
        self._root.mkdir(parents=True, exist_ok=True)
        paths = self._state_xlsx_paths()
        for state, dest in paths.items():
            if dest.exists() and not refresh:
                _log.debug("ABS BA LGA %s cached at %s", state, dest)
                continue
            url = self._state_urls.get(state)
            if not url:
                raise RuntimeError(
                    f"No resolved URL for state {state!r}; landing-page scrape "
                    f"may have missed this state. URLs found: "
                    f"{sorted(self._state_urls)}"
                )
            tmp = dest.with_suffix(dest.suffix + ".tmp")
            _log.info(
                "Downloading ABS BA LGA %s (%s) from %s",
                state,
                self.resolved_release,
                url,
            )
            with retry_stream_get(
                self._session,
                url,
                timeout=self._timeout,
                label=f"ABS BA LGA {state}",
            ) as response:
                response.raise_for_status()
                with tmp.open("wb") as f:
                    for chunk in response.iter_content(chunk_size=self._chunk_size):
                        if chunk:
                            f.write(chunk)
            tmp.replace(dest)
        return self._root

    def load(self) -> pd.DataFrame:
        """Return a DataFrame indexed by ``sa2_code_2021`` with one row
        per SA2 (downscaled from the LGA-level source via the attached
        :class:`LgaSa2Correspondence`).

        Requires :meth:`attach_correspondence` to have been called
        first — without it, raises a clear ``RuntimeError`` since
        LGA-keyed output isn't useful in the rest of the pipeline.
        """
        if self._correspondence is None:
            raise RuntimeError(
                "AbsBaLgaDataSource.load() requires a LgaSa2Correspondence "
                "to be attached first. Call attach_correspondence(corr) "
                "with the result of "
                "`census_augment.correspondence.compute_lga_sa2_correspondence("
                "sa2=<sa2_gdf>, lga=<lga_gdf>)`. "
                "Pipeline.from_config wires this automatically."
            )

        if self._resolved_release is not None and self._parquet_path.exists():
            return pd.read_parquet(self._parquet_path).set_index("sa2_code_2021")

        self.fetch()
        # Parse each per-state XLSX into a LGA-keyed DataFrame, then
        # concat. LGA codes are 5-digit numeric strings (matching the
        # boundary's LGA_CODE25 attribute).
        frames: list[pd.DataFrame] = []
        for state, xlsx_path in self._state_xlsx_paths().items():
            try:
                state_df = self._parse_state_xlsx(xlsx_path)
            except Exception as e:
                raise RuntimeError(
                    f"Failed to parse ABS BA LGA cube for {state} at {xlsx_path}: {e!r}"
                ) from e
            if not state_df.empty:
                frames.append(state_df)

        if not frames:
            raise RuntimeError(
                f"All {len(_STATE_PRODUCTS)} ABS BA LGA state cubes parsed "
                f"to empty DataFrames; data is missing for release "
                f"{self.resolved_release}"
            )

        lga_df = pd.concat(frames, ignore_index=True)
        if lga_df["lga_code"].duplicated().any():
            dupes = lga_df.loc[lga_df["lga_code"].duplicated(), "lga_code"].unique().tolist()
            raise RuntimeError(
                f"ABS BA LGA combined DataFrame has duplicate LGA codes "
                f"({len(dupes)} cases, e.g. {dupes[:3]!r}); an LGA's row "
                f"appears in multiple state cubes. Investigate."
            )
        lga_df = lga_df.set_index("lga_code")

        # Downscale every metric column from LGA to SA2 via the
        # correspondence. Each metric is additive (counts of approvals
        # or sums of $'000 values), so `downscale_counts` is the right
        # call — per-LGA sum is preserved across the SA2s that overlap
        # each LGA.
        from ..correspondence import LgaSa2Correspondence  # noqa: PLC0415

        # `_correspondence` is verified non-None above and validated as
        # an instance in attach_correspondence; the assertion narrows the
        # type for mypy.
        corr: LgaSa2Correspondence = self._correspondence  # type: ignore[assignment]

        # Track LGAs in the source that aren't in the correspondence so
        # we can warn loudly — a boundary-vs-publication mismatch
        # (e.g. ABS published values for an LGA that was abolished by a
        # recent boundary release we're using).
        corr_lgas = set(corr.weights["lga_code"].unique())
        source_lgas = set(lga_df.index)
        missing_in_corr = source_lgas - corr_lgas
        if missing_in_corr:
            _log.warning(
                "ABS BA LGA: %d source LGA(s) not in the LGA-SA2 "
                "correspondence (sample: %s); their values won't contribute "
                "to any SA2. Boundary-publication mismatch — check the "
                "LGA boundary release vs the ABS publication date.",
                len(missing_in_corr),
                sorted(missing_in_corr)[:5],
            )

        # Build a per-SA2 record by accumulating contributions from each
        # column. We could loop column-by-column calling
        # `corr.downscale_counts({lga: value for lga, value in column.items()})`
        # but doing all columns in one pass over the weights frame is
        # more efficient.
        weights = corr.weights[["sa2_code", "lga_code", "sa2_share_of_lga"]].copy()
        # Inner join — only LGAs that appear in both the source and the
        # correspondence contribute. LGAs in source but not in corr
        # were warned above; LGAs in corr but not in source just contribute
        # zero to those SA2s (handled implicitly by reindex below).
        joined = weights.merge(lga_df, left_on="lga_code", right_index=True, how="inner")
        per_sa2_records: dict[str, dict[str, float]] = {}
        for sa2_code, group in joined.groupby("sa2_code"):
            sa2_record: dict[str, float] = {}
            group_share = group["sa2_share_of_lga"].astype("float64")
            for col in _OUTPUT_COLUMNS:
                # value × area-share contribution per (sa2, lga) row,
                # then sum across the LGAs that contribute to this SA2.
                contribution = (group[col].astype("float64") * group_share).sum()
                sa2_record[col] = float(contribution)
            per_sa2_records[str(sa2_code)] = sa2_record

        if not per_sa2_records:
            raise RuntimeError(
                "ABS BA LGA: downscale produced zero SA2 records; the "
                "attached correspondence and the LGA cube data don't overlap "
                "anywhere. Check boundary release alignment."
            )

        out = pd.DataFrame.from_dict(per_sa2_records, orient="index")
        out.index.name = "sa2_code_2021"
        out["reference_financial_year"] = self.resolved_release
        # Cast to nullable dtypes that round-trip cleanly through parquet
        # and preserve NaN semantics.
        out.reset_index().to_parquet(self._parquet_path, index=False)
        return out

    # ---- release resolution -------------------------------------------

    def _resolve_release(self) -> None:
        if self._resolved_release is not None:
            return

        _log.debug("Resolving ABS BA LGA release via %s", self._landing_url)
        resp = self._session.get(self._landing_url, timeout=self._timeout)
        resp.raise_for_status()
        html = resp.text

        # Same URL pattern as the SA2 cube — only the product code
        # filter differs. Scrape both series; the routing logic below
        # picks the right one per the requested release.
        pattern = re.compile(
            r'href="([^"]*building-approvals-australia/'
            r"[a-z]{3}-\d{4}/"
            r"87310(?P<product>do\d{3})_(?P<yyyymm>\d{6})\.xlsx)\"?",
            re.IGNORECASE,
        )

        product_to_url: dict[str, tuple[str, str]] = {}
        for m in pattern.finditer(html):
            href = m.group(1)
            product = m.group("product").lower()
            yyyymm = m.group("yyyymm")
            url = (
                href
                if href.startswith("http")
                else ("https://www.abs.gov.au" + (href if href.startswith("/") else "/" + href))
            )
            existing = product_to_url.get(product)
            if existing is None or yyyymm > existing[0]:
                product_to_url[product] = (yyyymm, url)

        if not product_to_url:
            raise RuntimeError(
                f"Could not find any ABS BA per-state XLSX links on "
                f"{self._landing_url}. ABS may have changed the page layout."
            )

        # Take the latest yyyymm and decide the series the same way as
        # the SA2 cube. Mar 2026 release (yyyymm=202603) covers FY
        # 2024-25 complete + FY 2025-26 FYTD.
        latest_yyyymm = max(yyyymm for yyyymm, _ in product_to_url.values())
        year = int(latest_yyyymm[:4])
        month = int(latest_yyyymm[4:])
        complete_fy_end = year if month >= 7 else year - 1
        complete_fy_label = f"{complete_fy_end - 1}-{str(complete_fy_end)[-2:]}"
        ytd_fy_label = f"{complete_fy_end}-{str(complete_fy_end + 1)[-2:]}"

        if self._release_request == "latest":
            chosen_label = complete_fy_label
            chosen_series = "complete"
        elif self._release_request == complete_fy_label:
            chosen_label = complete_fy_label
            chosen_series = "complete"
        elif self._release_request == ytd_fy_label:
            chosen_label = ytd_fy_label
            chosen_series = "ytd"
        else:
            raise RuntimeError(
                f"ABS BA LGA release {self._release_request!r} not available "
                f"in the {latest_yyyymm} ABS release. Available: complete FY "
                f"{complete_fy_label!r} or FYTD {ytd_fy_label!r}."
            )

        for state, complete_product, ytd_product in _STATE_PRODUCTS:
            product = complete_product if chosen_series == "complete" else ytd_product
            entry = product_to_url.get(product)
            if entry is None:
                raise RuntimeError(
                    f"ABS BA LGA landing page is missing the {product!r} cube "
                    f"(state {state}) for release {chosen_label}. Products "
                    f"found: {sorted(product_to_url)}."
                )
            self._state_urls[state] = entry[1]

        self._resolved_release = chosen_label
        _log.info(
            "Resolved ABS BA LGA release=%s (series=%s, yyyymm=%s, %d states)",
            chosen_label,
            chosen_series,
            latest_yyyymm,
            len(self._state_urls),
        )

    # ---- parsing -------------------------------------------------------

    @staticmethod
    def _parse_state_xlsx(xlsx_path: Path) -> pd.DataFrame:
        """Parse one per-state LGA cube into a LGA-keyed DataFrame.

        Real-data layout (live-probed 2026-06-01 on NSW Mar 2026 cube):
        - Sheet ``Table 1`` (with space — DIFFERENT from the SA2 cube's
          ``Table_1`` with underscore)
        - Row 4 (0-indexed) = column headers
        - Row 5 = units row (``no.`` / ``$'000``)
        - Row 6 onwards = data
        - Column A holds 5-digit LGA codes (plus the single state-
          aggregate row with a 1-digit code at the top). Filter
          strictly to 5-digit numeric codes.
        """
        import openpyxl  # noqa: PLC0415

        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
        # Real-data finding: LGA cubes use "Table 1" (space), SA2 cubes
        # use "Table_1" (underscore). Don't share the constant.
        if "Table 1" not in wb.sheetnames:
            wb.close()
            raise RuntimeError(
                f"ABS BA LGA workbook {xlsx_path} has no 'Table 1' sheet "
                f"(note: space, not underscore). Sheets: {wb.sheetnames}"
            )
        ws = wb["Table 1"]
        rows: list[list[object]] = [list(r) for r in ws.iter_rows(values_only=True)]
        wb.close()

        # Locate the header row by matching expected prefix labels —
        # same resilient scan as the SA2 parser.
        header_idx = -1
        for i in range(min(15, len(rows))):
            row = rows[i]
            row_text = " | ".join("" if c is None else str(c).strip().lower() for c in row)
            hits = sum(1 for p in _EXPECTED_HEADER_PREFIXES if p in row_text)
            if hits >= 4:
                header_idx = i
                break
        if header_idx < 0:
            raise RuntimeError(
                f"Could not find ABS BA LGA column-header row in {xlsx_path}. "
                f"Looked for at least 4 of {_EXPECTED_HEADER_PREFIXES} in "
                f"the first 15 rows."
            )

        header_row = rows[header_idx]
        col_to_output: dict[int, str] = {}
        prefix_iter = iter(zip(_EXPECTED_HEADER_PREFIXES, _OUTPUT_COLUMNS, strict=False))
        for col_idx in range(2, len(header_row)):
            cell = header_row[col_idx]
            if cell is None:
                continue
            text = str(cell).strip().lower()
            if not text:
                continue
            try:
                prefix, output = next(prefix_iter)
            except StopIteration:
                break
            if prefix not in text:
                raise RuntimeError(
                    f"ABS BA LGA header column {col_idx} text {text!r} does "
                    f"not contain expected prefix {prefix!r} (mapped to "
                    f"output {output!r}). Layout may have shifted; check "
                    f"{xlsx_path}."
                )
            col_to_output[col_idx] = output

        if len(col_to_output) != len(_OUTPUT_COLUMNS):
            raise RuntimeError(
                f"ABS BA LGA parser matched {len(col_to_output)} of "
                f"{len(_OUTPUT_COLUMNS)} expected output columns in "
                f"{xlsx_path}. Header row was: {header_row!r}"
            )

        records: list[dict[str, object]] = []
        for row in rows[header_idx + 2 :]:
            if not row:
                continue
            lga_raw = row[0]
            lga = "" if lga_raw is None else str(lga_raw).strip()
            # Real ABS LGA codes are 5-digit numerics. Filter strictly
            # to drop the single state-aggregate row at the top
            # (1-digit code like "1" for NSW).
            if not (len(lga) == 5 and lga.isdigit()):
                continue
            rec: dict[str, object] = {"lga_code": lga}
            for col_idx, output_name in col_to_output.items():
                if col_idx < len(row):
                    rec[output_name] = _coerce_number(row[col_idx])
                else:
                    rec[output_name] = None
            records.append(rec)

        if not records:
            _log.warning(
                "ABS BA LGA cube %s contained no 5-digit LGA rows. Real ABS "
                "state cubes always have LGAs; investigate the upstream file.",
                xlsx_path,
            )
            return pd.DataFrame(
                {col: pd.Series(dtype="object") for col in ["lga_code", *_OUTPUT_COLUMNS]}
            )
        return pd.DataFrame.from_records(records)


def _coerce_number(cell: object) -> object:
    """LGA cube cells are int counts or float values; handle blank/dash
    sentinels as None for consistency with the other ABS datasets.
    """
    if cell is None:
        return None
    if isinstance(cell, bool):
        return int(cell)
    if isinstance(cell, (int, float)):
        return cell
    s = str(cell).strip()
    if not s or s.lower() in ("np", "na", "n/a", "-", "..", ".", "nan", "null"):
        return None
    if re.fullmatch(r"-?\d+", s):
        return int(s)
    if re.fullmatch(r"-?\d+\.\d+", s):
        return float(s)
    return None


# ---- fetcher registration ------------------------------------------------


def _build_fetcher(root: Path, release: str | None = None) -> AbsBaLgaDataSource:
    kwargs: dict[str, object] = {"root": root}
    if release is not None:
        kwargs["release"] = release
    return AbsBaLgaDataSource(**kwargs)  # type: ignore[arg-type]


def _register() -> None:
    from . import registry  # noqa: PLC0415

    registry.register_fetcher("abs_building_approvals_lga", _build_fetcher)


_register()
