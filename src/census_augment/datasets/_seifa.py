"""SEIFA fetcher — releases 2016 and 2021 (dataset id ``seifa``).

ABS publishes the Socio-Economic Indexes for Areas as a single workbook
per geographic level. This fetcher supports two Census releases:

- **2021** — SA2 .xlsx (~150 KB).  Direct-link download.
  Uses ASGS Edition 3 SA2 codes.
- **2016** — SA2 .xls (~700 KB).  Direct-link download.
  Uses ASGS Edition 2 SA2 codes. Parsed via python-calamine (Rust
  reader that handles both .xls and .xlsx without the legacy xlrd
  library).

Both workbooks share an identical sheet structure:

- **Contents** — table of contents (skipped).
- **Table 1** — summary view (Score + AusDecile for all four indexes).
- **Tables 2–5** — one per index (IRSD, IRSAD, IER, IEO), full flavour
  set (Score, Aus Rank/Decile/Percentile, State Rank/Decile/Percentile,
  SA1 min/max score, % URP without score).
- **Table 6** — Excluded SA2s (skipped).
- **Explanatory Notes** (skipped).

Both releases use the same column positions in Tables 2–5, confirmed
against the live ABS files on 2026-05-22.  The parser selects the
grid reader (openpyxl for .xlsx, CalamineWorkbook for .xls) based on
the release, then feeds the resulting ``dict[sheet_name → row grid]``
to the shared ``_parse_grids`` function.
"""

from __future__ import annotations

import logging
import re
from pathlib import Path

import pandas as pd
import requests

from ._xlsx_base import _AbsXlsxDataset

_log = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Release URLs
# ---------------------------------------------------------------------------

# Direct-link to the SA2 SEIFA 2021 workbook on the ABS site.
# Confirmed via the latest-release landing page on 2026-05-09.
DEFAULT_SEIFA_2021_URL = (
    "https://www.abs.gov.au/statistics/people/people-and-communities/"
    "socio-economic-indexes-areas-seifa-australia/2021/"
    "Statistical%20Area%20Level%202%2C%20Indexes%2C%20SEIFA%202021.xlsx"
)

# Direct-link to the SA2 SEIFA 2016 workbook on the ABS site.
# Confirmed via the 2016 release landing page on 2026-05-22.
DEFAULT_SEIFA_2016_URL = (
    "https://www.abs.gov.au/AUSSTATS/subscriber.nsf/log?"
    "openagent&2033055001%20-%20sa2%20indexes.xls&2033.0.55.001"
    "&Data%20Cubes&C9F7AD36397CB43DCA25825D000F917C&0&2016&27.03.2018&Latest"
)

_RELEASE_URLS: dict[str, str] = {
    "2016": DEFAULT_SEIFA_2016_URL,
    "2021": DEFAULT_SEIFA_2021_URL,
}

_SUPPORTED_RELEASES: frozenset[str] = frozenset(_RELEASE_URLS)

# ---------------------------------------------------------------------------
# Sheet / column layout — shared across 2016 and 2021
# ---------------------------------------------------------------------------

# Sheet → which index it details.  Position-indexed because sheet names
# are generic ("Table 2" / "Table 3" / ...).  Order is per ABS convention,
# confirmed against both the 2016 and 2021 releases.
_INDEX_SHEETS: list[tuple[str, str]] = [
    ("Table 2", "irsd"),
    ("Table 3", "irsad"),
    ("Table 4", "ier"),
    ("Table 5", "ieo"),
]

# Tables 2–5 share a fixed-position layout.  Column index → output field.
# Columns 4 and 8 are blank spacers in the source workbook.
_INDEX_TABLE_COLS: dict[str, int] = {
    "sa2_code": 0,
    "sa2_name": 1,
    "urp": 2,
    "score": 3,
    # col 4: blank spacer
    "aus_rank": 5,
    "aus_decile": 6,
    "aus_percentile": 7,
    # col 8: blank spacer
    "state_abbreviation": 9,
    "state_rank": 10,
    "state_decile": 11,
    "state_percentile": 12,
    "sa1_min": 13,
    "sa1_max": 14,
    "pct_urp_no_score": 15,
}

# Header row detection — scan the first N rows for this text.
_DEFAULT_HEADER_ROW = 5
_SA2_CODE_HEADER_FRAGMENTS = (
    "Statistical Area Level 2 (SA2) 9-Digit",
    "SA2 9-Digit",
    "SA2 9 Digit",
)


# ---------------------------------------------------------------------------
# Grid readers (format-agnostic)
# ---------------------------------------------------------------------------


def _read_grids(path: Path, release: str) -> dict[str, list[list[object]]]:
    """Read a SEIFA workbook into a sheet-name → row-grid dict.

    Selects the reader based on release:

    - ``"2021"`` — openpyxl (native .xlsx reader; already a project dep).
    - ``"2016"`` — python-calamine (Rust reader; handles .xls without xlrd).

    Returns a dict mapping each sheet name to a list of rows, where each
    row is a list of cell values (``None`` for empty/blank cells).
    """
    if release == "2021":
        import openpyxl  # noqa: PLC0415 — lazy import keeps cold start cheap

        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        try:
            return {
                name: [list(row) for row in wb[name].iter_rows(values_only=True)]
                for name in wb.sheetnames
            }
        finally:
            wb.close()
    else:  # "2016" — legacy .xls via python-calamine
        from python_calamine import CalamineWorkbook  # noqa: PLC0415

        wb = CalamineWorkbook.from_path(str(path))
        return {
            name: wb.get_sheet_by_name(name).to_python()
            for name in wb.sheet_names
        }


# ---------------------------------------------------------------------------
# Grid parsers (pure; testable without file I/O)
# ---------------------------------------------------------------------------


def _parse_grids(
    grids: dict[str, list[list[object]]],
    *,
    sa2_index_name: str,
) -> pd.DataFrame:
    """Parse SEIFA index sheets from a pre-read grid dict.

    ``grids`` maps sheet-name → row-list-of-lists as returned by
    :func:`_read_grids`.  ``sa2_index_name`` is the DataFrame index
    name (``"sa2_code_2021"`` or ``"sa2_code_2016"``).

    Returns a DataFrame indexed by ``sa2_index_name``.
    """
    per_index_dfs: list[pd.DataFrame] = []
    for sheet_name, prefix in _INDEX_SHEETS:
        if sheet_name not in grids:
            _log.warning(
                "Expected sheet %r missing from workbook; skipping %s",
                sheet_name,
                prefix,
            )
            continue
        df_one = _parse_index_sheet(
            grids[sheet_name], prefix, sa2_index_name=sa2_index_name
        )
        per_index_dfs.append(df_one)

    if not per_index_dfs:
        raise RuntimeError("No SEIFA index sheets (Tables 2–5) found in workbook")

    # Outer-join all four index frames on the SA2 code.  First frame
    # contributes URP and state_abbreviation (shared across all indexes).
    merged = per_index_dfs[0]
    for df_more in per_index_dfs[1:]:
        df_more = df_more.drop(
            columns=[
                c for c in ("urp", "state_abbreviation") if c in df_more.columns
            ],
            errors="ignore",
        )
        merged = merged.join(df_more, how="outer")

    return merged


def _parse_index_sheet(
    raw: list[list[object]],
    prefix: str,
    *,
    sa2_index_name: str,
) -> pd.DataFrame:
    """Parse one SEIFA index detail sheet from its raw row grid.

    ``raw`` is a list of rows (each row a list of cell values, ``None``
    for blanks) as produced by :func:`_read_grids`.

    Returns a DataFrame indexed by ``sa2_index_name`` with columns
    named ``{prefix}_{flavour}`` plus bare ``urp`` and
    ``state_abbreviation``.
    """
    header_row_idx = _find_header_row(raw)
    data_start = header_row_idx + 1

    records: list[dict[str, object]] = []
    for row in raw[data_start:]:
        if len(row) <= _INDEX_TABLE_COLS["sa2_code"]:
            continue
        sa2_raw = row[_INDEX_TABLE_COLS["sa2_code"]]
        sa2_code = _normalise_sa2_code(sa2_raw)
        if not (len(sa2_code) == 9 and sa2_code.isdigit()):
            continue  # skip aggregates / footers / blanks

        rec: dict[str, object] = {sa2_index_name: sa2_code}
        for field, col_idx in _INDEX_TABLE_COLS.items():
            if field == "sa2_code":
                continue
            if col_idx >= len(row):
                rec[_field_name(prefix, field)] = None
                continue
            value = _coerce(row[col_idx])
            rec[_field_name(prefix, field)] = value
        records.append(rec)

    if not records:
        raise RuntimeError(
            f"No data rows found below the header for prefix {prefix!r}"
        )

    df = pd.DataFrame.from_records(records)
    return df.set_index(sa2_index_name)


def _find_header_row(raw: list[list[object]]) -> int:
    """Locate the header row by scanning for SA2-code header text.

    Returns the row index, falling back to ``_DEFAULT_HEADER_ROW`` when
    the scan misses (some releases shift the preamble length by a row
    or two).
    """
    for i in range(min(15, len(raw))):
        row_text = " ".join("" if c is None else str(c) for c in raw[i])
        for fragment in _SA2_CODE_HEADER_FRAGMENTS:
            if fragment in row_text:
                return i
    return _DEFAULT_HEADER_ROW


def _normalise_sa2_code(raw: object) -> str:
    """Coerce a raw SA2 code cell to a plain digit string.

    Excel sometimes reads integer codes as floats (e.g. ``101021007.0``);
    we strip the ``.0`` suffix so the 9-digit filter works correctly.
    """
    if raw is None:
        return ""
    if isinstance(raw, float) and raw == int(raw):
        return str(int(raw))
    s = str(raw).strip()
    # Handle "101021007.0" artefacts from mixed-type reads.
    if re.fullmatch(r"\d+\.0", s):
        s = s[:-2]
    return s


def _coerce(cell: object) -> object:
    """Coerce a raw cell value to the right Python type.

    - ``None`` / empty / ``'np'`` / ``'NA'`` / ``'-'`` / similar → ``None``
    - Numeric (int or float) values stay as-is.
    - Otherwise convert to ``str``.

    ABS uses different null sentinels per release — ``np`` (not
    published) in 2016, ``NA`` in 2021, sometimes ``-`` or blank.
    """
    if cell is None:
        return None
    if isinstance(cell, (int, float)):
        return cell
    s = str(cell).strip()
    if not s or s.lower() in ("np", "na", "n/a", "-", "nan", "null", "..", "."):
        return None
    # Integer?
    if re.fullmatch(r"-?\d+", s):
        return int(s)
    # Float?
    if re.fullmatch(r"-?\d+\.\d+", s):
        return float(s)
    return s


def _field_name(prefix: str, field: str) -> str:
    """Apply the index prefix to per-flavour fields, leaving shared
    fields (``urp``, ``state_abbreviation``) bare.
    """
    if field in ("urp", "state_abbreviation"):
        return field
    return f"{prefix}_{field}"


# ---------------------------------------------------------------------------
# DatasetFetcher implementation
# ---------------------------------------------------------------------------


class SeifaDataSource(_AbsXlsxDataset):
    """Fetch + load SEIFA SA2 workbooks for releases 2016 and 2021.

    Implements the :class:`DatasetFetcher` Protocol via the shared
    :class:`_AbsXlsxDataset` base.

    Files land at ``<root>/seifa-{release}.{ext}``; parsed parquet at
    ``<root>/seifa-{release}.parquet`` (so subsequent ``load()`` calls
    skip the workbook parse).

    The SA2 index column is named ``sa2_code_{release}`` to reflect the
    ASGS edition:

    - ``seifa-2016`` → ``sa2_code_2016`` (ASGS Edition 2)
    - ``seifa-2021`` → ``sa2_code_2021`` (ASGS Edition 3)
    """

    _label = "SEIFA SA2 workbook"
    _cache_glob = "seifa-*.xlsx"

    def __init__(
        self,
        *,
        release: str = "2021",
        root: Path,
        url: str | None = None,
        session: requests.Session | None = None,
        chunk_size: int = 256 * 1024,
        timeout: float = 60.0,
    ) -> None:
        if release not in _SUPPORTED_RELEASES:
            raise ValueError(
                f"release must be one of {sorted(_SUPPORTED_RELEASES)}; "
                f"got {release!r}"
            )
        super().__init__(
            release=release,
            root=root,
            session=session,
            chunk_size=chunk_size,
            timeout=timeout,
        )
        # SEIFA has static URLs for all supported releases — populate
        # resolved attrs eagerly so _resolve_release() is a no-op.
        self._resolved_release = release
        self._resolved_url = url or _RELEASE_URLS[release]
        # SA2 index column name reflects the ASGS edition for this release.
        self._sa2_index_name = f"sa2_code_{release}"

    # ---- hooks -------------------------------------------------------------

    def _filename_stem(self, release: str) -> str:
        return f"seifa-{release}"

    @property
    def _xlsx_path(self) -> Path:
        """File path for the cached workbook.

        The 2016 release is a legacy .xls file; 2021 is .xlsx.
        """
        ext = ".xls" if self._resolved_release == "2016" else ".xlsx"
        return self._root / f"{self._filename_stem(self.resolved_release)}{ext}"

    def _resolve_release(self) -> None:
        # No-op: __init__ populated both resolved attrs eagerly.
        return

    # ---- parsing -----------------------------------------------------------

    def _parse_xlsx(self, xlsx_path: Path) -> pd.DataFrame:
        """Parse the SEIFA workbook into a single wide DataFrame.

        Delegates grid reading to :func:`_read_grids` (format-agnostic)
        and grid parsing to :func:`_parse_grids` (pure, testable).
        """
        grids = _read_grids(xlsx_path, self._resolved_release or "2021")
        return _parse_grids(grids, sa2_index_name=self._sa2_index_name)


# ---------------------------------------------------------------------------
# Fetcher registration
# ---------------------------------------------------------------------------
#
# Bind the SEIFA fetcher to the "seifa" dataset id on the process-wide
# registry.  The pipeline calls ``registry.make_fetcher("seifa", root=…)``
# to construct one during enrichment.


def _build_fetcher(root: Path, release: str | None = None) -> SeifaDataSource:
    kwargs: dict[str, object] = {"root": root}
    if release is not None:
        kwargs["release"] = release
    return SeifaDataSource(**kwargs)  # type: ignore[arg-type]


def _register() -> None:
    # Late import to dodge the circular: `datasets/__init__.py` imports
    # us, and we'd otherwise import it back at module top.
    from . import registry  # noqa: PLC0415

    registry.register_fetcher("seifa", _build_fetcher)


_register()
