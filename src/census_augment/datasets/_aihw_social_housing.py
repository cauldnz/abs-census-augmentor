"""AIHW Housing Assistance — Social housing dwellings fetcher (spec §20,
dataset id ``aihw_social_housing``).

AIHW publishes "Housing Assistance in Australia" with a data-tables XLSX
workbook; sheet **DWELLINGS.4** gives social-housing dwelling counts at
**SA4** level by program (public housing, SOMIH, community housing).
SA4-native, downscaled to SA2 via the boundary's ``SA4_CODE21``
attribute — the same cross-level inheritance the AIHW NMHSPF datasets
use, but this source is a single XLSX wide table rather than a ZIP+CSV
long-format cube, so it does NOT subclass ``AihwSa4Dataset``; it reuses
only the SA2→SA4 downscale contract (``attach_sa2_to_sa4_mapping``), which
``Pipeline.from_config`` / the enricher wire automatically by capability.

Real-data findings (live-probed 2026-06-10):

- Workbook ``AIHW-337-Data-tables-Social-housing-dwellings.xlsx``; the
  SA4 table is sheet ``DWELLINGS.4`` (title names "Statistical level 4
  (SA4) ... 2023"). Banner rows 1-2, header at row 4, data from row 5.
- Columns: ``State/territory, Region Code, Region Name, Public housing,
  SOMIH(a), Community housing, Total``. ``Region Code`` is the bare
  3-digit SA4 code (``101`` … ``801``); 88 SA4 rows.
- The ``SOMIH`` column uses the suppression sentinel ``". ."`` for the
  states without a SOMIH program (Vic / WA / ACT) — parsed to null.
- Footnote rows after the data have a blank Region Code, so a strict
  3-digit-code filter drops them.

URL uses an opaque AIHW ``getmedia`` UUID + a series number (e.g. 337)
that change per release — hardcoded per release (no HTML scrape); a new
release needs a new entry in ``_AIHW_SH_URLS_BY_RELEASE``.
"""

from __future__ import annotations

import logging
import re
from pathlib import Path

import pandas as pd
import requests

from .._http_retry import retry_stream_get

_log = logging.getLogger(__name__)

# Reference year -> (workbook url, sheet, title marker). Opaque getmedia
# UUID + series number change per release; add a new entry per release.
_AIHW_SH_URLS_BY_RELEASE: dict[str, dict[str, str]] = {
    "2023": {
        "url": (
            "https://www.aihw.gov.au/getmedia/"
            "47ce0fe9-8706-4991-9fa9-1b0770971ef8/"
            "AIHW-337-Data-tables-Social-housing-dwellings.xlsx"
        ),
        "sheet": "DWELLINGS.4",
        "title_marker": "2023",
    },
}

# Value-column header prefix (case-insensitive) -> output column, in the
# DWELLINGS.4 column order (cols D-G / 0-indexed 3-6).
_VALUE_HEADER_PREFIXES: list[str] = [
    "public housing",
    "somih",
    "community housing",
    "total",
]
_OUTPUT_COLUMNS: list[str] = [
    "social_housing_public_count",
    "social_housing_somih_count",
    "social_housing_community_count",
    "social_housing_total_count",
]
# Counts coerced to nullable Int64.
_COUNT_COLUMNS: tuple[str, ...] = tuple(_OUTPUT_COLUMNS)
_SA4_CODE_COL = 1
_FIRST_VALUE_COL = 3


class AihwSocialHousingDataSource:
    """Fetch + load AIHW social-housing dwelling counts at SA4 → SA2.

    Implements the :class:`DatasetFetcher` Protocol. SA4-native data
    downscaled to SA2 via a boundary-derived ``SA2 -> SA4`` mapping that
    callers attach before ``load()``.

    Args:
        release: Reference year (``"2023"``) or ``"latest"``.
        root: Cache directory for the workbook + parquet sidecar.
        session: Optional ``requests.Session`` (tests pass a hermetic one).
        chunk_size / timeout: As per the other ABS/AIHW fetchers.
    """

    def __init__(
        self,
        *,
        release: str = "latest",
        root: Path,
        session: requests.Session | None = None,
        chunk_size: int = 256 * 1024,
        timeout: float = 60.0,
    ) -> None:
        self._release_request = str(release)
        self._root = Path(root)
        self._session = session if session is not None else requests.Session()
        self._chunk_size = chunk_size
        self._timeout = timeout
        self._resolved_release: str | None = None
        self._resolved_url: str | None = None
        self._resolved_sheet: str | None = None
        self._resolved_marker: str | None = None
        self._sa2_to_sa4: dict[str, str] | None = None

    # ---- mapping attachment ------------------------------------------

    def attach_sa2_to_sa4_mapping(self, mapping: dict[str, str]) -> None:
        """Attach the boundary-derived ``{sa2_code: sa4_code}`` lookup.

        SA4 codes are the bare 3-digit form (``"101"``) on both sides.
        ``Pipeline.from_config`` wires this from
        ``compute_sa2_parent_codes(boundaries)["SA4"]``.
        """
        if not isinstance(mapping, dict):
            raise TypeError(
                f"attach_sa2_to_sa4_mapping expects a dict[str, str]; got {type(mapping).__name__}"
            )
        self._sa2_to_sa4 = dict(mapping)
        _log.debug(
            "AihwSocialHousingDataSource: attached %d SA2 -> SA4 mappings",
            len(self._sa2_to_sa4),
        )

    # ---- DatasetFetcher protocol --------------------------------------

    @property
    def resolved_release(self) -> str:
        if self._resolved_release is None:
            self._resolve_release()
        assert self._resolved_release is not None
        return self._resolved_release

    @property
    def is_cached(self) -> bool:
        if self._resolved_release is not None:
            return self._xlsx_path.exists()
        return self._root.exists() and any(self._root.glob("aihw-social-housing-*.xlsx"))

    @property
    def _xlsx_path(self) -> Path:
        return self._root / f"aihw-social-housing-{self.resolved_release}.xlsx"

    @property
    def _parquet_path(self) -> Path:
        return self._root / f"aihw-social-housing-{self.resolved_release}.parquet"

    def fetch(self, refresh: bool = False) -> Path:
        """Download the AIHW social-housing workbook for the release."""
        self._resolve_release()
        if self._xlsx_path.exists() and not refresh:
            _log.debug("AIHW Social Housing cached at %s", self._xlsx_path)
            return self._xlsx_path

        self._root.mkdir(parents=True, exist_ok=True)
        tmp = self._xlsx_path.with_suffix(self._xlsx_path.suffix + ".tmp")
        url = self._resolved_url or ""
        _log.info("Downloading AIHW Social Housing (%s) from %s", self.resolved_release, url)
        with retry_stream_get(
            self._session,
            url,
            timeout=self._timeout,
            label="AIHW Social Housing",
        ) as response:
            response.raise_for_status()
            with tmp.open("wb") as f:
                for chunk in response.iter_content(chunk_size=self._chunk_size):
                    if chunk:
                        f.write(chunk)
        tmp.replace(self._xlsx_path)
        _log.info("Saved AIHW Social Housing to %s", self._xlsx_path)
        return self._xlsx_path

    def load(self) -> pd.DataFrame:
        """Return a DataFrame indexed by ``sa2_code_2021`` with the
        per-SA2 social-housing dwelling counts (downscaled from SA4).
        """
        if self._sa2_to_sa4 is None:
            raise RuntimeError(
                "AihwSocialHousingDataSource.load() requires a SA2 -> SA4 "
                "mapping to be attached first. Call "
                "`attach_sa2_to_sa4_mapping(mapping)` with the lookup dict "
                "from `census_augment.spatial.compute_sa2_parent_codes("
                "boundaries)['SA4']`. Pipeline.from_config wires this "
                "automatically from the boundary GDF."
            )

        if self._resolved_release is not None and self._parquet_path.exists():
            df = pd.read_parquet(self._parquet_path)
            if "sa2_code_2021" in df.columns:
                return df.set_index("sa2_code_2021")

        xlsx_path = self.fetch()
        sa4_df = self._parse_workbook(
            xlsx_path,
            sheet=self._resolved_sheet or "",
            title_marker=self._resolved_marker or "",
        )

        # Cross-level downscale: every SA2 inherits its SA4's row values.
        records: list[dict[str, object]] = []
        for sa2_code, sa4_code in self._sa2_to_sa4.items():
            if sa4_code not in sa4_df.index:
                rec: dict[str, object] = {
                    "sa2_code_2021": str(sa2_code),
                    **{col: None for col in _OUTPUT_COLUMNS},
                }
            else:
                row = sa4_df.loc[sa4_code]
                rec = {
                    "sa2_code_2021": str(sa2_code),
                    **{col: row[col] for col in _OUTPUT_COLUMNS},
                }
            rec["reference_period"] = self.resolved_release
            records.append(rec)

        out = pd.DataFrame.from_records(records)
        out.to_parquet(self._parquet_path, index=False)
        return out.set_index("sa2_code_2021")

    # ---- release resolution -------------------------------------------

    def _resolve_release(self) -> None:
        if self._resolved_release is not None:
            return

        if self._release_request == "latest":
            picked = max(_AIHW_SH_URLS_BY_RELEASE)
        elif self._release_request in _AIHW_SH_URLS_BY_RELEASE:
            picked = self._release_request
        else:
            raise RuntimeError(
                f"AIHW Social Housing release {self._release_request!r} not in "
                f"the registry. Available: {sorted(_AIHW_SH_URLS_BY_RELEASE)}. "
                f"AIHW uses opaque getmedia UUIDs; new releases need an entry "
                f"in _AIHW_SH_URLS_BY_RELEASE in "
                f"src/census_augment/datasets/_aihw_social_housing.py."
            )
        entry = _AIHW_SH_URLS_BY_RELEASE[picked]
        self._resolved_release = picked
        self._resolved_url = entry["url"]
        self._resolved_sheet = entry["sheet"]
        self._resolved_marker = entry["title_marker"]
        _log.info(
            "Resolved AIHW Social Housing release=%s, sheet=%s",
            picked,
            entry["sheet"],
        )

    # ---- parsing -------------------------------------------------------

    @staticmethod
    def _parse_workbook(xlsx_path: Path, *, sheet: str, title_marker: str) -> pd.DataFrame:
        """Parse DWELLINGS.4 into a DataFrame indexed by bare SA4 code,
        one column per social-housing program.
        """
        import openpyxl  # noqa: PLC0415

        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
        if sheet not in wb.sheetnames:
            available = wb.sheetnames
            wb.close()
            raise RuntimeError(
                f"AIHW Social Housing workbook {xlsx_path} has no {sheet!r} sheet. "
                f"Sheets: {available}. The release→sheet map may be stale."
            )
        ws = wb[sheet]
        rows: list[list[object]] = [list(r) for r in ws.iter_rows(values_only=True)]
        wb.close()

        # Drift guard: the title row must name SA4 + the expected year.
        title_line: str | None = None
        for row in rows[:6]:
            joined = " ".join("" if c is None else str(c) for c in row)
            if "sa4" in joined.lower():
                title_line = joined
                break
        if title_line is None:
            raise RuntimeError(
                f"AIHW Social Housing sheet {sheet!r} in {xlsx_path} has no title "
                f"naming 'SA4' in its first 6 rows — layout may have changed."
            )
        if title_marker and title_marker not in title_line:
            raise RuntimeError(
                f"AIHW Social Housing sheet {sheet!r} title does not name "
                f"{title_marker!r} — the release→sheet map may be stale. Title: "
                f"{title_line[:160]!r}"
            )

        # Find the header row: a 'Region Code' cell in the SA4-code column.
        header_idx = -1
        for i in range(min(12, len(rows))):
            row = rows[i]
            if len(row) <= _FIRST_VALUE_COL:
                continue
            code_hdr = str(row[_SA4_CODE_COL]).strip().lower() if row[_SA4_CODE_COL] else ""
            if code_hdr == "region code":
                header_idx = i
                break
        if header_idx < 0:
            raise RuntimeError(
                f"Could not find the 'Region Code' header row in AIHW Social "
                f"Housing sheet {sheet!r} of {xlsx_path}."
            )

        # Validate the value-column headers against the expected order so a
        # column reshuffle fails loud rather than mislabelling silently.
        header_row = rows[header_idx]
        for offset, prefix in enumerate(_VALUE_HEADER_PREFIXES):
            col = _FIRST_VALUE_COL + offset
            cell = str(header_row[col]).strip().lower() if col < len(header_row) else ""
            if prefix not in cell:
                raise RuntimeError(
                    f"AIHW Social Housing value header col {col} is {cell!r}, "
                    f"expected to contain {prefix!r} (-> {_OUTPUT_COLUMNS[offset]!r}). "
                    f"Layout may have shifted in {xlsx_path}."
                )

        value_cols = list(range(_FIRST_VALUE_COL, _FIRST_VALUE_COL + len(_OUTPUT_COLUMNS)))
        records: list[dict[str, object]] = []
        for row in rows[header_idx + 1 :]:
            if len(row) <= _SA4_CODE_COL:
                continue
            code_raw = row[_SA4_CODE_COL]
            code = "" if code_raw is None else str(code_raw).strip()
            if not re.fullmatch(r"\d{3}", code):
                continue
            rec: dict[str, object] = {"sa4_code": code}
            for output_name, col in zip(_OUTPUT_COLUMNS, value_cols, strict=True):
                rec[output_name] = _coerce_count(row[col] if col < len(row) else None)
            records.append(rec)

        if not records:
            raise RuntimeError(
                f"AIHW Social Housing sheet {sheet!r} in {xlsx_path} produced no "
                f"3-digit SA4 rows. The layout may have changed — re-probe."
            )

        df = pd.DataFrame.from_records(records).set_index("sa4_code")
        for count_col in _COUNT_COLUMNS:
            df[count_col] = pd.to_numeric(df[count_col], errors="coerce").astype("Int64")
        return df


def _coerce_count(cell: object) -> object:
    """AIHW counts are integers; treat the ``". ."`` / ``n.a.`` sentinels
    (and blanks) as None."""
    if cell is None:
        return None
    if isinstance(cell, bool):
        return int(cell)
    if isinstance(cell, (int, float)):
        return cell
    s = str(cell).strip()
    if not s or s.lower() in (". .", "..", "n.a.", "na", "n/a", "np", "-", ".", "nan", "null"):
        return None
    s_nospace = s.replace(" ", "")
    if re.fullmatch(r"-?\d+", s_nospace):
        return int(s_nospace)
    if re.fullmatch(r"-?\d+\.\d+", s_nospace):
        return float(s_nospace)
    return None


# ---- fetcher registration ------------------------------------------------


def _build_fetcher(root: Path, release: str | None = None) -> AihwSocialHousingDataSource:
    kwargs: dict[str, object] = {"root": root}
    if release is not None:
        kwargs["release"] = release
    return AihwSocialHousingDataSource(**kwargs)  # type: ignore[arg-type]


def _register() -> None:
    from . import registry  # noqa: PLC0415

    registry.register_fetcher("aihw_social_housing", _build_fetcher)


_register()
