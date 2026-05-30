"""ABS Estimated Resident Population by SA2 fetcher (spec §20, dataset ``erp_by_sa2``).

ABS publishes annual ERP estimates as two complementary multi-sheet
XLSX workbooks under different catalogues:

- **3218.0 — Regional Population.**  The long-history total-population
  workbook (``32180DS0003_*.xlsx``); carries the full 2001-onwards series
  on the current ASGS edition. The augmentor uses Table 1 for SA2 totals.
- **3235.0 — Regional Population by Age and Sex.**  Median age, sex
  ratio, broad age groups (0-14 / 15-64 / 65+), and male/female totals,
  for the latest reference year only (``32350DS0002_*.xlsx``). Added in
  the ERP-wishlist PR — unblocks cross-dataset PRESETs like
  ``pct_age_pension_recipients = DSS.age_pension_recipients /
  ERP.population_65_plus``.

The download URLs are constructed from two distinct ABS landing pages:

- 3218.0 (``ERP_LANDING_URL``) uses financial-year paths like
  ``/regional-population/2024-25/``.
- 3235.0 (``ERP_AGE_SEX_LANDING_URL``) uses calendar-year paths like
  ``/regional-population-age-and-sex/2024/``.

Both are scraped once each on ``fetch()``; URLs are then cached on
the source instance.  When ``release="latest"`` the per-page latest
is picked independently — they usually publish together but the
fetcher tolerates a one-release skew (the resolved age/sex release is
recorded but not enforced against DS0003's release).
"""

from __future__ import annotations

import logging
import re
from pathlib import Path

import pandas as pd
import requests

from .._http_retry import retry_stream_get
from ._xlsx_base import _AbsXlsxDataset

_log = logging.getLogger(__name__)

ERP_LANDING_URL = (
    "https://www.abs.gov.au/statistics/people/population/regional-population/latest-release"
)
ERP_AGE_SEX_LANDING_URL = (
    "https://www.abs.gov.au/statistics/people/population/"
    "regional-population-age-and-sex/latest-release"
)

# DS-product number on the ABS download page; this is the long-history
# SA2 ERP file. The same page also has DS0005 (2021-onwards SA2) and
# higher levels (SA3/SA4/etc.); we want the long-history SA2 version.
_LONG_HISTORY_SA2_FRAGMENT = "32180DS0003"

# DS-product number for the 3235.0 age-and-sex SA2 cube. Table 1 of this
# workbook is wide-format: one row per SA2 with Males/Females/Persons
# counts, sex ratio, median age, and the three broad age-group
# percentages (0-14, 15-64, 65+). We multiply the percentages by the
# Persons total to recover counts.
_AGE_SEX_SA2_FRAGMENT = "32350DS0002"

# Column positions in Table 1 of DS0003 (zero-indexed). The header
# is split across two rows in current releases:
#   * Row 4 has the year integers in cols 10+ (cols 0-9 blank).
#   * Row 5 has the geography labels in cols 0-9 and "no." (units)
#     in cols 10+.
# Data starts at row 6. We scan the first ~10 rows for the row that
# contains year integers in the year-position range, so a small layout
# shift between releases doesn't break us.
_TABLE1_SA2_CODE_COL = 8
_TABLE1_SA2_NAME_COL = 9
_TABLE1_FIRST_YEAR_COL = 10

# Column positions in Table 1 of DS0002 (zero-indexed). Header at row 5
# (geography labels) + row 4 (column-group labels). Data starts at row 6.
# Cells are: 0-9 geography (same as DS0003), then 10..17 values.
_DS0002_SA2_CODE_COL = 8
_DS0002_MALES_COL = 10
_DS0002_FEMALES_COL = 11
_DS0002_PERSONS_COL = 12
_DS0002_SEX_RATIO_COL = 13
_DS0002_MEDIAN_AGE_COL = 14
_DS0002_PCT_0_14_COL = 15
_DS0002_PCT_15_64_COL = 16
_DS0002_PCT_65_PLUS_COL = 17


class ErpDataSource(_AbsXlsxDataset):
    """Fetch + load ABS Regional Population (SA2 long-history series).

    Implements the :class:`DatasetFetcher` Protocol. ``release`` is the
    *reference year* (e.g. ``2024`` for the 2024-25 release covering
    populations through 30 June 2024). ``"latest"`` resolves the
    landing page's most recent release.
    """

    _label = "ABS Regional Population SA2 (long history)"
    _cache_glob = "erp-*.xlsx"

    def __init__(
        self,
        *,
        release: str | int = "latest",
        root: Path,
        landing_url: str = ERP_LANDING_URL,
        age_sex_landing_url: str = ERP_AGE_SEX_LANDING_URL,
        session: requests.Session | None = None,
        chunk_size: int = 256 * 1024,
        timeout: float = 60.0,
    ) -> None:
        super().__init__(
            release=str(release),
            root=root,
            session=session,
            chunk_size=chunk_size,
            timeout=timeout,
        )
        self._landing_url = landing_url
        self._age_sex_landing_url = age_sex_landing_url
        # Resolved lazily on first parse() — separate from the DS0003
        # release because the two ABS products publish on slightly
        # different cadences. Tracked so consumers can audit the
        # actual age/sex release year if they care.
        self._resolved_age_sex_release: str | None = None
        self._resolved_age_sex_url: str | None = None
        # Issue #92 fix: ERP is structurally different from SEIFA / GCP /
        # DSS. ABS publishes ONE annual workbook per cycle that contains
        # the full 2001-onwards history in ``population_history_<year>``
        # columns. Temporal mode resolves releases per row date, so we
        # accept any historical year as a logical release and project
        # the right history column into ``population_total`` at load()
        # time. ``_physical_release_year`` tracks the actual workbook
        # year (latest); ``_resolved_release`` tracks the logical
        # (possibly historical) year the caller asked for.
        self._physical_release_year: str | None = None
        # Population-density support: optional SA2-code -> area-km² lookup.
        # When attached (via ``attach_sa2_areas``), ``load()`` emits the
        # ``population_density_per_km2`` column = population_total / area.
        # When ``None`` the density column is omitted — keeps the fetcher
        # usable without a boundary dependency for callers who don't need
        # density (the cross-sectional notebook path, tests, etc.).
        self._sa2_areas_km2: dict[str, float] | None = None

    # ---- hooks ---------------------------------------------------------

    def _filename_stem(self, release: str) -> str:
        return f"erp-sa2-{release}"

    def _resolve_release(self) -> None:
        if self._resolved_release is not None:
            return

        # Fetch the landing page and pull the latest period directory
        # from any DS0003 link.
        _log.debug("Resolving ERP release via %s", self._landing_url)
        resp = self._session.get(self._landing_url, timeout=self._timeout)
        resp.raise_for_status()
        html = resp.text

        # Match links like /regional-population/2024-25/32180DS0003_2001-25.xlsx
        pattern = re.compile(
            r'href="([^"]*regional-population/'
            r"(?P<period>\d{4}-\d{2,4})/"
            r'[^"]*' + _LONG_HISTORY_SA2_FRAGMENT + r'[^"]*\.xlsx)"',
            re.IGNORECASE,
        )

        all_matches = pattern.finditer(html)
        candidates: list[tuple[str, str]] = []
        for m in all_matches:
            href = m.group(1)
            period = m.group("period")
            release_year = period.split("-", 1)[0]
            url = (
                href
                if href.startswith("http")
                else ("https://www.abs.gov.au" + (href if href.startswith("/") else "/" + href))
            )
            candidates.append((release_year, url))

        if not candidates:
            raise RuntimeError(
                f"Could not find any {_LONG_HISTORY_SA2_FRAGMENT} link on "
                f"{self._landing_url}. Page layout may have changed."
            )

        # Always pick the latest workbook — it's the only one ABS reliably
        # hosts and its data goes back to 2001 via the
        # ``population_history_<year>`` columns. Historical years
        # requested by the temporal-mode resolver project from those
        # columns at load() time (issue #92 fix).
        latest_year, latest_url = max(candidates, key=lambda t: t[0])
        self._physical_release_year = latest_year
        self._resolved_url = latest_url

        if self._release_request == "latest":
            self._resolved_release = latest_year
        elif self._release_request.isdigit():
            requested_year = int(self._release_request)
            latest_year_int = int(latest_year)
            if requested_year > latest_year_int:
                raise RuntimeError(
                    f"ERP release {self._release_request!r} is more recent "
                    f"than the latest published workbook ({latest_year}). "
                    f"Available historical range: 2001..{latest_year}."
                )
            # Lower bound (2001) is validated at load() time against the
            # actual ``population_history_*`` columns the workbook
            # contains — keeps this method's only network round-trip the
            # landing-page scrape.
            self._resolved_release = self._release_request
        else:
            raise RuntimeError(f"ERP release {self._release_request!r} is not a year or 'latest'.")

        _log.info(
            "Resolved ERP release=%s (physical workbook: %s), url=%s",
            self._resolved_release,
            self._physical_release_year,
            self._resolved_url,
        )

    # ---- optional SA2-area attachment (population_density_per_km2) -----

    def attach_sa2_areas(self, areas: dict[str, float]) -> None:
        """Attach an SA2-code → area-km² lookup so ``load()`` emits
        the ``population_density_per_km2`` column.

        The caller computes the lookup once per pipeline run (typically
        from the boundary GeoDataFrame via
        :func:`census_augment.spatial.compute_sa2_areas_km2`) and
        attaches the same dict to every ErpDataSource instance that
        will produce density output.

        Safe to call repeatedly — last attachment wins. Safe to leave
        unattached: density column simply doesn't appear in the
        ``load()`` output.
        """
        self._sa2_areas_km2 = dict(areas)

    # ---- cache paths use the *physical* year (issue #92) ---------------

    @property
    def _xlsx_path(self) -> Path:
        """Cache the XLSX under the *physical* release year, not the
        logical release. Every historical release shares the same
        on-disk workbook — ABS only publishes one file per cycle that
        contains the full back-series.
        """
        # Force resolve so _physical_release_year is populated.
        self._resolve_release()
        assert self._physical_release_year is not None
        return self._root / f"erp-sa2-{self._physical_release_year}.xlsx"

    @property
    def _parquet_path(self) -> Path:
        """Same — one parsed-result cache per *physical* workbook, not
        per logical release. The projection in :meth:`load` picks the
        right historical column at read time.
        """
        self._resolve_release()
        assert self._physical_release_year is not None
        return self._root / f"erp-sa2-{self._physical_release_year}.parquet"

    # ---- load() with historical-year projection (issue #92) ------------

    def load(self) -> pd.DataFrame:
        """Return ERP data for the configured ``release``.

        For ``release="latest"`` (or when the resolved release matches
        the physical workbook year): returns the full data as parsed
        from the workbook — same shape as before this fix.

        For a historical release (e.g. ``release="2017"``): projects
        the ``population_history_<year>`` column into
        ``population_total`` and updates ``reference_year``.

        Age/sex columns (3235.0 DS0002) are sourced from the latest
        workbook regardless of release. For historical releases we
        null them out so users don't accidentally use current
        demographics for historical population totals. This is
        documented in ``datasets/erp_by_sa2.md`` and on issue #92.
        """
        df = super().load()
        # Defensive: super().load() returns a DataFrame indexed by
        # sa2_code_2021. _resolve_release has run by this point, so
        # both _resolved_release and _physical_release_year are
        # populated.
        assert self._resolved_release is not None
        assert self._physical_release_year is not None
        if self._resolved_release != self._physical_release_year:
            # Issue #92 projection: swap population_total for the
            # requested historical year and null the age/sex columns.
            requested = self._resolved_release
            hist_col = f"population_history_{requested}"
            if hist_col not in df.columns:
                available = sorted(
                    c.removeprefix("population_history_")
                    for c in df.columns
                    if c.startswith("population_history_")
                )
                raise RuntimeError(
                    f"ERP release {requested!r} is not in the workbook's "
                    f"historical coverage. Available years: {available}. "
                    f"This usually means the requested year predates ABS's "
                    f"published series start (typically 2001)."
                )
            df = df.copy()
            df["population_total"] = df[hist_col]
            df["reference_year"] = int(requested)
            # Age/sex columns reflect the latest workbook only — null
            # them for historical releases. See class docstring + #92.
            for col in (
                "population_male",
                "population_female",
                "median_age",
                "population_0_14",
                "population_15_64",
                "population_65_plus",
            ):
                if col in df.columns:
                    df[col] = None
        # Population density (independent of historical-projection above).
        # Emitted when ``attach_sa2_areas`` has been called. NaN for any
        # SA2 the lookup doesn't cover (no area available).
        if self._sa2_areas_km2 is not None:
            df = df.copy()
            areas = pd.Series(self._sa2_areas_km2, name="_sa2_area_km2")
            areas.index.name = df.index.name
            # Right-align so we only annotate rows where both inputs
            # exist; coerces missing SA2s to NaN density.
            joined = df.join(areas, how="left")
            density = (joined["population_total"] / joined["_sa2_area_km2"]).astype(float)
            # Inf shows up for zero-area SA2s (shouldn't happen with
            # real ABS boundaries but keeps the column clean if it does).
            density = density.where(density.abs() != float("inf"))
            df["population_density_per_km2"] = density
        return df

    # ---- parsing -------------------------------------------------------

    def _parse_xlsx(self, xlsx_path: Path) -> pd.DataFrame:
        """Parse the DS0003 totals workbook and merge in DS0002 age/sex columns.

        The merge is non-fatal: if the age/sex workbook can't be fetched
        or parsed (older releases when 3235.0 wasn't yet published, or a
        layout shift), we log a warning and emit just the DS0003
        columns. This keeps the long-history series usable in isolation.
        """
        df_totals = _parse_ds0003_workbook(xlsx_path)
        try:
            df_age_sex = self._fetch_and_parse_age_sex()
        except Exception as e:  # noqa: BLE001 — best-effort enrichment
            _log.warning(
                "ERP age/sex enrichment skipped (%s: %s). Output will not "
                "include population_male / population_female / "
                "population_0_14 / population_15_64 / population_65_plus / "
                "median_age columns.",
                type(e).__name__,
                e,
            )
            return df_totals
        # left-join: keep every SA2 the totals file has; age/sex columns
        # come in via the shared sa2_code_2021 index. Mismatches (DS0002
        # missing an SA2) result in NaN for the age/sex columns, which
        # parquet handles cleanly.
        return df_totals.join(df_age_sex, how="left")

    def _fetch_and_parse_age_sex(self) -> pd.DataFrame:
        """Resolve, download, and parse the 3235.0 DS0002 age/sex workbook.

        Returns a DataFrame indexed by ``sa2_code_2021`` with columns:

        - ``population_male`` (int)
        - ``population_female`` (int)
        - ``population_0_14`` (int — derived: persons × pct/100)
        - ``population_15_64`` (int)
        - ``population_65_plus`` (int)
        - ``median_age`` (float — years, one decimal)
        """
        self._resolve_age_sex_release()
        path = self._age_sex_xlsx_path
        if not path.exists():
            url = self._resolved_age_sex_url or ""
            tmp = path.with_suffix(path.suffix + ".tmp")
            _log.info(
                "Downloading %s (%s) from %s",
                "ABS Regional Population by Age and Sex SA2",
                self._resolved_age_sex_release,
                url,
            )
            with retry_stream_get(
                self._session,
                url,
                timeout=self._timeout,
                label="ABS Regional Population by Age and Sex SA2",
            ) as response:
                response.raise_for_status()
                with tmp.open("wb") as f:
                    for chunk in response.iter_content(chunk_size=self._chunk_size):
                        if chunk:
                            f.write(chunk)
            tmp.replace(path)
        return _parse_ds0002_workbook(path)

    def _resolve_age_sex_release(self) -> None:
        """Populate ``self._resolved_age_sex_release`` and matching URL."""
        if self._resolved_age_sex_release is not None:
            return

        _log.debug("Resolving ERP age/sex release via %s", self._age_sex_landing_url)
        resp = self._session.get(self._age_sex_landing_url, timeout=self._timeout)
        resp.raise_for_status()
        html = resp.text

        # Match links like /regional-population-age-and-sex/2024/32350DS0002_2024.xlsx
        pattern = re.compile(
            r'href="([^"]*regional-population-age-and-sex/'
            r"(?P<period>\d{4})/"
            r'[^"]*' + _AGE_SEX_SA2_FRAGMENT + r'[^"]*\.xlsx)"',
            re.IGNORECASE,
        )

        candidates: list[tuple[str, str]] = []
        for m in pattern.finditer(html):
            href = m.group(1)
            period = m.group("period")
            url = (
                href
                if href.startswith("http")
                else ("https://www.abs.gov.au" + (href if href.startswith("/") else "/" + href))
            )
            candidates.append((period, url))

        if not candidates:
            raise RuntimeError(
                f"Could not find any {_AGE_SEX_SA2_FRAGMENT} link on "
                f"{self._age_sex_landing_url}. Page layout may have changed."
            )

        picked = max(candidates, key=lambda t: t[0])
        self._resolved_age_sex_release = picked[0]
        self._resolved_age_sex_url = picked[1]
        _log.info(
            "Resolved ERP age/sex release=%s, url=%s",
            self._resolved_age_sex_release,
            self._resolved_age_sex_url,
        )

    @property
    def _age_sex_xlsx_path(self) -> Path:
        assert self._resolved_age_sex_release is not None
        return self._root / f"erp-age-sex-{self._resolved_age_sex_release}.xlsx"


def _parse_ds0003_workbook(xlsx_path: Path) -> pd.DataFrame:
    """Parse the 3218.0 DS0003 totals workbook (Table 1).

    Pure function so unit tests can exercise it without touching the
    network / disk via the fetcher class.
    """
    import openpyxl  # noqa: PLC0415

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    if "Table 1" not in wb.sheetnames:
        wb.close()
        raise RuntimeError(
            f"ERP workbook {xlsx_path} has no 'Table 1' sheet. Sheets: {wb.sheetnames}"
        )
    ws = wb["Table 1"]
    rows: list[list[object]] = []
    for r in ws.iter_rows(values_only=True):
        rows.append(list(r))
    wb.close()

    # Locate the year-header row. Cells at column index
    # _TABLE1_FIRST_YEAR_COL onwards should be year integers.
    # Sometimes the layout has years on the same row as the
    # geography labels; sometimes on the row above. Scan the
    # first ~10 rows.
    year_header_row_idx = -1
    year_cols: dict[int, int] = {}
    for ridx in range(min(10, len(rows))):
        r = rows[ridx]
        if len(r) <= _TABLE1_FIRST_YEAR_COL:
            continue
        candidate: dict[int, int] = {}
        for col_idx in range(_TABLE1_FIRST_YEAR_COL, len(r)):
            cell = r[col_idx]
            if isinstance(cell, int) and 1900 < cell < 2100:
                candidate[int(cell)] = col_idx
            elif isinstance(cell, str) and cell.strip().isdigit():
                yr = int(cell.strip())
                if 1900 < yr < 2100:
                    candidate[yr] = col_idx
        if candidate:
            year_header_row_idx = ridx
            year_cols = candidate
            break

    if not year_cols or year_header_row_idx < 0:
        raise RuntimeError(
            "No year columns found in any ERP header row. Inspected "
            f"first {min(10, len(rows))} rows."
        )

    # Data starts at the row *after* the geography-labels row,
    # which is right below the year-header row in the current
    # releases. (Row 4 = years, row 5 = labels+units, row 6 =
    # data.) Skip ahead until we find a row whose SA2-code cell
    # is actually a 9-digit code.
    data_start = year_header_row_idx + 1
    while data_start < len(rows):
        r = rows[data_start]
        if len(r) > _TABLE1_SA2_CODE_COL:
            cell = r[_TABLE1_SA2_CODE_COL]
            if cell is not None:
                s = str(cell).strip()
                if len(s) == 9 and s.isdigit():
                    break
        data_start += 1

    latest_year = max(year_cols)

    records: list[dict[str, object]] = []
    for row in rows[data_start:]:
        if len(row) <= _TABLE1_SA2_CODE_COL:
            continue
        sa2_raw = row[_TABLE1_SA2_CODE_COL]
        sa2 = "" if sa2_raw is None else str(sa2_raw).strip()
        if not (len(sa2) == 9 and sa2.isdigit()):
            continue
        state_name = str(row[1]).strip() if row[1] is not None else ""
        rec: dict[str, object] = {
            "sa2_code_2021": sa2,
            "state_abbreviation": _state_to_abbreviation(state_name),
            "reference_year": latest_year,
        }
        # Latest year's population goes to a stable column name.
        if latest_year in year_cols:
            rec["population_total"] = _coerce_number(row[year_cols[latest_year]])
        # Full year history.
        for year, col_idx in year_cols.items():
            key = f"population_history_{year}"
            rec[key] = _coerce_number(row[col_idx]) if col_idx < len(row) else None
        records.append(rec)

    if not records:
        raise RuntimeError(f"No SA2 data rows in {xlsx_path}")

    df = pd.DataFrame.from_records(records)
    return df.set_index("sa2_code_2021")


def _parse_ds0002_workbook(xlsx_path: Path) -> pd.DataFrame:
    """Parse the 3235.0 DS0002 age/sex workbook (Table 1).

    Returns a DataFrame indexed by ``sa2_code_2021`` with the six
    wishlist columns. Counts for the three age bands are derived by
    multiplying the published percentage by the persons total (DS0002
    columns 12 & 15-17); ABS only publishes the percentages directly.

    Header layout matches DS0003: title rows + row 4 (group labels) +
    row 5 (geography labels + units row). Data starts at row 6.
    """
    import openpyxl  # noqa: PLC0415

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    if "Table 1" not in wb.sheetnames:
        wb.close()
        raise RuntimeError(
            f"ERP age/sex workbook {xlsx_path} has no 'Table 1' sheet. Sheets: {wb.sheetnames}"
        )
    ws = wb["Table 1"]
    rows: list[list[object]] = []
    for r in ws.iter_rows(values_only=True):
        rows.append(list(r))
    wb.close()

    # Find the first row whose SA2-code cell is a 9-digit code.
    data_start = 0
    for ridx in range(min(15, len(rows))):
        r = rows[ridx]
        if len(r) > _DS0002_SA2_CODE_COL:
            cell = r[_DS0002_SA2_CODE_COL]
            if cell is not None:
                s = str(cell).strip()
                if len(s) == 9 and s.isdigit():
                    data_start = ridx
                    break

    if data_start == 0:
        raise RuntimeError(
            f"No SA2-shaped data rows in first 15 rows of {xlsx_path}. Layout may have shifted."
        )

    records: list[dict[str, object]] = []
    for row in rows[data_start:]:
        if len(row) <= _DS0002_PCT_65_PLUS_COL:
            continue
        sa2_raw = row[_DS0002_SA2_CODE_COL]
        sa2 = "" if sa2_raw is None else str(sa2_raw).strip()
        if not (len(sa2) == 9 and sa2.isdigit()):
            continue

        males = _coerce_number(row[_DS0002_MALES_COL])
        females = _coerce_number(row[_DS0002_FEMALES_COL])
        persons = _coerce_number(row[_DS0002_PERSONS_COL])
        median_age = _coerce_number(row[_DS0002_MEDIAN_AGE_COL])
        pct_0_14 = _coerce_number(row[_DS0002_PCT_0_14_COL])
        pct_15_64 = _coerce_number(row[_DS0002_PCT_15_64_COL])
        pct_65 = _coerce_number(row[_DS0002_PCT_65_PLUS_COL])

        rec: dict[str, object] = {
            "sa2_code_2021": sa2,
            "population_male": males,
            "population_female": females,
            "median_age": median_age,
        }
        # Derive band counts from persons × pct/100. Round to int —
        # ABS publishes percentages to one decimal so the derived
        # counts are accurate to about ±0.5% of persons.
        if isinstance(persons, (int, float)) and isinstance(pct_0_14, (int, float)):
            rec["population_0_14"] = int(round(persons * pct_0_14 / 100))
        else:
            rec["population_0_14"] = None
        if isinstance(persons, (int, float)) and isinstance(pct_15_64, (int, float)):
            rec["population_15_64"] = int(round(persons * pct_15_64 / 100))
        else:
            rec["population_15_64"] = None
        if isinstance(persons, (int, float)) and isinstance(pct_65, (int, float)):
            rec["population_65_plus"] = int(round(persons * pct_65 / 100))
        else:
            rec["population_65_plus"] = None
        records.append(rec)

    if not records:
        raise RuntimeError(f"No SA2 data rows in {xlsx_path}")

    df = pd.DataFrame.from_records(records)
    return df.set_index("sa2_code_2021")


# ---- helpers ------------------------------------------------------------


_STATE_NAMES = {
    "new south wales": "NSW",
    "victoria": "VIC",
    "queensland": "QLD",
    "south australia": "SA",
    "western australia": "WA",
    "tasmania": "TAS",
    "northern territory": "NT",
    "australian capital territory": "ACT",
    "other territories": "OT",
}


def _state_to_abbreviation(name: str) -> str:
    return _STATE_NAMES.get(name.lower().strip(), name.strip()[:3].upper())


def _coerce_number(cell: object) -> object:
    """Coerce a numeric-looking cell to int / float / None."""
    if cell is None:
        return None
    if isinstance(cell, (int, float)):
        return cell
    s = str(cell).strip()
    if not s or s.lower() in ("np", "na", "n/a", "-", "..", "."):
        return None
    if re.fullmatch(r"-?\d+", s):
        return int(s)
    if re.fullmatch(r"-?\d+\.\d+", s):
        return float(s)
    return None


# ---- fetcher registration ------------------------------------------------


def _build_fetcher(root: Path, release: str | None = None) -> ErpDataSource:
    kwargs: dict[str, object] = {"root": root}
    if release is not None:
        kwargs["release"] = release
    return ErpDataSource(**kwargs)  # type: ignore[arg-type]


def _register() -> None:
    from . import registry  # noqa: PLC0415

    registry.register_fetcher("erp_by_sa2", _build_fetcher)


_register()
