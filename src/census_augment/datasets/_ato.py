"""ABS Personal Income (ATO administrative) fetcher (spec §20, dataset id ``ato_personal_income``).

ABS publishes Personal Income annually as a series of XLSX workbooks
(Tables 1–14). Table 1 carries the **total-income summary statistics**
(earners, median age, sum, median, mean) — the most-used columns for
augmentation. Tables 2–9 carry breakdowns by age/sex, income
distribution, and per-income-type (employee, investment, super,
own-business). v1.3 implements Table 1 only; the other tables can land
as follow-ups using the same parser pattern.

Within each Table file there's a sheet per geography level:

- Table 1.1 — GCCSA
- Table 1.2 — SA4
- Table 1.3 — SA3
- Table 1.4 — **SA2** (what we want)
- Table 1.5 — LGA

We scrape the landing page to find the Table 1 URL, download the XLSX
(~1 MB), then parse Table 1.4 into a SA2-keyed DataFrame.
"""

from __future__ import annotations

import logging
import re
from pathlib import Path

import pandas as pd
import requests

_log = logging.getLogger(__name__)

ATO_LANDING_URL = (
    "https://www.abs.gov.au/statistics/labour/earnings-and-working-conditions/"
    "personal-income-australia/latest-release"
)

# Table 1 covers total-income summary stats — the file we download.
_TABLE1_URL_FRAGMENT = "Table%201"

# Within Table 1's workbook, sheet 1.4 holds the SA2 series.
_SA2_SHEET = "Table 1.4"

# Group header → output column for that group's *latest year* value.
# Row 5 of Table 1.4 has group labels at fixed columns. We resolve
# each group's last-year cell by reading the financial-year row (row 6)
# and picking the rightmost year per group.
_GROUP_COLUMNS: dict[str, str] = {
    "Earners (persons)": "income_earners_count",
    "Median age of earners (years)": "median_age_of_earners",
    "Sum ($)": "sum_total_income",
    "Median ($)": "median_total_income",
    "Mean ($)": "mean_total_income",
}


class AtoDataSource:
    """Fetch + load ABS Personal Income SA2 (Table 1 summary).

    Implements the :class:`DatasetFetcher` Protocol. ``release`` is the
    reference financial year (e.g. ``"2022-23"`` for the FY 2022-23
    release). ``"latest"`` resolves the landing page's most recent.
    """

    _label = "ABS Personal Income SA2 (Table 1)"

    def __init__(
        self,
        *,
        release: str = "latest",
        root: Path,
        landing_url: str = ATO_LANDING_URL,
        session: requests.Session | None = None,
        chunk_size: int = 256 * 1024,
        timeout: float = 60.0,
    ) -> None:
        self._release_request = release
        self._root = Path(root)
        self._landing_url = landing_url
        self._session = session if session is not None else requests.Session()
        self._chunk_size = chunk_size
        self._timeout = timeout
        self._resolved_release: str | None = None
        self._resolved_url: str | None = None

    # ---- protocol -------------------------------------------------------

    @property
    def resolved_release(self) -> str:
        if self._resolved_release is None:
            self._resolve_release()
        assert self._resolved_release is not None
        return self._resolved_release

    @property
    def is_cached(self) -> bool:
        if self._resolved_release is not None:
            return self._xlsx_path.exists()
        return self._root.exists() and any(self._root.glob("ato-*.xlsx"))

    @property
    def _xlsx_path(self) -> Path:
        return self._root / f"ato-personal-income-{self.resolved_release}.xlsx"

    @property
    def _parquet_path(self) -> Path:
        return self._root / f"ato-personal-income-{self.resolved_release}.parquet"

    def fetch(self, refresh: bool = False) -> Path:
        self._resolve_release()
        if self._xlsx_path.exists() and not refresh:
            _log.debug("ATO cached at %s", self._xlsx_path)
            return self._xlsx_path

        self._root.mkdir(parents=True, exist_ok=True)
        tmp = self._xlsx_path.with_suffix(self._xlsx_path.suffix + ".tmp")
        url = self._resolved_url or ""
        _log.info(
            "Downloading %s (%s) from %s",
            self._label,
            self.resolved_release,
            url,
        )
        with self._session.get(url, stream=True, timeout=self._timeout) as response:
            response.raise_for_status()
            with tmp.open("wb") as f:
                for chunk in response.iter_content(chunk_size=self._chunk_size):
                    if chunk:
                        f.write(chunk)
        tmp.replace(self._xlsx_path)
        _log.info("Saved %s to %s", self._label, self._xlsx_path)
        return self._xlsx_path

    def load(self) -> pd.DataFrame:
        if self._resolved_release is not None and self._parquet_path.exists():
            return pd.read_parquet(self._parquet_path).set_index("sa2_code_2021")

        xlsx = self.fetch()
        df = self._parse_xlsx(xlsx)
        df["reference_financial_year"] = self.resolved_release
        df.reset_index().to_parquet(self._parquet_path, index=False)
        return df

    # ---- release resolution --------------------------------------------

    def _resolve_release(self) -> None:
        if self._resolved_release is not None:
            return

        _log.debug("Resolving ATO release via %s", self._landing_url)
        resp = self._session.get(self._landing_url, timeout=self._timeout)
        resp.raise_for_status()
        html = resp.text

        # Match links like
        # /personal-income-australia/2022-23/Table%201%20-%20...xlsx
        pattern = re.compile(
            r'href="([^"]*personal-income-australia/'
            r"(?P<period>\d{4}-\d{2,4})/"
            r'[^"]*' + _TABLE1_URL_FRAGMENT + r'[^"]*\.xlsx)"',
            re.IGNORECASE,
        )

        candidates: list[tuple[str, str]] = []
        for m in pattern.finditer(html):
            href = m.group(1)
            period = m.group("period")
            url = (
                href
                if href.startswith("http")
                else ("https://www.abs.gov.au" + (href if href.startswith("/") else "/" + href))
            )
            candidates.append((period, url))

        if not candidates:
            raise RuntimeError(f"Could not find any Table 1 link on {self._landing_url}")

        if self._release_request == "latest":
            picked = max(candidates, key=lambda t: t[0])
        else:
            matching = [c for c in candidates if c[0] == self._release_request]
            if not matching:
                raise RuntimeError(
                    f"ATO release {self._release_request!r} not found. "
                    f"Available: {sorted({c[0] for c in candidates}, reverse=True)}"
                )
            picked = matching[0]

        self._resolved_release = picked[0]
        self._resolved_url = picked[1]
        _log.info(
            "Resolved ATO release=%s, url=%s",
            self._resolved_release,
            self._resolved_url,
        )

    # ---- parsing -------------------------------------------------------

    @staticmethod
    def _parse_xlsx(xlsx_path: Path) -> pd.DataFrame:
        import openpyxl  # noqa: PLC0415

        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
        if _SA2_SHEET not in wb.sheetnames:
            wb.close()
            raise RuntimeError(
                f"ATO workbook {xlsx_path} has no '{_SA2_SHEET}' sheet. Sheets: {wb.sheetnames}"
            )

        ws = wb[_SA2_SHEET]
        rows: list[list[object]] = []
        for r in ws.iter_rows(values_only=True):
            rows.append(list(r))
        wb.close()

        # Locate the group-header row (first non-blank row that contains
        # any of the known group labels).
        group_row_idx = -1
        for i, row in enumerate(rows[:10]):
            text = " ".join("" if c is None else str(c) for c in row)
            if any(label in text for label in _GROUP_COLUMNS):
                group_row_idx = i
                break

        if group_row_idx < 0:
            raise RuntimeError("Could not find ATO group header row in Table 1.4")

        # Map group label → starting column index (the column where the
        # group label appears).
        group_starts: list[tuple[int, str]] = []
        group_row = rows[group_row_idx]
        for col_idx, cell in enumerate(group_row):
            if cell is None:
                continue
            text = str(cell).strip()
            if text in _GROUP_COLUMNS:
                group_starts.append((col_idx, text))
        # Sort by col index so we can find the year-cells per group.
        group_starts.sort()

        # Below the group header is the year header (row group_row_idx+1).
        year_row = rows[group_row_idx + 1]
        # Sniff the financial-year cells (e.g. "2022-23") in each group.
        # For each group, we want the *rightmost* year cell — that's the
        # latest reference period.
        latest_col_per_group: dict[str, int] = {}
        # group_starts is sorted; each group spans from its start col
        # to the start of the next group (exclusive).
        for i, (col_idx, label) in enumerate(group_starts):
            end = group_starts[i + 1][0] if i + 1 < len(group_starts) else len(year_row)
            # Walk year_row[col_idx:end] looking for the last
            # financial-year-shaped cell.
            last_year_col = -1
            for j in range(col_idx, min(end, len(year_row))):
                cell = year_row[j]
                if cell is None:
                    continue
                s = str(cell).strip()
                if re.fullmatch(r"\d{4}-\d{2}", s):
                    last_year_col = j
            if last_year_col >= 0:
                latest_col_per_group[label] = last_year_col

        # Read data rows below the year row.
        data_start = group_row_idx + 2
        records: list[dict[str, object]] = []
        for row in rows[data_start:]:
            if not row:
                continue
            sa2_raw = row[0]
            sa2 = "" if sa2_raw is None else str(sa2_raw).strip()
            if not (len(sa2) == 9 and sa2.isdigit()):
                continue
            rec: dict[str, object] = {"sa2_code_2021": sa2}
            for label, col_idx in latest_col_per_group.items():
                output_name = _GROUP_COLUMNS[label]
                if col_idx < len(row):
                    rec[output_name] = _coerce_number(row[col_idx])
                else:
                    rec[output_name] = None
            records.append(rec)

        if not records:
            raise RuntimeError(f"No SA2 data rows in {xlsx_path}")

        df = pd.DataFrame.from_records(records)
        return df.set_index("sa2_code_2021")


def _coerce_number(cell: object) -> object:
    if cell is None:
        return None
    if isinstance(cell, (int, float)):
        return cell
    s = str(cell).strip()
    if not s or s.lower() in ("np", "na", "n/a", "-", "..", ".", "<5", "nan", "null"):
        return None
    if re.fullmatch(r"-?\d+", s):
        return int(s)
    if re.fullmatch(r"-?\d+\.\d+", s):
        return float(s)
    return None
