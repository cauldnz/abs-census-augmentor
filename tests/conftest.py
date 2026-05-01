"""Shared test fixtures used across the suite.

These are pytest fixtures (auto-discovered by pytest from a top-level
``conftest.py``); test files reference them by name without imports.
"""

from __future__ import annotations

import io
import zipfile
from pathlib import Path
from typing import Any

import geopandas as gpd
import openpyxl
import pandas as pd
import pyarrow as pa
import pyarrow.parquet as pq
import pytest
from shapely.geometry import Polygon


# ---- Boundary fixtures (synthetic SA2 polygons) ----------------------------

# Three synthetic SA2 polygons covering parts of inner Sydney. Coordinates
# are in EPSG:7844 (GDA2020) to match the real ABS boundary CRS.
_FAKE_SA2_RECORDS = [
    {
        "SA2_CODE21": "117011326",
        "SA2_NAME21": "Sydney CBD",
        "polygon": [
            (151.20, -33.87),
            (151.22, -33.87),
            (151.22, -33.85),
            (151.20, -33.85),
        ],
    },
    {
        "SA2_CODE21": "117011327",
        "SA2_NAME21": "North Sydney",
        "polygon": [
            (151.19, -33.84),
            (151.21, -33.84),
            (151.21, -33.82),
            (151.19, -33.82),
        ],
    },
    {
        "SA2_CODE21": "117011328",
        "SA2_NAME21": "Eastern Suburbs",
        "polygon": [
            (151.23, -33.89),
            (151.26, -33.89),
            (151.26, -33.86),
            (151.23, -33.86),
        ],
    },
]


@pytest.fixture
def fake_sa2_gdf() -> gpd.GeoDataFrame:
    """Three-polygon synthetic SA2 GeoDataFrame in EPSG:7844 (GDA2020)."""
    return gpd.GeoDataFrame(
        {
            "SA2_CODE21": [r["SA2_CODE21"] for r in _FAKE_SA2_RECORDS],
            "SA2_NAME21": [r["SA2_NAME21"] for r in _FAKE_SA2_RECORDS],
            "geometry": [Polygon(r["polygon"]) for r in _FAKE_SA2_RECORDS],
        },
        crs="EPSG:7844",
    )


@pytest.fixture
def fake_boundary_zip_bytes(
    tmp_path: Path, fake_sa2_gdf: gpd.GeoDataFrame
) -> bytes:
    """In-memory ZIP containing fake SA2 shapefile + sidecars (.shp/.dbf/.prj/.shx).

    Mirrors the real ABS layout: filename inside the ZIP is the bare
    ``SA2_2021_AUST_GDA2020.shp`` (no ``SHP`` token; that only appears
    on the *ZIP* filename per spec §4.1).
    """
    work_dir = tmp_path / "_fixture_boundary"
    work_dir.mkdir(parents=True, exist_ok=True)
    shp_path = work_dir / "SA2_2021_AUST_GDA2020.shp"
    fake_sa2_gdf.to_file(shp_path, driver="ESRI Shapefile")

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w") as zf:
        for sidecar in work_dir.iterdir():
            zf.write(sidecar, arcname=sidecar.name)
    return buf.getvalue()


# ---- DataPack fixtures (synthetic 3-row tables + realistic metadata) -------

# Synthetic G01 table (population) for the same 3 SA2 codes used in fake_sa2_gdf
_FAKE_G01_RECORDS = [
    {"SA2_CODE_2021": "117011326", "Tot_P_M": 5000, "Tot_P_F": 5200, "Tot_P_P": 10200},
    {"SA2_CODE_2021": "117011327", "Tot_P_M": 4500, "Tot_P_F": 4800, "Tot_P_P": 9300},
    {"SA2_CODE_2021": "117011328", "Tot_P_M": 6000, "Tot_P_F": 6300, "Tot_P_P": 12300},
]

# Synthetic G02 table (selected medians) for the same 3 SA2 codes
_FAKE_G02_RECORDS = [
    {
        "SA2_CODE_2021": "117011326",
        "Median_age_persons": 35,
        "Median_tot_hhd_inc_weekly": 1800,
        "Median_rent_weekly": 550,
    },
    {
        "SA2_CODE_2021": "117011327",
        "Median_age_persons": 38,
        "Median_tot_hhd_inc_weekly": 2200,
        "Median_rent_weekly": 650,
    },
    {
        "SA2_CODE_2021": "117011328",
        "Median_age_persons": 42,
        "Median_tot_hhd_inc_weekly": 2500,
        "Median_rent_weekly": 700,
    },
]

# Real ABS descriptor-sheet rows: (Sequential, Short, Long, DataPackfile,
# Profiletable, Columnheadingdescriptioninprofile). Note the deliberate
# divergence between Long (underscored) and Columnheadingdescriptioninprofile
# (proper text) — that's what real 2021 GCP looks like.
_FAKE_DESCRIPTOR_ROWS = [
    ("G1", "Tot_P_M", "Total_Persons_Male", "G01", "G01", "Males"),
    ("G2", "Tot_P_F", "Total_Persons_Female", "G01", "G01", "Females"),
    ("G3", "Tot_P_P", "Total_Persons_Persons", "G01", "G01", "Persons"),
    ("G109", "Median_age_persons", "Median_age_of_persons",
     "G02", "G02", "Median age of persons"),
    ("G115", "Median_tot_hhd_inc_weekly", "Median_total_household_income_weekly",
     "G02", "G02", "Median total household income ($/weekly)"),
    ("G112", "Median_rent_weekly", "Median_rent_weekly",
     "G02", "G02", "Median rent ($/weekly)"),
]

# Real ABS Table-Number sheet rows: (Table Number, Table Name, Table population)
_FAKE_TABLE_ROWS = [
    ("G01", "Selected Person Characteristics by Sex", "Persons"),
    ("G02", "Selected Medians and Averages", None),
]


@pytest.fixture
def fake_g01_df() -> pd.DataFrame:
    return pd.DataFrame(_FAKE_G01_RECORDS)


@pytest.fixture
def fake_g02_df() -> pd.DataFrame:
    return pd.DataFrame(_FAKE_G02_RECORDS)


@pytest.fixture
def fake_descriptor_rows() -> list[tuple[Any, ...]]:
    return list(_FAKE_DESCRIPTOR_ROWS)


@pytest.fixture
def fake_table_rows() -> list[tuple[Any, ...]]:
    return list(_FAKE_TABLE_ROWS)


def build_metadata_xlsx(
    descriptor_rows: list[tuple[Any, ...]],
    table_rows: list[tuple[Any, ...]] | None = None,
    *,
    descriptor_sheet_name: str = "Cell Descriptors Information",
    table_sheet_name: str = "Table Number, Name, Population",
    descriptor_columns: list[str] | None = None,
    title_row_count: int = 6,
) -> bytes:
    """Build an in-memory metadata workbook that mirrors real ABS layout.

    Title rows come first (blank padding + decorative text), then the
    column header row, then data rows. The "Table population " trailing
    whitespace in the table sheet header is intentional (matches real ABS).
    """
    if descriptor_columns is None:
        descriptor_columns = [
            "Sequential",
            "Short",
            "Long",
            "DataPackfile",
            "Profiletable",
            "Columnheadingdescriptioninprofile",
        ]
    wb = openpyxl.Workbook()

    ws_d = wb.active
    ws_d.title = descriptor_sheet_name
    for _ in range(title_row_count):
        ws_d.append([None] * len(descriptor_columns))
    ws_d.append(["2021 Census of Population and Housing"])
    ws_d.append(["General Community Profile DataPack Metadata"])
    ws_d.append([None] * len(descriptor_columns))
    ws_d.append([None, None, "Celldescriptors", None, None, None])
    ws_d.append(descriptor_columns)
    for row in descriptor_rows:
        ws_d.append(list(row))

    if table_rows is not None:
        ws_t = wb.create_sheet(table_sheet_name)
        for _ in range(title_row_count):
            ws_t.append([None, None, None])
        ws_t.append(["2021 Census of Population and Housing"])
        ws_t.append(["General Community Profile Tables"])
        ws_t.append(["Table Number", "Table Name", "Table population "])
        for row in table_rows:
            ws_t.append(list(row))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@pytest.fixture
def fake_metadata_xlsx_bytes(
    fake_descriptor_rows: list[tuple[Any, ...]],
    fake_table_rows: list[tuple[Any, ...]],
) -> bytes:
    return build_metadata_xlsx(fake_descriptor_rows, fake_table_rows)


def _empty_xlsx_bytes() -> bytes:
    wb = openpyxl.Workbook()
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@pytest.fixture
def fake_datapack_zip_bytes(
    fake_g01_df: pd.DataFrame,
    fake_g02_df: pd.DataFrame,
    fake_metadata_xlsx_bytes: bytes,
) -> bytes:
    """In-memory ZIP mirroring real 2021 GCP DataPack layout.

    Includes:
    - CSVs in a long-named subdirectory with realistic ABS filenames.
    - Real metadata Excel under ``Metadata/Metadata_2021_GCP_DataPack_R1_R2.xlsx``.
    - Two "noise" sibling xlsx files that the parser must ignore.
    """
    csv_dir = "2021 Census GCP Statistical Area 2 for AUS"
    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, "w") as zf:
        zf.writestr(
            f"{csv_dir}/2021Census_G01_AUST_SA2.csv",
            fake_g01_df.to_csv(index=False),
        )
        zf.writestr(
            f"{csv_dir}/2021Census_G02_AUST_SA2.csv",
            fake_g02_df.to_csv(index=False),
        )
        zf.writestr(
            "Metadata/Metadata_2021_GCP_DataPack_R1_R2.xlsx",
            fake_metadata_xlsx_bytes,
        )
        zf.writestr(
            "Metadata/2021Census_geog_desc_test.xlsx", _empty_xlsx_bytes()
        )
        zf.writestr(
            "Metadata/2021_GCP_Sequential_Template_Test.xlsx",
            _empty_xlsx_bytes(),
        )
    return zip_buf.getvalue()


# ---- G-NAF fixtures (synthetic Parquet matching the SA2 polygons above) ----
#
# Synthetic G-NAF Core "addresses" with MB_CODE values that resolve (via the
# fake_mb_correspondence fixture below) to the same 3 SA2 codes used in
# fake_sa2_gdf. Coordinates fall inside those polygons. Address labels
# exercise the four matching tiers (Phase 4):
#
#   * Tier 1 (exact ADDRESS_LABEL): match the input verbatim.
#   * Tier 2 (component): differ only in punctuation / abbreviation.
#   * Tier 3 (fuzzy): typo'd input maps to a known label.
#   * Miss: addresses guaranteed to fall through every G-NAF tier
#     (e.g. PO box surrogate) — exercised in Phase 4.

_FAKE_GNAF_RECORDS = [
    # (PID, ADDRESS_LABEL, LATITUDE, LONGITUDE, MB_CODE, POSTCODE)
    # POSTCODE is the Tier 2 / Tier 3 pre-filter key (spec §19.3).
    # SA2 117011326 (Sydney CBD area) — lat/lon inside fake_sa2_gdf polygon 1
    (
        "GANSW000000001",
        "1 GEORGE STREET SYDNEY NSW 2000",
        -33.864,
        151.211,
        "11701132601",
        "2000",
    ),
    (
        "GANSW000000002",
        "100 PITT STREET SYDNEY NSW 2000",
        -33.866,
        151.211,
        "11701132602",
        "2000",
    ),
    (
        "GANSW000000003",
        "200 KENT STREET SYDNEY NSW 2000",
        -33.868,
        151.205,
        "11701132603",
        "2000",
    ),
    # SA2 117011327 (North Sydney area) — inside polygon 2
    (
        "GANSW000000004",
        "10 BLUES POINT ROAD MCMAHONS POINT NSW 2060",
        -33.835,
        151.200,
        "11701132701",
        "2060",
    ),
    # SA2 117011328 (Eastern Suburbs) — inside polygon 3
    (
        "GANSW000000005",
        "5 BONDI ROAD BONDI NSW 2026",
        -33.880,
        151.245,
        "11701132801",
        "2026",
    ),
]


@pytest.fixture
def fake_gnaf_parquet_bytes() -> bytes:
    """In-memory G-NAF Core Parquet (5 synthetic Sydney addresses).

    Schema matches the v1.0 spec §4.3 minimum:
    ``ADDRESS_DETAIL_PID``, ``ADDRESS_LABEL``, ``LATITUDE``, ``LONGITUDE``,
    ``MB_CODE``.
    """
    table = pa.table(
        {
            "ADDRESS_DETAIL_PID": [r[0] for r in _FAKE_GNAF_RECORDS],
            "ADDRESS_LABEL": [r[1] for r in _FAKE_GNAF_RECORDS],
            "LATITUDE": [r[2] for r in _FAKE_GNAF_RECORDS],
            "LONGITUDE": [r[3] for r in _FAKE_GNAF_RECORDS],
            "MB_CODE": [r[4] for r in _FAKE_GNAF_RECORDS],
            "POSTCODE": [r[5] for r in _FAKE_GNAF_RECORDS],
        }
    )
    buf = io.BytesIO()
    pq.write_table(table, buf)
    return buf.getvalue()


@pytest.fixture
def fake_gnaf_data_dir(
    tmp_path: Path, fake_gnaf_parquet_bytes: bytes
) -> Path:
    """Set up a tmp data_dir with one G-NAF release cached.

    Creates ``<tmp>/data/gnaf/202602/addresses.parquet`` and returns
    the ``data_dir`` to pass into :class:`GnafDataSource`.
    """
    data_dir = tmp_path / "data"
    release_dir = data_dir / "gnaf" / "202602"
    release_dir.mkdir(parents=True)
    (release_dir / "addresses.parquet").write_bytes(fake_gnaf_parquet_bytes)
    return data_dir


@pytest.fixture
def fake_gnaf_data_dir_with_two_releases(
    tmp_path: Path, fake_gnaf_parquet_bytes: bytes
) -> Path:
    """A data_dir with two cached releases — useful for testing
    ``release='latest'`` resolution."""
    data_dir = tmp_path / "data"
    for release in ("202511", "202602"):
        rel_dir = data_dir / "gnaf" / release
        rel_dir.mkdir(parents=True)
        (rel_dir / "addresses.parquet").write_bytes(fake_gnaf_parquet_bytes)
    return data_dir


# ---- MB → SA2 correspondence fixtures (synthetic Mesh Block shapefile) ----
#
# The real MB→SA2 mapping (spec §15.1 resolution) is the .dbf attribute
# table of ``MB_2021_AUST_SHP_GDA2020.zip``. We emit a tiny synthetic
# shapefile with the same year-suffixed column names ABS uses, so the
# parser exercises the column-detection logic. MB_CODE values match the
# G-NAF fixture above so end-to-end (G-NAF mb_code → SA2) tests can
# round-trip.

_FAKE_MB_RECORDS = [
    # (MB_CODE21, SA2_CODE21, SA2_NAME21)
    # SA2 117011326 (Sydney CBD) — three MBs
    ("11701132601", "117011326", "Sydney CBD"),
    ("11701132602", "117011326", "Sydney CBD"),
    ("11701132603", "117011326", "Sydney CBD"),
    # SA2 117011327 (North Sydney) — one MB
    ("11701132701", "117011327", "North Sydney"),
    # SA2 117011328 (Eastern Suburbs) — one MB
    ("11701132801", "117011328", "Eastern Suburbs"),
]


@pytest.fixture
def fake_mb_gdf() -> gpd.GeoDataFrame:
    """Synthetic Mesh Block GeoDataFrame with 5 MBs across 3 SA2s.

    Geometry is a 1-degree square dummy per row (we never read it from
    the .dbf; it's only present so geopandas can write a valid shapefile).
    Column names follow the real ABS Mesh Block shapefile convention
    (``MB_CODE21``, ``SA2_CODE21``, ``SA2_NAME21`` — 10-char DBF limit),
    so the parser exercises its column-detection logic against the same
    names production data uses.
    """
    return gpd.GeoDataFrame(
        {
            "MB_CODE21": [r[0] for r in _FAKE_MB_RECORDS],
            "SA2_CODE21": [r[1] for r in _FAKE_MB_RECORDS],
            "SA2_NAME21": [r[2] for r in _FAKE_MB_RECORDS],
            "geometry": [
                Polygon(
                    [
                        (151.20 + i * 0.01, -33.87),
                        (151.21 + i * 0.01, -33.87),
                        (151.21 + i * 0.01, -33.86),
                        (151.20 + i * 0.01, -33.86),
                    ]
                )
                for i in range(len(_FAKE_MB_RECORDS))
            ],
        },
        crs="EPSG:7844",
    )


@pytest.fixture
def fake_mb_correspondence_zip_bytes(
    tmp_path: Path, fake_mb_gdf: gpd.GeoDataFrame
) -> bytes:
    """In-memory ZIP containing the synthetic Mesh Block shapefile.

    Filename inside the ZIP follows the ABS convention:
    ``MB_2021_AUST_GDA2020.shp`` (no ``SHP`` token in the inner name —
    that token only appears on the outer ZIP filename, mirroring the
    boundary fixture).
    """
    work_dir = tmp_path / "_fixture_mb"
    work_dir.mkdir(parents=True, exist_ok=True)
    shp_path = work_dir / "MB_2021_AUST_GDA2020.shp"
    fake_mb_gdf.to_file(shp_path, driver="ESRI Shapefile")

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w") as zf:
        for sidecar in work_dir.iterdir():
            zf.write(sidecar, arcname=sidecar.name)
    return buf.getvalue()

