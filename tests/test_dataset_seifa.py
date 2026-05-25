"""Tests for the SEIFA dataset fetcher (spec §20, dataset id ``seifa``).

Hermetic tests use a synthetic XLSX whose layout mirrors the real ABS
publication (Tables 2-5 each carry one index with the full flavour
set). Real-network checks live in ``tools/verify_real_parsers.py``.

2016-release parsing is tested via ``_parse_grids`` directly (the pure
function that accepts a pre-read grid dict) to avoid needing a real .xls
binary in the test suite.  Download-path tests for 2016 verify URL and
filename behaviour only (no actual calamine parse).
"""

from __future__ import annotations

import io
from pathlib import Path

import openpyxl
import pandas as pd
import pytest
import requests
import responses

from census_augment.datasets._seifa import (
    DEFAULT_SEIFA_2016_URL,
    DEFAULT_SEIFA_2021_URL,
    SeifaDataSource,
    _parse_grids,
)


def _index_sheet_rows(prefix: str, sa2_records: list[tuple[str, str, dict]]) -> list[list[object]]:
    """Build the rows of a single Table 2/3/4/5-style sheet.

    ``sa2_records`` is a list of (sa2_code, sa2_name, fields_dict)
    tuples. ``fields_dict`` keys: urp, score, aus_rank, aus_decile,
    aus_percentile, state, state_rank, state_decile, state_percentile,
    sa1_min, sa1_max, pct_urp_no_score.
    """
    rows: list[list[object]] = [
        ["Australian Bureau of Statistics"],
        [f"Index of {prefix.upper()} — SEIFA 2021"],
        ["Released 27 April 2023"],
        [f"Table for {prefix} index"],
        # Group row (top of header) — we don't actually use it but
        # mirror the real layout.
        [
            "",
            "",
            "",
            "",
            "",
            "Ranking within Australia",
            "",
            "",
            "",
            "Ranking within State or Territory",
            "",
            "",
            "",
            "",
            "",
            "",
        ],
        # Header row.
        [
            "2021 Statistical Area Level 2 (SA2) 9-Digit Code",
            "2021 Statistical Area Level 2 (SA2) Name",
            "Usual Resident Population",
            "Score",
            "",
            "Rank",
            "Decile",
            "Percentile",
            "",
            "State",
            "Rank",
            "Decile",
            "Percentile",
            "Minimum score for SA1s in area",
            "Maximum score for SA1s in area",
            "% Usual Resident Population without a score",
        ],
    ]
    for sa2_code, sa2_name, f in sa2_records:
        rows.append(
            [
                sa2_code,
                sa2_name,
                f["urp"],
                f["score"],
                "",
                f["aus_rank"],
                f["aus_decile"],
                f["aus_percentile"],
                "",
                f["state"],
                f["state_rank"],
                f["state_decile"],
                f["state_percentile"],
                f.get("sa1_min", 0),
                f.get("sa1_max", 0),
                f.get("pct_urp_no_score", 0),
            ]
        )
    return rows


def _build_synthetic_seifa_xlsx() -> bytes:
    """A SEIFA-shaped workbook with Tables 1-5 mirroring the real layout.

    Two SA2 records per sheet, with deterministic values per index so
    tests can assert specific cells.
    """
    wb = openpyxl.Workbook()
    # Replace the default sheet with a Contents sheet (skipped at parse).
    ws = wb.active
    ws.title = "Contents"
    ws.append(["Contents page (skipped)"])

    # Table 1 — summary (we don't parse this in v1.3, but include it
    # for fidelity to the real workbook).
    summary = wb.create_sheet("Table 1")
    summary.append(["Summary table"])
    summary.append([])
    summary.append([])
    summary.append([])
    summary.append([])
    summary.append(
        [
            "2021 Statistical Area Level 2 (SA2) 9-Digit Code",
            "2021 Statistical Area Level 2 (SA2) Name",
            "Score",
            "Decile",
            "Score",
            "Decile",
            "Score",
            "Decile",
            "Score",
            "Decile",
            "Usual Resident Population",
        ]
    )
    summary.append(["117011326", "Sydney CBD", 1042, 8, 1080, 9, 1010, 7, 1100, 9, 12000])

    sa2_a = {
        "urp": 12000,
        "score": 1042,
        "aus_rank": 1500,
        "aus_decile": 8,
        "aus_percentile": 75,
        "state": "NSW",
        "state_rank": 800,
        "state_decile": 8,
        "state_percentile": 78,
        "sa1_min": 1000,
        "sa1_max": 1080,
        "pct_urp_no_score": 0,
    }
    sa2_b = {
        "urp": 9500,
        "score": 950,
        "aus_rank": 7000,
        "aus_decile": 5,
        "aus_percentile": 50,
        "state": "NSW",
        "state_rank": 3000,
        "state_decile": 5,
        "state_percentile": 50,
        "sa1_min": 920,
        "sa1_max": 985,
        "pct_urp_no_score": 0,
    }

    for sheet_name, prefix in [
        ("Table 2", "irsd"),
        ("Table 3", "irsad"),
        ("Table 4", "ier"),
        ("Table 5", "ieo"),
    ]:
        sheet = wb.create_sheet(sheet_name)
        # Each index gets distinguishably different scores so we can
        # tell them apart in assertions.
        offset = {"irsd": 0, "irsad": 38, "ier": -32, "ieo": 58}[prefix]
        a = dict(sa2_a, score=sa2_a["score"] + offset)
        b = dict(sa2_b, score=sa2_b["score"] + offset)
        rows = _index_sheet_rows(
            prefix,
            [
                ("117011326", "Sydney CBD", a),
                ("117011327", "North Sydney", b),
                # An aggregate row that should be filtered (not 9 digits).
                ("Australia", "Australia", dict(sa2_a, score=1000)),
            ],
        )
        for row in rows:
            sheet.append(row)

    # Table 6 — excluded SA2s (skipped).
    excluded = wb.create_sheet("Table 6")
    excluded.append(["Excluded SA2s (skipped)"])

    notes = wb.create_sheet("Explanatory Notes")
    notes.append(["Methodology notes (skipped)"])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@pytest.fixture
def synthetic_seifa_bytes() -> bytes:
    return _build_synthetic_seifa_xlsx()


@pytest.fixture
def seifa_data_dir(tmp_path: Path) -> Path:
    return tmp_path / "seifa-cache"


# ---- fetch ---------------------------------------------------------------


@responses.activate
def test_fetch_downloads_and_caches(
    tmp_path: Path,
    synthetic_seifa_bytes: bytes,
    seifa_data_dir: Path,
) -> None:
    responses.add(
        responses.GET,
        DEFAULT_SEIFA_2021_URL,
        body=synthetic_seifa_bytes,
        status=200,
    )
    ds = SeifaDataSource(root=seifa_data_dir)
    path = ds.fetch()
    assert path.exists()
    assert path.suffix == ".xlsx"
    assert ds.is_cached
    assert ds.resolved_release == "2021"


@responses.activate
def test_fetch_idempotent(
    synthetic_seifa_bytes: bytes,
    seifa_data_dir: Path,
) -> None:
    responses.add(
        responses.GET,
        DEFAULT_SEIFA_2021_URL,
        body=synthetic_seifa_bytes,
        status=200,
    )
    ds = SeifaDataSource(root=seifa_data_dir)
    ds.fetch()
    ds.fetch()
    assert len(responses.calls) == 1


@responses.activate
def test_fetch_refresh_redownloads(
    synthetic_seifa_bytes: bytes,
    seifa_data_dir: Path,
) -> None:
    responses.add(
        responses.GET,
        DEFAULT_SEIFA_2021_URL,
        body=synthetic_seifa_bytes,
        status=200,
    )
    responses.add(
        responses.GET,
        DEFAULT_SEIFA_2021_URL,
        body=synthetic_seifa_bytes,
        status=200,
    )
    ds = SeifaDataSource(root=seifa_data_dir)
    ds.fetch()
    ds.fetch(refresh=True)
    assert len(responses.calls) == 2


@responses.activate
def test_fetch_404_raises(seifa_data_dir: Path) -> None:
    responses.add(responses.GET, DEFAULT_SEIFA_2021_URL, status=404)
    ds = SeifaDataSource(root=seifa_data_dir)
    with pytest.raises(requests.HTTPError):
        ds.fetch()


def test_invalid_release_raises(seifa_data_dir: Path) -> None:
    with pytest.raises(ValueError, match="release must be"):
        SeifaDataSource(release="2026", root=seifa_data_dir)


def test_supported_releases(seifa_data_dir: Path) -> None:
    """2016 and 2021 are the only supported releases; others raise."""
    # These should not raise.
    SeifaDataSource(release="2021", root=seifa_data_dir)
    SeifaDataSource(release="2016", root=seifa_data_dir)
    # Anything else should.
    with pytest.raises(ValueError, match="release must be"):
        SeifaDataSource(release="2011", root=seifa_data_dir)


# ---- parse ---------------------------------------------------------------


@responses.activate
def test_load_returns_sa2_indexed_dataframe(
    synthetic_seifa_bytes: bytes,
    seifa_data_dir: Path,
) -> None:
    responses.add(
        responses.GET,
        DEFAULT_SEIFA_2021_URL,
        body=synthetic_seifa_bytes,
        status=200,
    )
    ds = SeifaDataSource(root=seifa_data_dir)
    df = ds.load()

    assert df.index.name == "sa2_code_2021"
    assert "117011326" in df.index
    assert "117011327" in df.index
    # Aggregate "Australia" row filtered out.
    assert "Australia" not in df.index


@responses.activate
def test_load_extracts_irsd_columns(
    synthetic_seifa_bytes: bytes,
    seifa_data_dir: Path,
) -> None:
    responses.add(
        responses.GET,
        DEFAULT_SEIFA_2021_URL,
        body=synthetic_seifa_bytes,
        status=200,
    )
    ds = SeifaDataSource(root=seifa_data_dir)
    df = ds.load()

    # IRSD: score offset is 0, so score == 1042 for Sydney CBD
    assert "irsd_score" in df.columns
    assert df.loc["117011326", "irsd_score"] == 1042
    assert df.loc["117011326", "irsd_aus_decile"] == 8
    assert df.loc["117011326", "irsd_aus_percentile"] == 75
    assert df.loc["117011326", "irsd_state_rank"] == 800
    assert df.loc["117011326", "irsd_state_decile"] == 8
    assert df.loc["117011326", "irsd_state_percentile"] == 78


@responses.activate
def test_load_extracts_all_four_indexes(
    synthetic_seifa_bytes: bytes,
    seifa_data_dir: Path,
) -> None:
    responses.add(
        responses.GET,
        DEFAULT_SEIFA_2021_URL,
        body=synthetic_seifa_bytes,
        status=200,
    )
    ds = SeifaDataSource(root=seifa_data_dir)
    df = ds.load()

    for prefix in ("irsd", "irsad", "ier", "ieo"):
        for flavour in (
            "score",
            "aus_rank",
            "aus_decile",
            "aus_percentile",
            "state_rank",
            "state_decile",
            "state_percentile",
        ):
            col = f"{prefix}_{flavour}"
            assert col in df.columns, f"missing {col}"


@responses.activate
def test_load_distinguishes_indexes_by_score(
    synthetic_seifa_bytes: bytes,
    seifa_data_dir: Path,
) -> None:
    """Each index has a different offset in the synthetic fixture; verify
    the parser keeps them separate."""
    responses.add(
        responses.GET,
        DEFAULT_SEIFA_2021_URL,
        body=synthetic_seifa_bytes,
        status=200,
    )
    ds = SeifaDataSource(root=seifa_data_dir)
    df = ds.load()

    row = df.loc["117011326"]
    assert row["irsd_score"] == 1042
    assert row["irsad_score"] == 1080
    assert row["ier_score"] == 1010
    assert row["ieo_score"] == 1100


@responses.activate
def test_load_caches_parquet_for_repeat_calls(
    synthetic_seifa_bytes: bytes,
    seifa_data_dir: Path,
) -> None:
    responses.add(
        responses.GET,
        DEFAULT_SEIFA_2021_URL,
        body=synthetic_seifa_bytes,
        status=200,
    )
    ds = SeifaDataSource(root=seifa_data_dir)
    df1 = ds.load()
    df2 = ds.load()
    pd.testing.assert_frame_equal(df1, df2)


@responses.activate
def test_load_includes_urp_and_state(
    synthetic_seifa_bytes: bytes,
    seifa_data_dir: Path,
) -> None:
    responses.add(
        responses.GET,
        DEFAULT_SEIFA_2021_URL,
        body=synthetic_seifa_bytes,
        status=200,
    )
    ds = SeifaDataSource(root=seifa_data_dir)
    df = ds.load()
    assert "urp" in df.columns
    assert df.loc["117011326", "urp"] == 12000
    assert "state_abbreviation" in df.columns
    assert df.loc["117011326", "state_abbreviation"] == "NSW"


@responses.activate
def test_load_handles_suppressed_cells_as_null(
    seifa_data_dir: Path,
) -> None:
    """``np`` cells in the source come through as NaN, not zero."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Contents"
    # Add Table 2 with one suppressed and one normal SA2.
    table2 = wb.create_sheet("Table 2")
    rows = _index_sheet_rows(
        "irsd",
        [
            (
                "117011326",
                "Sydney CBD",
                {
                    "urp": "np",
                    "score": "np",
                    "aus_rank": "np",
                    "aus_decile": "np",
                    "aus_percentile": "np",
                    "state": "NSW",
                    "state_rank": "np",
                    "state_decile": "np",
                    "state_percentile": "np",
                    "sa1_min": "np",
                    "sa1_max": "np",
                    "pct_urp_no_score": "np",
                },
            ),
            (
                "117011327",
                "North Sydney",
                {
                    "urp": 9500,
                    "score": 950,
                    "aus_rank": 7000,
                    "aus_decile": 5,
                    "aus_percentile": 50,
                    "state": "NSW",
                    "state_rank": 3000,
                    "state_decile": 5,
                    "state_percentile": 50,
                    "sa1_min": 920,
                    "sa1_max": 985,
                    "pct_urp_no_score": 0,
                },
            ),
        ],
    )
    for row in rows:
        table2.append(row)

    buf = io.BytesIO()
    wb.save(buf)

    responses.add(
        responses.GET,
        DEFAULT_SEIFA_2021_URL,
        body=buf.getvalue(),
        status=200,
    )
    ds = SeifaDataSource(root=seifa_data_dir)
    df = ds.load()
    # Suppressed → NaN
    assert pd.isna(df.loc["117011326", "irsd_score"])
    assert pd.isna(df.loc["117011326", "irsd_aus_decile"])
    # Real values intact
    assert df.loc["117011327", "irsd_score"] == 950


# ===========================================================================
# 2016 release tests
# ===========================================================================


def _build_synthetic_grid(
    sa2_records: list[tuple[object, str, dict]],
    prefix: str,
) -> list[list[object]]:
    """Build a raw row grid (list of lists) for one SEIFA index sheet.

    Mirrors the structure that :func:`_read_grids` returns for a real
    workbook — suitable for testing :func:`_parse_grids` directly without
    any file I/O.

    ``sa2_records`` is ``[(sa2_code, sa2_name, fields_dict), ...]``.
    ``sa2_code`` can be an integer (as calamine returns for .xls) or a
    string (as openpyxl returns for .xlsx).
    """
    preamble: list[list[object]] = [
        ["Australian Bureau of Statistics"],
        [f"Index of {prefix.upper()} -- SEIFA 2016"],
        ["Released 27 March 2018"],
        [f"Table for {prefix} index"],
        # Group row (spacer).
        [
            "",
            "",
            "",
            "",
            "",
            "Ranking within Australia",
            "",
            "",
            "",
            "Ranking within State or Territory",
            "",
            "",
            "",
            "",
            "",
            "",
        ],
        # Header row (row index 5 == _DEFAULT_HEADER_ROW).
        [
            "2016 Statistical Area Level 2 (SA2) 9-Digit Code",
            "2016 Statistical Area Level 2 (SA2) Name",
            "Usual Resident Population",
            "Score",
            "",
            "Rank",
            "Decile",
            "Percentile",
            "",
            "State",
            "Rank",
            "Decile",
            "Percentile",
            "Minimum score for SA1s in area",
            "Maximum score for SA1s in area",
            "% Usual Resident Population without a score",
        ],
    ]
    data: list[list[object]] = []
    for sa2_code, sa2_name, f in sa2_records:
        data.append(
            [
                sa2_code,
                sa2_name,
                f["urp"],
                f["score"],
                None,
                f["aus_rank"],
                f["aus_decile"],
                f["aus_percentile"],
                None,
                f["state"],
                f["state_rank"],
                f["state_decile"],
                f["state_percentile"],
                f.get("sa1_min", 0),
                f.get("sa1_max", 0),
                f.get("pct_urp_no_score", 0),
            ]
        )
    return preamble + data


def _build_synthetic_grids_2016() -> dict[str, list[list[object]]]:
    """Build a minimal synthetic grid dict for 2016-style tests.

    SA2 codes are integers (as calamine returns from an .xls integer cell).
    Two SA2s; one aggregate row (non-9-digit) that should be filtered.
    """
    # Edition-2 SA2 codes (9-digit integers).
    sa2_a_code = 101021007  # integer, as calamine returns from .xls
    sa2_b_code = 101021008

    fields_a = {
        "urp": 12000,
        "score": 1042,
        "aus_rank": 1500,
        "aus_decile": 8,
        "aus_percentile": 75,
        "state": "NSW",
        "state_rank": 800,
        "state_decile": 8,
        "state_percentile": 78,
        "sa1_min": 1000,
        "sa1_max": 1080,
        "pct_urp_no_score": 0,
    }
    fields_b = {
        "urp": 9500,
        "score": 950,
        "aus_rank": 7000,
        "aus_decile": 5,
        "aus_percentile": 50,
        "state": "NSW",
        "state_rank": 3000,
        "state_decile": 5,
        "state_percentile": 50,
        "sa1_min": 920,
        "sa1_max": 985,
        "pct_urp_no_score": 0,
    }
    aggregate = {
        "urp": 999999,
        "score": 1000,
        "aus_rank": 0,
        "aus_decile": 5,
        "aus_percentile": 50,
        "state": "AUS",
        "state_rank": 0,
        "state_decile": 5,
        "state_percentile": 50,
    }

    grids: dict[str, list[list[object]]] = {"Contents": [["Contents (skipped)"]]}
    offsets = {"irsd": 0, "irsad": 38, "ier": -32, "ieo": 58}
    for sheet_idx, (sheet_name, prefix) in enumerate(
        [("Table 2", "irsd"), ("Table 3", "irsad"), ("Table 4", "ier"), ("Table 5", "ieo")],
        start=2,
    ):
        offset = offsets[prefix]
        a = dict(fields_a, score=fields_a["score"] + offset)
        b = dict(fields_b, score=fields_b["score"] + offset)
        grids[sheet_name] = _build_synthetic_grid(
            [
                (sa2_a_code, "Area A", a),
                (sa2_b_code, "Area B", b),
                ("Australia", "Australia", aggregate),  # should be filtered
            ],
            prefix,
        )
    return grids


# ---- _parse_grids unit tests (pure; no file I/O) --------------------------


def test_parse_grids_2016_integer_codes() -> None:
    """Calamine returns integer SA2 codes from .xls; verify they are
    converted to 9-digit strings and used as the DataFrame index."""
    grids = _build_synthetic_grids_2016()
    df = _parse_grids(grids, sa2_index_name="sa2_code_2016")

    assert df.index.name == "sa2_code_2016"
    assert "101021007" in df.index
    assert "101021008" in df.index
    assert "Australia" not in df.index


def test_parse_grids_2016_all_four_indexes() -> None:
    """All four index prefixes appear in the columns."""
    grids = _build_synthetic_grids_2016()
    df = _parse_grids(grids, sa2_index_name="sa2_code_2016")
    for prefix in ("irsd", "irsad", "ier", "ieo"):
        for flavour in (
            "score",
            "aus_rank",
            "aus_decile",
            "aus_percentile",
            "state_rank",
            "state_decile",
            "state_percentile",
        ):
            col = f"{prefix}_{flavour}"
            assert col in df.columns, f"missing column {col}"


def test_parse_grids_2016_score_values() -> None:
    """Each index gets the offset applied in the fixture."""
    grids = _build_synthetic_grids_2016()
    df = _parse_grids(grids, sa2_index_name="sa2_code_2016")
    row = df.loc["101021007"]
    assert row["irsd_score"] == 1042  # offset 0
    assert row["irsad_score"] == 1080  # offset +38
    assert row["ier_score"] == 1010  # offset -32
    assert row["ieo_score"] == 1100  # offset +58


def test_parse_grids_2016_suppressed_cells() -> None:
    """None cells from calamine (empty/np) come through as NaN."""
    grids = _build_synthetic_grids_2016()
    # Patch one cell in the irsd grid to None (simulates suppression).
    # Data rows start after the 6-row preamble; first data row = index 6.
    data_row = grids["Table 2"][6]
    data_row[3] = None  # score column (index 3)
    df = _parse_grids(grids, sa2_index_name="sa2_code_2016")
    assert pd.isna(df.loc["101021007", "irsd_score"])


# ---- 2016 fetcher-level tests (URL / filename; no calamine parse) ---------


def test_seifa_2016_uses_correct_url(seifa_data_dir: Path) -> None:
    ds = SeifaDataSource(release="2016", root=seifa_data_dir)
    assert ds._resolved_url == DEFAULT_SEIFA_2016_URL


def test_seifa_2016_sa2_index_name(seifa_data_dir: Path) -> None:
    ds = SeifaDataSource(release="2016", root=seifa_data_dir)
    assert ds._sa2_index_name == "sa2_code_2016"


def test_seifa_2016_filename_has_xls_extension(seifa_data_dir: Path) -> None:
    """The cached file for the 2016 release is saved with .xls suffix."""
    ds = SeifaDataSource(release="2016", root=seifa_data_dir)
    assert ds._xlsx_path.suffix == ".xls"
    assert "seifa-2016" in ds._xlsx_path.name


def test_seifa_2021_filename_has_xlsx_extension(seifa_data_dir: Path) -> None:
    ds = SeifaDataSource(release="2021", root=seifa_data_dir)
    assert ds._xlsx_path.suffix == ".xlsx"
    assert "seifa-2021" in ds._xlsx_path.name


@responses.activate
def test_fetch_2016_downloads_to_xls_path(
    seifa_data_dir: Path,
) -> None:
    """The 2016 fetcher hits the right URL and saves to seifa-2016.xls."""
    dummy_bytes = b"PK\x03\x04"  # minimal ZIP magic (just needs non-empty bytes)
    responses.add(
        responses.GET,
        DEFAULT_SEIFA_2016_URL,
        body=dummy_bytes,
        status=200,
    )
    ds = SeifaDataSource(release="2016", root=seifa_data_dir)
    path = ds.fetch()
    assert path.exists()
    assert path.suffix == ".xls"
    assert path.name == "seifa-2016.xls"
    assert len(responses.calls) == 1
