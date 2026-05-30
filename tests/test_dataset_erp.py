"""Tests for the ERP dataset fetcher (spec §20, dataset id ``erp_by_sa2``)."""

from __future__ import annotations

import io
from pathlib import Path

import openpyxl
import pandas as pd
import pytest
import responses

from census_augment.datasets._erp import (
    ERP_AGE_SEX_LANDING_URL,
    ERP_LANDING_URL,
    ErpDataSource,
    _state_to_abbreviation,
)


def _make_landing_html(release_periods: list[str]) -> str:
    """Synthetic landing-page HTML with DS0003 links for the given periods.

    Each entry is a "YYYY-YY" or "YYYY-YYYY" string like "2024-25".
    """
    body_parts = ["<html><body>"]
    for period in release_periods:
        body_parts.append(
            f'<a href="/statistics/people/population/regional-population/'
            f'{period}/32180DS0003_2001-25.xlsx">DS0003</a>'
        )
    body_parts.append("</body></html>")
    return "\n".join(body_parts)


def _make_erp_xlsx(
    sa2_records: list[tuple[str, str, str, dict[int, int]]],
) -> bytes:
    """Build an ERP-shaped XLSX with Table 1 mirroring the real layout.

    sa2_records: list of (sa2_code, sa2_name, state_name, year→population dict)
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Contents"
    ws.append(["Contents (skipped)"])

    t1 = wb.create_sheet("Table 1")
    t1.append(["This tab has one table"])
    t1.append(["Table 1. ERP estimates"])
    t1.append([])
    t1.append([])
    t1.append([])
    # Row 5 (header) — ABS uses cols 0-7 for the geography hierarchy
    # (state code/name, GCCSA, SA4, SA3) and cols 8-9 for SA2.
    years_in_data = sorted({y for *_, year_dict in sa2_records for y in year_dict})
    header = [
        "S/T code",
        "S/T name",
        "GCCSA code",
        "GCCSA name",
        "SA4 code",
        "SA4 name",
        "SA3 code",
        "SA3 name",
        "SA2 code",
        "SA2 name",
    ] + list(years_in_data)
    t1.append(header)

    for sa2_code, sa2_name, state_name, populations in sa2_records:
        row: list[object] = [
            "1",
            state_name,
            "1RNSW",
            "Rest of NSW",
            "101",
            "Capital Region",
            "10102",
            "Queanbeyan",
            sa2_code,
            sa2_name,
        ]
        for year in years_in_data:
            row.append(populations.get(year, None))
        t1.append(row)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _make_age_sex_landing_html(release_years: list[str]) -> str:
    """Synthetic 3235.0 landing-page HTML.

    Each entry is a calendar-year string like "2024".
    """
    body_parts = ["<html><body>"]
    for year in release_years:
        body_parts.append(
            f'<a href="/statistics/people/population/regional-population-age-and-sex/'
            f'{year}/32350DS0002_{year}.xlsx">DS0002</a>'
        )
    body_parts.append("</body></html>")
    return "\n".join(body_parts)


def _make_age_sex_xlsx(
    sa2_records: list[tuple[str, str, int, int, float, float, float, float, float]],
) -> bytes:
    """Build a 3235.0 DS0002-shaped XLSX with Table 1 mirroring the real layout.

    sa2_records: list of tuples
        (sa2_code, sa2_name, males, females, sex_ratio, median_age,
         pct_0_14, pct_15_64, pct_65_plus)

    The persons total is computed as males + females (matches the real
    file's invariant). The fetcher derives age-band counts by multiplying
    persons * pct/100 — synthetic test rows respect that contract.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Contents"
    ws.append(["Contents (skipped)"])

    t1 = wb.create_sheet("Table 1")
    # Mirror the real layout: 5 title rows, then row 5 with the geography
    # labels + values, then data rows starting at row 6 (index 5).
    t1.append(["This tab has one table"])
    t1.append(["Australian Bureau of Statistics"])
    t1.append(["Table 1. Median age, sex ratio and broad age groups, SA2"])
    t1.append([])
    # Row 4 = column-group labels (we use blanks since the parser
    # ignores them — it only needs the data start row).
    t1.append(
        [
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "Males",
            "Females",
            "Persons",
            "Sex ratio",
            "Median age",
            "People aged 0-14 years",
            "People aged 15-64 years",
            "People aged 65 years and over",
        ]
    )
    # Row 5 = geography + units labels.
    t1.append(
        [
            "S/T code",
            "S/T name",
            "GCCSA code",
            "GCCSA name",
            "SA4 code",
            "SA4 name",
            "SA3 code",
            "SA3 name",
            "SA2 code",
            "SA2 name",
            "no.",
            "no.",
            "no.",
            "males per 100 females",
            "years",
            "%",
            "%",
            "%",
        ]
    )

    for (
        sa2_code,
        sa2_name,
        males,
        females,
        sex_ratio,
        median_age,
        pct_0_14,
        pct_15_64,
        pct_65_plus,
    ) in sa2_records:
        persons = males + females
        t1.append(
            [
                "1",
                "New South Wales",
                "1RNSW",
                "Rest of NSW",
                "101",
                "Capital Region",
                "10102",
                "Queanbeyan",
                sa2_code,
                sa2_name,
                males,
                females,
                persons,
                sex_ratio,
                median_age,
                pct_0_14,
                pct_15_64,
                pct_65_plus,
            ]
        )

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@pytest.fixture
def erp_data_dir(tmp_path: Path) -> Path:
    return tmp_path / "erp-cache"


def _add_age_sex_mocks(
    sa2_records: list[tuple[str, str, int, int, float, float, float, float, float]],
    release_year: str = "2024",
) -> None:
    """Register the two HTTP mocks required for the age/sex enrichment leg.

    Pass `responses.activate` test the same SA2 records you want
    `ds.load()` to merge in; this saves repeating the URL boilerplate in
    every test.
    """
    responses.add(
        responses.GET,
        ERP_AGE_SEX_LANDING_URL,
        body=_make_age_sex_landing_html([release_year]),
        status=200,
    )
    responses.add(
        responses.GET,
        f"https://www.abs.gov.au/statistics/people/population/"
        f"regional-population-age-and-sex/{release_year}/32350DS0002_{release_year}.xlsx",
        body=_make_age_sex_xlsx(sa2_records),
        status=200,
    )


# ---- helpers -------------------------------------------------------------


def test_state_to_abbreviation_known() -> None:
    assert _state_to_abbreviation("New South Wales") == "NSW"
    assert _state_to_abbreviation("victoria") == "VIC"
    assert _state_to_abbreviation("Australian Capital Territory") == "ACT"


def test_state_to_abbreviation_unknown_falls_back() -> None:
    # Unknown name: take the first 3 chars uppercase.
    assert _state_to_abbreviation("Vibing Wonderland") == "VIB"


# ---- release resolution --------------------------------------------------


@responses.activate
def test_resolve_latest_picks_highest_period(erp_data_dir: Path) -> None:
    responses.add(
        responses.GET,
        ERP_LANDING_URL,
        body=_make_landing_html(["2023-24", "2024-25"]),
        status=200,
    )
    ds = ErpDataSource(release="latest", root=erp_data_dir)
    assert ds.resolved_release == "2024"


@responses.activate
def test_resolve_specific_release(erp_data_dir: Path) -> None:
    responses.add(
        responses.GET,
        ERP_LANDING_URL,
        body=_make_landing_html(["2023-24", "2024-25"]),
        status=200,
    )
    ds = ErpDataSource(release="2023", root=erp_data_dir)
    assert ds.resolved_release == "2023"


@responses.activate
def test_resolve_release_more_recent_than_latest_raises(erp_data_dir: Path) -> None:
    """Requesting a release year newer than the latest published workbook
    raises a clear RuntimeError naming both the requested year and the
    available range. Historical years (≤ latest) are now accepted via
    the load() projection (issue #92 fix); only future years still
    raise here.
    """
    responses.add(
        responses.GET,
        ERP_LANDING_URL,
        body=_make_landing_html(["2024-25"]),
        status=200,
    )
    ds = ErpDataSource(release="2030", root=erp_data_dir)
    with pytest.raises(RuntimeError, match="more recent than the latest"):
        _ = ds.resolved_release


@responses.activate
def test_resolve_no_links_raises(erp_data_dir: Path) -> None:
    responses.add(
        responses.GET,
        ERP_LANDING_URL,
        body="<html><body>Nothing to see</body></html>",
        status=200,
    )
    ds = ErpDataSource(release="latest", root=erp_data_dir)
    with pytest.raises(RuntimeError, match="Could not find"):
        _ = ds.resolved_release


# ---- fetch + parse -------------------------------------------------------


@responses.activate
def test_load_returns_sa2_indexed_dataframe(erp_data_dir: Path) -> None:
    fake_xlsx = _make_erp_xlsx(
        [
            ("117011326", "Sydney CBD", "New South Wales", {2001: 5000, 2010: 7500, 2024: 12000}),
            ("117011327", "North Sydney", "New South Wales", {2001: 4500, 2010: 6800, 2024: 9500}),
            # Aggregate row that should be skipped.
            ("Australia", "Australia", "Australia", {2001: 19000000, 2024: 27000000}),
        ]
    )
    responses.add(
        responses.GET,
        ERP_LANDING_URL,
        body=_make_landing_html(["2024-25"]),
        status=200,
    )
    download_url = (
        "https://www.abs.gov.au/statistics/people/population/"
        "regional-population/2024-25/32180DS0003_2001-25.xlsx"
    )
    responses.add(responses.GET, download_url, body=fake_xlsx, status=200)
    _add_age_sex_mocks(
        [
            # (sa2, name, males, females, sex_ratio, median_age,
            #  pct_0_14, pct_15_64, pct_65_plus)
            ("117011326", "Sydney CBD", 6000, 6000, 100.0, 35.0, 15.0, 70.0, 15.0),
            ("117011327", "North Sydney", 4800, 4700, 102.0, 14.0, 14.0, 72.0, 14.0),
        ]
    )

    ds = ErpDataSource(root=erp_data_dir)
    df = ds.load()

    assert df.index.name == "sa2_code_2021"
    assert "117011326" in df.index
    assert "117011327" in df.index
    assert "Australia" not in df.index
    assert "population_total" in df.columns
    assert "reference_year" in df.columns
    assert df.loc["117011326", "population_total"] == 12000
    assert df.loc["117011326", "reference_year"] == 2024
    assert df.loc["117011326", "population_history_2001"] == 5000
    assert df.loc["117011326", "state_abbreviation"] == "NSW"


@responses.activate
def test_load_handles_suppressed_cells(erp_data_dir: Path) -> None:
    fake_xlsx = _make_erp_xlsx(
        [
            ("117011326", "Sydney CBD", "New South Wales", {2001: "np", 2024: 12000}),
        ]
    )
    responses.add(
        responses.GET,
        ERP_LANDING_URL,
        body=_make_landing_html(["2024-25"]),
        status=200,
    )
    download_url = (
        "https://www.abs.gov.au/statistics/people/population/"
        "regional-population/2024-25/32180DS0003_2001-25.xlsx"
    )
    responses.add(responses.GET, download_url, body=fake_xlsx, status=200)
    _add_age_sex_mocks([("117011326", "Sydney CBD", 6000, 6000, 100.0, 35.0, 15.0, 70.0, 15.0)])

    ds = ErpDataSource(root=erp_data_dir)
    df = ds.load()
    assert pd.isna(df.loc["117011326", "population_history_2001"])
    assert df.loc["117011326", "population_history_2024"] == 12000


@responses.activate
def test_load_caches_parquet(erp_data_dir: Path) -> None:
    fake_xlsx = _make_erp_xlsx(
        [
            ("117011326", "Sydney CBD", "New South Wales", {2024: 12000}),
        ]
    )
    responses.add(
        responses.GET,
        ERP_LANDING_URL,
        body=_make_landing_html(["2024-25"]),
        status=200,
    )
    download_url = (
        "https://www.abs.gov.au/statistics/people/population/"
        "regional-population/2024-25/32180DS0003_2001-25.xlsx"
    )
    responses.add(responses.GET, download_url, body=fake_xlsx, status=200)
    _add_age_sex_mocks([("117011326", "Sydney CBD", 6000, 6000, 100.0, 35.0, 15.0, 70.0, 15.0)])

    ds = ErpDataSource(root=erp_data_dir)
    df1 = ds.load()
    df2 = ds.load()
    pd.testing.assert_frame_equal(df1, df2)


# ---- age/sex enrichment (wishlist columns) -------------------------------


@responses.activate
def test_load_includes_age_sex_columns_when_available(erp_data_dir: Path) -> None:
    """ERP load merges the 3235.0 DS0002 age/sex columns onto the totals.

    Verifies the derived band counts (persons × pct/100) round correctly
    and that gender / median age land verbatim.
    """
    fake_xlsx = _make_erp_xlsx(
        [
            ("117011326", "Sydney CBD", "New South Wales", {2024: 12000}),
            ("117011327", "North Sydney", "New South Wales", {2024: 9500}),
        ]
    )
    responses.add(
        responses.GET,
        ERP_LANDING_URL,
        body=_make_landing_html(["2024-25"]),
        status=200,
    )
    responses.add(
        responses.GET,
        "https://www.abs.gov.au/statistics/people/population/"
        "regional-population/2024-25/32180DS0003_2001-25.xlsx",
        body=fake_xlsx,
        status=200,
    )
    _add_age_sex_mocks(
        [
            # Sydney CBD: 6000 M + 6000 F = 12000 persons (matches totals).
            # 15% × 12000 = 1800 aged 0-14; 70% × 12000 = 8400 aged 15-64;
            # 15% × 12000 = 1800 aged 65+.
            ("117011326", "Sydney CBD", 6000, 6000, 100.0, 35.0, 15.0, 70.0, 15.0),
            # North Sydney: 4800 M + 4700 F = 9500 persons.
            # 10% × 9500 = 950 aged 0-14; 75% × 9500 = 7125 aged 15-64;
            # 15% × 9500 = 1425 aged 65+.
            ("117011327", "North Sydney", 4800, 4700, 102.0, 38.5, 10.0, 75.0, 15.0),
        ]
    )

    ds = ErpDataSource(root=erp_data_dir)
    df = ds.load()

    # All six wishlist columns present
    for col in (
        "population_male",
        "population_female",
        "population_0_14",
        "population_15_64",
        "population_65_plus",
        "median_age",
    ):
        assert col in df.columns, f"missing wishlist column: {col}"

    # Sydney CBD: gender + median age verbatim
    assert df.loc["117011326", "population_male"] == 6000
    assert df.loc["117011326", "population_female"] == 6000
    assert df.loc["117011326", "median_age"] == 35.0

    # Sydney CBD: derived age-band counts (persons × pct/100, rounded)
    assert df.loc["117011326", "population_0_14"] == 1800
    assert df.loc["117011326", "population_15_64"] == 8400
    assert df.loc["117011326", "population_65_plus"] == 1800

    # North Sydney: rounded derivations
    assert df.loc["117011327", "population_0_14"] == 950
    assert df.loc["117011327", "population_15_64"] == 7125
    assert df.loc["117011327", "population_65_plus"] == 1425


@responses.activate
def test_load_silently_omits_age_sex_columns_when_landing_fails(
    erp_data_dir: Path,
) -> None:
    """If the age/sex landing page can't be reached, fall back to DS0003-only.

    The augmentor's existing pre-3235.0 contract (population_total +
    population_history_*) is preserved. The age/sex columns are simply
    absent — downstream consumers using only the core columns continue
    working. A WARNING goes to the log.
    """
    fake_xlsx = _make_erp_xlsx([("117011326", "Sydney CBD", "New South Wales", {2024: 12000})])
    responses.add(
        responses.GET,
        ERP_LANDING_URL,
        body=_make_landing_html(["2024-25"]),
        status=200,
    )
    responses.add(
        responses.GET,
        "https://www.abs.gov.au/statistics/people/population/"
        "regional-population/2024-25/32180DS0003_2001-25.xlsx",
        body=fake_xlsx,
        status=200,
    )
    # Age/sex landing page returns 503 — simulates transient ABS outage.
    responses.add(
        responses.GET,
        ERP_AGE_SEX_LANDING_URL,
        body="<html>broken</html>",
        status=503,
    )

    ds = ErpDataSource(root=erp_data_dir)
    df = ds.load()

    # Core columns still present
    assert df.loc["117011326", "population_total"] == 12000
    # Age/sex columns absent
    for col in (
        "population_male",
        "population_female",
        "median_age",
        "population_0_14",
        "population_15_64",
        "population_65_plus",
    ):
        assert col not in df.columns, f"unexpected age/sex column present: {col}"


# ---- issue #92: historical-year projection (release="<historical-year>") ----


@responses.activate
def test_load_with_historical_release_projects_population_total(
    erp_data_dir: Path,
) -> None:
    """Requesting a historical release year (e.g. 2017) returns the same
    underlying workbook but with ``population_total`` projected from
    ``population_history_2017`` and ``reference_year`` set to 2017.

    Issue #92 fix: ERP publishes ONE annual workbook per cycle that
    carries the full 2001-onwards history in
    ``population_history_<year>`` columns. Temporal-mode resolution
    asks for a historical release per row date; the fetcher now
    serves those via column projection rather than failing.
    """
    # Realistic synthetic data: 2017 and 2024 history columns with
    # distinct values so we can verify the projection picks the right
    # column.
    fake_xlsx = _make_erp_xlsx(
        [
            (
                "117011326",
                "Sydney CBD",
                "New South Wales",
                {2017: 11000, 2024: 12500},
            ),
        ]
    )
    responses.add(
        responses.GET,
        ERP_LANDING_URL,
        body=_make_landing_html(["2024-25"]),
        status=200,
    )
    responses.add(
        responses.GET,
        "https://www.abs.gov.au/statistics/people/population/"
        "regional-population/2024-25/32180DS0003_2001-25.xlsx",
        body=fake_xlsx,
        status=200,
    )
    _add_age_sex_mocks([("117011326", "Sydney CBD", 6250, 6250, 100.0, 35.0, 15.0, 70.0, 15.0)])

    ds = ErpDataSource(release="2017", root=erp_data_dir)
    df = ds.load()

    # Logical release the user asked for.
    assert ds.resolved_release == "2017"
    # Projected: population_total comes from population_history_2017
    # and reference_year is 2017 (not the workbook's latest 2024).
    assert df.loc["117011326", "population_total"] == 11000
    assert df.loc["117011326", "reference_year"] == 2017
    # The history columns themselves stay intact for downstream use.
    assert df.loc["117011326", "population_history_2017"] == 11000
    assert df.loc["117011326", "population_history_2024"] == 12500


@responses.activate
def test_load_with_historical_release_nulls_age_sex_columns(
    erp_data_dir: Path,
) -> None:
    """Age/sex columns reflect the latest 3235.0 publication only —
    there's no historical age/sex breakdown in the products the
    augmentor fetches. For historical releases the load() projection
    nulls them out so users don't accidentally pair 2017 totals with
    2024 demographics.
    """
    fake_xlsx = _make_erp_xlsx(
        [("117011326", "Sydney CBD", "New South Wales", {2017: 11000, 2024: 12500})]
    )
    responses.add(
        responses.GET,
        ERP_LANDING_URL,
        body=_make_landing_html(["2024-25"]),
        status=200,
    )
    responses.add(
        responses.GET,
        "https://www.abs.gov.au/statistics/people/population/"
        "regional-population/2024-25/32180DS0003_2001-25.xlsx",
        body=fake_xlsx,
        status=200,
    )
    _add_age_sex_mocks([("117011326", "Sydney CBD", 6250, 6250, 100.0, 35.0, 15.0, 70.0, 15.0)])

    ds = ErpDataSource(release="2017", root=erp_data_dir)
    df = ds.load()

    # All age/sex columns null for the historical release.
    for col in (
        "population_male",
        "population_female",
        "median_age",
        "population_0_14",
        "population_15_64",
        "population_65_plus",
    ):
        assert pd.isna(df.loc["117011326", col]), (
            f"expected null {col!r} for historical release; got {df.loc['117011326', col]!r}"
        )


@responses.activate
def test_load_with_historical_release_outside_coverage_raises(
    erp_data_dir: Path,
) -> None:
    """A release year that's ≤ latest but not in the workbook's
    ``population_history_*`` columns raises a clear error naming the
    available years. Catches "user asked for 1990 but ABS series
    starts at 2001" without hardcoding the lower bound.
    """
    # 2024 workbook contains 2017 + 2024 (we control the synthetic
    # fixture). Requesting 2015 should raise — not in history.
    fake_xlsx = _make_erp_xlsx(
        [("117011326", "Sydney CBD", "New South Wales", {2017: 11000, 2024: 12500})]
    )
    responses.add(
        responses.GET,
        ERP_LANDING_URL,
        body=_make_landing_html(["2024-25"]),
        status=200,
    )
    responses.add(
        responses.GET,
        "https://www.abs.gov.au/statistics/people/population/"
        "regional-population/2024-25/32180DS0003_2001-25.xlsx",
        body=fake_xlsx,
        status=200,
    )
    _add_age_sex_mocks([("117011326", "Sydney CBD", 6250, 6250, 100.0, 35.0, 15.0, 70.0, 15.0)])

    ds = ErpDataSource(release="2015", root=erp_data_dir)
    with pytest.raises(RuntimeError, match="not in the workbook's historical coverage"):
        ds.load()


@responses.activate
def test_load_latest_release_unchanged_by_projection(erp_data_dir: Path) -> None:
    """Regression-prevention companion: ``release="latest"`` (the
    default) returns the workbook's data verbatim — no projection
    happens when the resolved release equals the physical release
    year.
    """
    fake_xlsx = _make_erp_xlsx(
        [("117011326", "Sydney CBD", "New South Wales", {2017: 11000, 2024: 12500})]
    )
    responses.add(
        responses.GET,
        ERP_LANDING_URL,
        body=_make_landing_html(["2024-25"]),
        status=200,
    )
    responses.add(
        responses.GET,
        "https://www.abs.gov.au/statistics/people/population/"
        "regional-population/2024-25/32180DS0003_2001-25.xlsx",
        body=fake_xlsx,
        status=200,
    )
    _add_age_sex_mocks([("117011326", "Sydney CBD", 6250, 6250, 100.0, 35.0, 15.0, 70.0, 15.0)])

    ds = ErpDataSource(release="latest", root=erp_data_dir)
    df = ds.load()

    # Latest = 2024. population_total reflects 2024.
    assert ds.resolved_release == "2024"
    assert df.loc["117011326", "population_total"] == 12500
    assert df.loc["117011326", "reference_year"] == 2024
    # Age/sex columns present (not nulled).
    assert df.loc["117011326", "population_male"] == 6250
    assert df.loc["117011326", "median_age"] == 35.0


# ---- population_density_per_km2 column (attach_sa2_areas) -------------------


@responses.activate
def test_load_emits_population_density_when_areas_attached(
    erp_data_dir: Path,
) -> None:
    """``attach_sa2_areas`` enables the ``population_density_per_km2``
    column. Density = ``population_total / area_km2``. Test uses
    convenient round numbers so the assertion is unambiguous.
    """
    fake_xlsx = _make_erp_xlsx(
        [
            ("117011326", "Sydney CBD", "New South Wales", {2024: 10_000}),
            ("117011327", "North Sydney", "New South Wales", {2024: 50_000}),
        ]
    )
    responses.add(
        responses.GET,
        ERP_LANDING_URL,
        body=_make_landing_html(["2024-25"]),
        status=200,
    )
    responses.add(
        responses.GET,
        "https://www.abs.gov.au/statistics/people/population/"
        "regional-population/2024-25/32180DS0003_2001-25.xlsx",
        body=fake_xlsx,
        status=200,
    )
    _add_age_sex_mocks(
        [
            ("117011326", "Sydney CBD", 5000, 5000, 100.0, 35.0, 15.0, 70.0, 15.0),
            ("117011327", "North Sydney", 25000, 25000, 100.0, 35.0, 15.0, 70.0, 15.0),
        ]
    )

    ds = ErpDataSource(release="latest", root=erp_data_dir)
    ds.attach_sa2_areas(
        {
            "117011326": 5.0,  # 10,000 / 5.0 km² = 2,000/km² (dense)
            "117011327": 100.0,  # 50,000 / 100 km² = 500/km² (sparse)
        }
    )
    df = ds.load()

    assert "population_density_per_km2" in df.columns
    assert df.loc["117011326", "population_density_per_km2"] == 2000.0
    assert df.loc["117011327", "population_density_per_km2"] == 500.0


@responses.activate
def test_load_omits_population_density_without_attach(
    erp_data_dir: Path,
) -> None:
    """Without ``attach_sa2_areas``, the density column is absent —
    no implicit dependency on the boundary file, keeps the fetcher
    standalone-usable.
    """
    fake_xlsx = _make_erp_xlsx([("117011326", "Sydney CBD", "New South Wales", {2024: 10_000})])
    responses.add(
        responses.GET,
        ERP_LANDING_URL,
        body=_make_landing_html(["2024-25"]),
        status=200,
    )
    responses.add(
        responses.GET,
        "https://www.abs.gov.au/statistics/people/population/"
        "regional-population/2024-25/32180DS0003_2001-25.xlsx",
        body=fake_xlsx,
        status=200,
    )
    _add_age_sex_mocks([("117011326", "Sydney CBD", 5000, 5000, 100.0, 35.0, 15.0, 70.0, 15.0)])

    ds = ErpDataSource(release="latest", root=erp_data_dir)
    # Deliberately no attach_sa2_areas call.
    df = ds.load()

    assert "population_density_per_km2" not in df.columns


@responses.activate
def test_load_emits_density_for_historical_release(
    erp_data_dir: Path,
) -> None:
    """The density projection lines up with the historical-release
    projection from #92: density = projected_population_total /
    area_km2. So density for ``release="2017"`` uses
    ``population_history_2017`` as the numerator.
    """
    fake_xlsx = _make_erp_xlsx(
        [("117011326", "Sydney CBD", "New South Wales", {2017: 8_000, 2024: 10_000})]
    )
    responses.add(
        responses.GET,
        ERP_LANDING_URL,
        body=_make_landing_html(["2024-25"]),
        status=200,
    )
    responses.add(
        responses.GET,
        "https://www.abs.gov.au/statistics/people/population/"
        "regional-population/2024-25/32180DS0003_2001-25.xlsx",
        body=fake_xlsx,
        status=200,
    )
    _add_age_sex_mocks([("117011326", "Sydney CBD", 5000, 5000, 100.0, 35.0, 15.0, 70.0, 15.0)])

    ds = ErpDataSource(release="2017", root=erp_data_dir)
    ds.attach_sa2_areas({"117011326": 4.0})
    df = ds.load()

    # Projected total: 8000 (from population_history_2017).
    assert df.loc["117011326", "population_total"] == 8000
    assert df.loc["117011326", "reference_year"] == 2017
    # Density: 8000 / 4.0 km² = 2000/km².
    assert df.loc["117011326", "population_density_per_km2"] == 2000.0


@responses.activate
def test_load_density_nan_for_sa2_missing_from_areas(
    erp_data_dir: Path,
) -> None:
    """When the SA2 areas lookup doesn't cover every SA2 in the ERP
    data (e.g. partial lookup), the density column is NaN for the
    uncovered SA2s — never crashes.
    """
    fake_xlsx = _make_erp_xlsx(
        [
            ("117011326", "Sydney CBD", "New South Wales", {2024: 10_000}),
            ("117011327", "North Sydney", "New South Wales", {2024: 50_000}),
        ]
    )
    responses.add(
        responses.GET,
        ERP_LANDING_URL,
        body=_make_landing_html(["2024-25"]),
        status=200,
    )
    responses.add(
        responses.GET,
        "https://www.abs.gov.au/statistics/people/population/"
        "regional-population/2024-25/32180DS0003_2001-25.xlsx",
        body=fake_xlsx,
        status=200,
    )
    _add_age_sex_mocks(
        [
            ("117011326", "Sydney CBD", 5000, 5000, 100.0, 35.0, 15.0, 70.0, 15.0),
            ("117011327", "North Sydney", 25000, 25000, 100.0, 35.0, 15.0, 70.0, 15.0),
        ]
    )

    ds = ErpDataSource(release="latest", root=erp_data_dir)
    # Only one SA2 in the areas lookup — the other should produce NaN.
    ds.attach_sa2_areas({"117011326": 5.0})
    df = ds.load()

    assert df.loc["117011326", "population_density_per_km2"] == 2000.0
    assert pd.isna(df.loc["117011327", "population_density_per_km2"])
